"""Private JSON/XLSX/PDF exports with bounded revision snippets."""

from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import datetime
from typing import cast
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction, Permission, has_permission
from app.core.config import Settings
from app.core.exceptions import AuthorizationError
from app.models.revision_change import RevisionChange
from app.models.revision_comparison import RevisionComparison
from app.models.user import User
from app.repositories.revision_change_repository import (
    RevisionChangeRepository,
)
from app.repositories.revision_comparison_repository import (
    RevisionComparisonRepository,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.documents.base import DocumentServiceBase, document_error
from app.services.revision_comparison.revision_comparison_service import (
    revision_comparison_not_found,
)

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@dataclass(frozen=True, slots=True)
class RevisionExportArtifact:
    content: bytes
    media_type: str
    filename: str


class RevisionExportService(DocumentServiceBase):
    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        user: User,
        metadata: RequestMetadata,
    ) -> None:
        super().__init__(session, user, metadata)
        self.settings = settings
        self.comparisons = RevisionComparisonRepository(session)
        self.changes = RevisionChangeRepository(session)

    async def export(
        self, comparison_id: UUID, *, export_format: str
    ) -> RevisionExportArtifact:
        self._ensure_export_permission()
        normalized = export_format.strip().lower()
        if normalized not in {"json", "xlsx", "pdf"}:
            raise document_error(
                "Revision export format must be json, xlsx, or pdf.",
                field="format",
                code="REVISION_COMPARISON_EXPORT_FORMAT_INVALID",
            )
        comparison = await self.comparisons.get_by_id(
            comparison_id,
            department_ids=self._scope_department_ids(),
        )
        if comparison is None:
            raise revision_comparison_not_found()
        rows = await self.changes.list_all(
            comparison_id,
            maximum=self.settings.revision_comparison_max_changes + 1,
        )
        if len(rows) > self.settings.revision_comparison_max_changes:
            raise document_error(
                "The revision comparison exceeds the configured export limit.",
                code="REVISION_COMPARISON_EXPORT_LIMIT",
                status_code=413,
                title="Revision comparison export is too large.",
            )
        if normalized == "json":
            content = self._json(comparison, rows)
            media_type = "application/json"
        elif normalized == "xlsx":
            content = self._xlsx(comparison, rows)
            media_type = XLSX_CONTENT_TYPE
        else:
            content = self._pdf(comparison, rows)
            media_type = "application/pdf"
        await self.audit(
            action=AuditAction.EXPORT_REVISION_COMPARISON,
            entity_type="RevisionComparison",
            entity_id=comparison.id,
            description="Revision comparison exported.",
            new_values={"format": normalized, "changeCount": len(rows)},
        )
        await self.session.commit()
        return RevisionExportArtifact(
            content=content,
            media_type=media_type,
            filename=f"revision-comparison-{comparison.id}.{normalized}",
        )

    def _json(
        self,
        comparison: RevisionComparison,
        changes: list[RevisionChange],
    ) -> bytes:
        payload = {
            "metadata": {
                "comparisonId": str(comparison.id),
                "documentId": str(comparison.document_id),
                "baseRevisionId": str(comparison.base_revision_id),
                "targetRevisionId": str(comparison.target_revision_id),
                "generatedAt": datetime.now().astimezone().isoformat(),
                "disclaimer": self._disclaimer(),
            },
            "summary": self._summary(comparison),
            "languageChanges": comparison.language_coverage_change_json,
            "findingChanges": comparison.summary_json.get(
                "findingChanges", []
            ),
            "changes": [self._change_payload(item) for item in changes],
        }
        return json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
            default=str,
        ).encode("utf-8")

    def _xlsx(
        self,
        comparison: RevisionComparison,
        changes: list[RevisionChange],
    ) -> bytes:
        workbook = Workbook()
        default = workbook.active
        assert default is not None
        workbook.remove(default)
        self._sheet_from_mapping(
            workbook, "Summary", [self._summary(comparison)]
        )
        self._sheet_from_mapping(
            workbook,
            "Section Changes",
            self._section_rows(changes),
        )
        language_values = comparison.language_coverage_change_json.get(
            "languages", []
        )
        self._sheet_from_mapping(
            workbook,
            "Language Changes",
            [
                item
                for item in language_values
                if isinstance(item, dict)
            ],
        )
        self._sheet_from_mapping(
            workbook,
            "Content Changes",
            [self._change_payload(item) for item in changes],
        )
        self._sheet_from_mapping(
            workbook,
            "Compliance Changes",
            [
                {
                    "baseStatus": comparison.summary_json.get(
                        "baseComplianceStatus"
                    ),
                    "targetStatus": comparison.summary_json.get(
                        "targetComplianceStatus"
                    ),
                    "scoreChange": self._decimal(
                        comparison.compliance_score_change
                    ),
                }
            ],
        )
        self._sheet_from_mapping(
            workbook,
            "Similarity Changes",
            [
                {
                    "scoreChange": self._decimal(
                        comparison.similarity_score_change
                    )
                }
            ],
        )
        self._sheet_from_mapping(
            workbook,
            "Glossary Changes",
            [
                {
                    "baseGlossaryRunId": comparison.base_glossary_run_id,
                    "targetGlossaryRunId": comparison.target_glossary_run_id,
                    "status": "NOT_EVALUATED"
                    if (
                        comparison.base_glossary_run_id is None
                        or comparison.target_glossary_run_id is None
                    )
                    else "AVAILABLE",
                }
            ],
        )
        finding_values = comparison.summary_json.get("findingChanges", [])
        self._sheet_from_mapping(
            workbook,
            "Finding Changes",
            [
                item
                for item in finding_values
                if isinstance(item, dict)
            ],
        )
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _pdf(
        self,
        comparison: RevisionComparison,
        changes: list[RevisionChange],
    ) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        output = io.BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title="Revision Comparison",
        )
        styles = getSampleStyleSheet()
        story: list[object] = [
            Paragraph("Revision Comparison", styles["Title"]),
            Spacer(1, 8),
            Paragraph(
                f"Document: {comparison.document_id}", styles["Normal"]
            ),
            Paragraph(
                f"Base revision: {comparison.base_revision_id}",
                styles["Normal"],
            ),
            Paragraph(
                f"Target revision: {comparison.target_revision_id}",
                styles["Normal"],
            ),
            Spacer(1, 10),
            Paragraph("Executive summary", styles["Heading2"]),
        ]
        summary_rows = [
            ["Classification", comparison.classification.value],
            ["Total changes", comparison.total_changes],
            ["Added", comparison.added_blocks],
            ["Removed", comparison.removed_blocks],
            ["Modified", comparison.modified_blocks],
            ["Moved", comparison.moved_blocks],
            ["New findings", comparison.new_findings],
            ["No longer reproduced", comparison.removed_findings],
        ]
        summary_table = Table(summary_rows, colWidths=[70 * mm, 85 * mm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.extend([summary_table, PageBreak()])
        story.append(Paragraph("Bounded content changes", styles["Heading2"]))
        maximum_pdf_rows = min(
            self.settings.report_pdf_max_table_rows, len(changes)
        )
        for item in changes[:maximum_pdf_rows]:
            story.append(
                Paragraph(
                    (
                        f"<b>{item.change_type.value}</b> "
                        f"{self._safe_pdf_text(item.source_reference_target or item.source_reference_base or '')}"
                    ),
                    styles["Normal"],
                )
            )
            base = self._snippet(item.base_text_snapshot)
            target = self._snippet(item.target_text_snapshot)
            if base:
                story.append(
                    Paragraph(
                        f"Base: {self._safe_pdf_text(base)}",
                        styles["BodyText"],
                    )
                )
            if target:
                story.append(
                    Paragraph(
                        f"Target: {self._safe_pdf_text(target)}",
                        styles["BodyText"],
                    )
                )
            story.append(Spacer(1, 5))
        story.extend(
            [
                Spacer(1, 10),
                Paragraph("Disclaimer", styles["Heading2"]),
                Paragraph(self._disclaimer(), styles["BodyText"]),
            ]
        )
        document.build(story)
        return output.getvalue()

    def _change_payload(self, item: RevisionChange) -> dict[str, object]:
        return {
            "changeType": item.change_type.value,
            "entityType": item.entity_type.value,
            "languageCode": item.language_code,
            "sourceReferenceBase": item.source_reference_base,
            "sourceReferenceTarget": item.source_reference_target,
            "baseSnippet": self._snippet(item.base_text_snapshot),
            "targetSnippet": self._snippet(item.target_text_snapshot),
            "textSimilarity": self._decimal(item.text_similarity),
            "structuralSimilarity": self._decimal(
                item.structural_similarity
            ),
            "alignmentConfidence": self._decimal(
                item.alignment_confidence
            ),
            "characterChangeCount": item.character_change_count,
            "wordChangeCount": item.word_change_count,
        }

    @staticmethod
    def _summary(item: RevisionComparison) -> dict[str, object]:
        return {
            "comparisonId": str(item.id),
            "classification": item.classification.value,
            "totalChanges": item.total_changes,
            "added": item.added_blocks,
            "removed": item.removed_blocks,
            "modified": item.modified_blocks,
            "moved": item.moved_blocks,
            "unchanged": item.unchanged_blocks,
            "newFindings": item.new_findings,
            "noLongerReproduced": item.removed_findings,
        }

    @staticmethod
    def _section_rows(
        changes: list[RevisionChange],
    ) -> list[dict[str, object]]:
        output: dict[str, dict[str, object]] = {}
        for item in changes:
            key = str(
                item.metadata_json.get("targetSectionCode")
                or item.metadata_json.get("baseSectionCode")
                or "UNMAPPED"
            )
            row = output.setdefault(
                key,
                {
                    "section": key,
                    "added": 0,
                    "removed": 0,
                    "modified": 0,
                    "moved": 0,
                    "unchanged": 0,
                },
            )
            bucket = {
                "ADDED": "added",
                "REMOVED": "removed",
                "MODIFIED": "modified",
                "MOVED": "moved",
                "UNCHANGED": "unchanged",
                "SPLIT": "modified",
                "MERGED": "modified",
            }[item.change_type.value]
            row[bucket] = cast(int, row[bucket]) + 1
        return list(output.values())

    @staticmethod
    def _sheet_from_mapping(
        workbook: Workbook,
        name: str,
        rows: list[dict[str, object]],
    ) -> None:
        sheet = workbook.create_sheet(name[:31])
        if not rows:
            sheet.append(["No data"])
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append(
                [_safe_spreadsheet_value(row.get(header)) for header in headers]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, header in enumerate(headers, start=1):
            maximum = max(
                len(str(header)),
                *(
                    len(str(sheet.cell(row=row, column=index).value or ""))
                    for row in range(2, sheet.max_row + 1)
                ),
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(
                maximum + 2, 60
            )

    def _snippet(self, value: str | None) -> str | None:
        if value is None:
            return None
        return value[: self.settings.report_text_snippet_max_characters]

    @staticmethod
    def _decimal(value: float | None) -> float | None:
        return float(value) if value is not None else None

    @staticmethod
    def _safe_pdf_text(value: str) -> str:
        return (
            value.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    @staticmethod
    def _disclaimer() -> str:
        return (
            "This automated comparison is a review aid. It does not prove "
            "legal or technical equivalence and does not modify either "
            "source document."
        )

    def _scope_department_ids(self) -> list[UUID] | None:
        if has_permission(
            self.user.role,
            Permission.REVISION_COMPARISON_VIEW_ALL_DEPARTMENTS,
            is_superuser=self.user.is_superuser,
        ):
            return None
        return [self.user.department_id] if self.user.department_id else []

    def _ensure_export_permission(self) -> None:
        if not has_permission(
            self.user.role,
            Permission.REVISION_COMPARISON_EXPORT,
            is_superuser=self.user.is_superuser,
        ):
            raise AuthorizationError(
                "You do not have permission to export this comparison."
            )


def _safe_spreadsheet_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
