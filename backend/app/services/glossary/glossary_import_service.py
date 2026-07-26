"""Safe five-sheet XLSX glossary preview and confirmation."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, cast
from uuid import UUID
from zipfile import is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.authorization import AuditAction
from app.models.glossary_enums import (
    GlossaryExceptionScopeType,
    GlossaryExceptionType,
    GlossaryLanguageCode,
    GlossaryScopeType,
    GlossaryTermSeverity,
    GlossaryTermType,
    GlossaryVariantType,
)
from app.models.glossary_exception import GlossaryException
from app.models.glossary_profile import GlossaryProfile
from app.models.glossary_term import GlossaryTerm
from app.models.glossary_term_variant import GlossaryTermVariant
from app.models.glossary_translation import GlossaryTranslation
from app.models.user import User
from app.schemas.glossary import (
    GlossaryExceptionCreate,
    GlossaryImportConfirmResponse,
    GlossaryImportIssue,
    GlossaryImportMode,
    GlossaryImportPreviewResponse,
    GlossaryImportSheetSummary,
    GlossaryProfileCreate,
    GlossaryTermCreate,
    GlossaryTranslationCreate,
    GlossaryVariantCreate,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.glossary.base import GlossaryServiceBase, glossary_error
from app.services.glossary.glossary_matching_service import (
    GlossaryMatchingService,
)
from app.services.glossary.matching.term_normalizer import normalize_term

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

SHEET_HEADERS: dict[str, tuple[str, ...]] = {
    "Profiles": (
        "code",
        "name",
        "description",
        "scope_type",
        "department_id",
        "document_type_id",
        "is_default",
        "is_active",
    ),
    "Terms": (
        "profile_code",
        "term_code",
        "concept_name",
        "description",
        "term_type",
        "severity",
        "is_case_sensitive",
        "match_whole_word",
        "allow_inflection",
        "is_regex",
        "is_active",
        "notes",
    ),
    "Translations": (
        "term_code",
        "language_code",
        "term_text",
        "is_preferred",
        "is_forbidden",
        "is_required",
        "priority",
        "usage_note",
        "example_text",
        "is_active",
    ),
    "Variants": (
        "term_code",
        "language_code",
        "preferred_term_text",
        "variant_text",
        "variant_type",
        "is_allowed",
        "is_active",
    ),
    "Exceptions": (
        "term_code",
        "scope_type",
        "department_id",
        "document_id",
        "document_revision_id",
        "document_file_id",
        "section_definition_id",
        "language_code",
        "exception_type",
        "reason",
        "effective_from",
        "effective_to",
        "is_active",
        "approved_by",
    ),
}


@dataclass(slots=True)
class ValidatedImportRow:
    sheet: str
    row_number: int
    key: str
    payload: object
    context: dict[str, object]


@dataclass(slots=True)
class GlossaryWorkbookValidation:
    rows: dict[str, list[ValidatedImportRow]]
    preview: GlossaryImportPreviewResponse


def _bool(value: object, *, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    normalized = str(value).strip().casefold()
    if normalized in {"true", "yes", "y", "1"}:
        return True
    if normalized in {"false", "no", "n", "0"}:
        return False
    raise ValueError("Value must be a boolean.")


def _text(value: object, *, required: bool = False) -> str | None:
    normalized = str(value).strip() if value is not None else ""
    if required and not normalized:
        raise ValueError("Value is required.")
    return normalized or None


def _required_text(value: object) -> str:
    normalized = _text(value, required=True)
    assert normalized is not None
    return normalized


def _integer(value: object, *, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        raise ValueError("Boolean is not an integer.")  # noqa: TRY004
    result = int(cast(Any, value))
    if isinstance(value, float) and result != value:
        raise ValueError("Integer value must not contain decimals.")
    return result


def _uuid(value: object) -> UUID | None:
    text = _text(value)
    return UUID(text) if text else None


def _date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def _json_value(value: object) -> object:
    if isinstance(value, (UUID, date, datetime)):
        return value.isoformat() if not isinstance(value, UUID) else str(value)
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    return value


class GlossaryImportService(GlossaryServiceBase):
    """Validate the complete workbook before any database mutation."""

    def __init__(
        self,
        session: AsyncSession,
        user: User,
        metadata: RequestMetadata,
        *,
        maximum_rows: int = 100_000,
        maximum_bytes: int = 25 * 1024 * 1024,
        regex_max_length: int = 500,
        regex_timeout_ms: int = 100,
    ) -> None:
        super().__init__(session, user, metadata)
        self.maximum_rows = maximum_rows
        self.maximum_bytes = maximum_bytes
        self.matching = GlossaryMatchingService(
            regex_max_length=regex_max_length,
            regex_timeout_ms=regex_timeout_ms,
        )

    @staticmethod
    def template() -> tuple[bytes, str]:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in SHEET_HEADERS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1F4E78",
                )
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = (
                f"A1:{sheet.cell(row=1, column=len(headers)).coordinate}"
            )
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), "glossary_import_template.xlsx"

    async def preview(
        self,
        *,
        filename: str | None,
        content: bytes,
    ) -> GlossaryImportPreviewResponse:
        return (
            await self._validate(filename=filename, content=content)
        ).preview

    async def confirm(
        self,
        *,
        mode: GlossaryImportMode,
        filename: str | None,
        content: bytes,
    ) -> GlossaryImportConfirmResponse:
        validation = await self._validate(
            filename=filename,
            content=content,
        )
        if not validation.preview.valid:
            raise glossary_error(
                "Workbook contains invalid rows. Run preview and fix all "
                "issues before confirmation.",
                field="file",
            )
        created: Counter[str] = Counter()
        updated: Counter[str] = Counter()
        skipped: Counter[str] = Counter()

        profiles = {
            item.code: item
            for item in (
                await self.session.scalars(select(GlossaryProfile))
            ).all()
        }
        for row in validation.rows["Profiles"]:
            profile_payload = row.payload
            assert isinstance(profile_payload, GlossaryProfileCreate)
            profile_item = profiles.get(profile_payload.code.upper())
            if (
                profile_item is not None
                and mode is GlossaryImportMode.CREATE_ONLY
            ):
                skipped["profiles"] += 1
                continue
            if profile_item is None:
                profile_item = GlossaryProfile(
                    **profile_payload.model_dump(by_alias=False),
                    created_by=self.user.id,
                    updated_by=self.user.id,
                )
                self.session.add(profile_item)
                profiles[profile_item.code] = profile_item
                created["profiles"] += 1
            else:
                for field_name, value in profile_payload.model_dump(
                    by_alias=False
                ).items():
                    setattr(profile_item, field_name, value)
                profile_item.version += 1
                profile_item.updated_by = self.user.id
                updated["profiles"] += 1
        await self.session.flush()

        terms = {
            (item.profile.code, item.term_code): item
            for item in (
                await self.session.scalars(
                    select(GlossaryTerm).join(GlossaryTerm.profile)
                )
            ).all()
        }
        term_by_code: dict[str, GlossaryTerm] = {}
        for row in validation.rows["Terms"]:
            term_payload = row.payload
            assert isinstance(term_payload, GlossaryTermCreate)
            profile_code = str(row.context["profileCode"])
            term_key = (profile_code, term_payload.term_code.upper())
            term_item = terms.get(term_key)
            if (
                term_item is not None
                and mode is GlossaryImportMode.CREATE_ONLY
            ):
                skipped["terms"] += 1
                term_by_code[term_payload.term_code.upper()] = term_item
                continue
            if term_item is None:
                term_item = GlossaryTerm(
                    **term_payload.model_dump(
                        by_alias=False,
                        exclude={"glossary_profile_id"},
                    ),
                    glossary_profile_id=profiles[profile_code].id,
                    created_by=self.user.id,
                    updated_by=self.user.id,
                )
                self.session.add(term_item)
                terms[term_key] = term_item
                created["terms"] += 1
            else:
                for field_name, value in term_payload.model_dump(
                    by_alias=False,
                    exclude={"glossary_profile_id"}
                ).items():
                    setattr(term_item, field_name, value)
                term_item.updated_by = self.user.id
                updated["terms"] += 1
            term_by_code[term_payload.term_code.upper()] = term_item
        await self.session.flush()

        translations: dict[
            tuple[str, str, str],
            GlossaryTranslation,
        ] = {}
        existing_translations = (
            await self.session.scalars(
                select(GlossaryTranslation).options(
                    selectinload(GlossaryTranslation.variants)
                )
            )
        ).all()
        term_code_by_id = {
            term.id: term.term_code for term in term_by_code.values()
        }
        for existing_translation in existing_translations:
            term_code = term_code_by_id.get(
                existing_translation.glossary_term_id
            )
            if term_code:
                translations[
                    (
                        term_code,
                        existing_translation.language_code.value,
                        existing_translation.normalised_term,
                    )
                ] = existing_translation
        for row in validation.rows["Translations"]:
            translation_payload = row.payload
            assert isinstance(
                translation_payload,
                GlossaryTranslationCreate,
            )
            term_code = row.key
            term = term_by_code[term_code]
            normalized = normalize_term(
                translation_payload.term_text,
                case_sensitive=term.is_case_sensitive,
            )
            translation_key = (
                term_code,
                translation_payload.language_code.value,
                normalized,
            )
            translation_item = translations.get(translation_key)
            if (
                translation_item is not None
                and mode is GlossaryImportMode.CREATE_ONLY
            ):
                skipped["translations"] += 1
                continue
            if translation_item is None:
                translation_item = GlossaryTranslation(
                    **translation_payload.model_dump(by_alias=False),
                    glossary_term_id=term.id,
                    normalised_term=normalized,
                )
                translation_item.variants = []
                self.session.add(translation_item)
                translations[translation_key] = translation_item
                created["translations"] += 1
            else:
                for field_name, value in translation_payload.model_dump(
                    by_alias=False
                ).items():
                    setattr(translation_item, field_name, value)
                updated["translations"] += 1
        await self.session.flush()

        for row in validation.rows["Variants"]:
            variant_payload = row.payload
            assert isinstance(variant_payload, GlossaryVariantCreate)
            translation_key = (
                row.key,
                str(row.context["languageCode"]),
                str(row.context["preferredNormalised"]),
            )
            translation = translations[translation_key]
            term = term_by_code[row.key]
            normalized = normalize_term(
                variant_payload.variant_text,
                case_sensitive=term.is_case_sensitive,
            )
            existing = next(
                (
                    item
                    for item in translation.variants
                    if item.normalised_variant == normalized
                ),
                None,
            )
            if existing is not None and mode is GlossaryImportMode.CREATE_ONLY:
                skipped["variants"] += 1
                continue
            values = variant_payload.model_dump(by_alias=False)
            if (
                variant_payload.variant_type.value == "FORBIDDEN_VARIANT"
            ):
                values["is_allowed"] = False
            if existing is None:
                self.session.add(
                    GlossaryTermVariant(
                        **values,
                        glossary_translation_id=translation.id,
                        normalised_variant=normalized,
                    )
                )
                created["variants"] += 1
            else:
                for field, value in values.items():
                    setattr(existing, field, value)
                updated["variants"] += 1

        for row in validation.rows["Exceptions"]:
            exception_payload = row.payload
            assert isinstance(exception_payload, GlossaryExceptionCreate)
            term = term_by_code[row.key]
            self.session.add(
                GlossaryException(
                    **exception_payload.model_dump(
                        by_alias=False,
                        exclude={"glossary_term_id"},
                    ),
                    glossary_term_id=term.id,
                    created_by=self.user.id,
                )
            )
            created["exceptions"] += 1
        await self.session.flush()
        result = GlossaryImportConfirmResponse(
            mode=mode,
            total_rows=validation.preview.total_rows,
            created=dict(created),
            updated=dict(updated),
            skipped=dict(skipped),
        )
        await self.audit(
            action=AuditAction.IMPORT_GLOSSARY,
            entity_type="Glossary",
            entity_id=None,
            description="Glossary workbook imported.",
            new_values=result.model_dump(mode="json", by_alias=True),
        )
        await self.commit_or_conflict(
            message="Glossary import conflicts with existing data.",
            field="file",
        )
        return result

    async def _validate(
        self,
        *,
        filename: str | None,
        content: bytes,
    ) -> GlossaryWorkbookValidation:
        workbook = self._load(filename=filename, content=content)
        raw_rows = self._read_rows(workbook)
        existing_profiles = {
            item.code: item
            for item in (
                await self.session.scalars(select(GlossaryProfile))
            ).all()
        }
        validated: dict[str, list[ValidatedImportRow]] = {
            name: [] for name in SHEET_HEADERS
        }
        issues: list[GlossaryImportIssue] = []
        preview: dict[str, list[dict[str, object]]] = {
            name: [] for name in SHEET_HEADERS
        }
        profile_codes = set(existing_profiles)
        seen_profile_codes: set[str] = set()
        term_definitions: dict[str, GlossaryTermCreate] = {}
        term_profile_codes: dict[str, str] = {}

        for row_number, row in raw_rows["Profiles"]:
            try:
                code = str(_text(row["code"], required=True)).upper()
                if code in seen_profile_codes:
                    raise ValueError("Duplicate profile code in workbook.")
                profile_payload = GlossaryProfileCreate(
                    code=code,
                    name=_required_text(row["name"]),
                    description=_text(row["description"]),
                    scope_type=GlossaryScopeType(
                        _required_text(row["scope_type"])
                    ),
                    department_id=_uuid(row["department_id"]),
                    document_type_id=_uuid(row["document_type_id"]),
                    is_default=_bool(row["is_default"], default=False),
                    is_active=_bool(row["is_active"], default=True),
                )
                seen_profile_codes.add(code)
                profile_codes.add(code)
                validated["Profiles"].append(
                    ValidatedImportRow(
                        "Profiles",
                        row_number,
                        code,
                        profile_payload,
                        {},
                    )
                )
                preview["Profiles"].append(
                    profile_payload.model_dump(mode="json", by_alias=True)
                )
            except (ValueError, ValidationError) as exc:
                issues.append(self._issue("Profiles", row_number, exc))

        seen_term_codes: set[str] = set()
        for row_number, row in raw_rows["Terms"]:
            try:
                profile_code = str(
                    _text(row["profile_code"], required=True)
                ).upper()
                if profile_code not in profile_codes:
                    raise ValueError("Referenced profile does not exist.")
                term_code = str(
                    _text(row["term_code"], required=True)
                ).upper()
                if term_code in seen_term_codes:
                    raise ValueError("Duplicate term code in workbook.")
                profile = existing_profiles.get(profile_code)
                profile_id = (
                    profile.id
                    if profile is not None
                    else UUID(int=0)
                )
                term_payload = GlossaryTermCreate(
                    glossary_profile_id=profile_id,
                    term_code=term_code,
                    concept_name=_required_text(row["concept_name"]),
                    description=_text(row["description"]),
                    term_type=GlossaryTermType(
                        _required_text(row["term_type"])
                    ),
                    severity=GlossaryTermSeverity(
                        _required_text(row["severity"])
                    ),
                    is_case_sensitive=_bool(
                        row["is_case_sensitive"],
                        default=False,
                    ),
                    match_whole_word=_bool(
                        row["match_whole_word"],
                        default=True,
                    ),
                    allow_inflection=_bool(
                        row["allow_inflection"],
                        default=False,
                    ),
                    is_regex=_bool(row["is_regex"], default=False),
                    is_active=_bool(row["is_active"], default=True),
                    notes=_text(row["notes"]),
                )
                seen_term_codes.add(term_code)
                term_definitions[term_code] = term_payload
                term_profile_codes[term_code] = profile_code
                validated["Terms"].append(
                    ValidatedImportRow(
                        "Terms",
                        row_number,
                        term_code,
                        term_payload,
                        {"profileCode": profile_code},
                    )
                )
                data = term_payload.model_dump(mode="json", by_alias=True)
                data["profileCode"] = profile_code
                preview["Terms"].append(data)
            except (ValueError, ValidationError) as exc:
                issues.append(self._issue("Terms", row_number, exc))

        seen_translations: set[tuple[str, str, str]] = set()
        preferred_counts: Counter[tuple[str, str]] = Counter()
        languages_by_term: dict[str, set[str]] = defaultdict(set)
        translation_rows: dict[
            tuple[str, str, str],
            GlossaryTranslationCreate,
        ] = {}
        for row_number, row in raw_rows["Translations"]:
            try:
                term_code = str(
                    _text(row["term_code"], required=True)
                ).upper()
                term = term_definitions.get(term_code)
                if term is None:
                    raise ValueError("Referenced term does not exist.")
                translation_payload = GlossaryTranslationCreate(
                    language_code=GlossaryLanguageCode(
                        _required_text(row["language_code"])
                    ),
                    term_text=_required_text(row["term_text"]),
                    is_preferred=_bool(
                        row["is_preferred"],
                        default=False,
                    ),
                    is_forbidden=_bool(
                        row["is_forbidden"],
                        default=False,
                    ),
                    is_required=_bool(
                        row["is_required"],
                        default=False,
                    ),
                    priority=_integer(row["priority"]),
                    usage_note=_text(row["usage_note"]),
                    example_text=_text(row["example_text"]),
                    is_active=_bool(row["is_active"], default=True),
                )
                normalized = normalize_term(
                    translation_payload.term_text,
                    case_sensitive=term.is_case_sensitive,
                )
                key = (
                    term_code,
                    translation_payload.language_code.value,
                    normalized,
                )
                if key in seen_translations:
                    raise ValueError("Duplicate translation in workbook.")
                if term.is_regex:
                    self.matching.regex.validate(
                        translation_payload.term_text,
                        case_sensitive=term.is_case_sensitive,
                    )
                seen_translations.add(key)
                translation_rows[key] = translation_payload
                languages_by_term[term_code].add(
                    translation_payload.language_code.value
                )
                if translation_payload.is_preferred:
                    preferred_counts[
                        (
                            term_code,
                            translation_payload.language_code.value,
                        )
                    ] += 1
                    if preferred_counts[
                        (
                            term_code,
                            translation_payload.language_code.value,
                        )
                    ] > 1:
                        raise ValueError(
                            "Conflicting preferred translations."
                        )
                validated["Translations"].append(
                    ValidatedImportRow(
                        "Translations",
                        row_number,
                        term_code,
                        translation_payload,
                        {"normalisedTerm": normalized},
                    )
                )
                preview["Translations"].append(
                    translation_payload.model_dump(
                        mode="json",
                        by_alias=True,
                    )
                    | {"termCode": term_code}
                )
            except (ValueError, ValidationError) as exc:
                issues.append(
                    self._issue("Translations", row_number, exc)
                )

        for term_code, term in term_definitions.items():
            if term.term_type is not GlossaryTermType.REQUIRED:
                continue
            missing = {
                language.value for language in GlossaryLanguageCode
            } - languages_by_term[term_code]
            if missing:
                row_number = next(
                    row.row_number
                    for row in validated["Terms"]
                    if row.key == term_code
                )
                issues.append(
                    GlossaryImportIssue(
                        sheet="Terms",
                        row_number=row_number,
                        code="MISSING_REQUIRED_TRANSLATIONS",
                        field="termType",
                        message=(
                            "Required term is missing translations: "
                            + ", ".join(sorted(missing))
                        ),
                    )
                )

        seen_variants: set[tuple[str, str, str]] = set()
        for row_number, row in raw_rows["Variants"]:
            try:
                term_code = str(
                    _text(row["term_code"], required=True)
                ).upper()
                term = term_definitions.get(term_code)
                if term is None:
                    raise ValueError("Referenced term does not exist.")
                language = str(
                    _text(row["language_code"], required=True)
                ).casefold()
                preferred_text = str(
                    _text(row["preferred_term_text"], required=True)
                )
                preferred_normalized = normalize_term(
                    preferred_text,
                    case_sensitive=term.is_case_sensitive,
                )
                translation_key = (
                    term_code,
                    language,
                    preferred_normalized,
                )
                if translation_key not in translation_rows:
                    raise ValueError(
                        "Referenced preferred translation does not exist."
                    )
                variant_payload = GlossaryVariantCreate(
                    variant_text=_required_text(row["variant_text"]),
                    variant_type=GlossaryVariantType(
                        _required_text(row["variant_type"])
                    ),
                    is_allowed=_bool(row["is_allowed"], default=True),
                    is_active=_bool(row["is_active"], default=True),
                )
                normalized = normalize_term(
                    variant_payload.variant_text,
                    case_sensitive=term.is_case_sensitive,
                )
                key = (term_code, language, normalized)
                if key in seen_variants:
                    raise ValueError("Duplicate variant in workbook.")
                if normalized == preferred_normalized:
                    raise ValueError(
                        "Variant must differ from its translation."
                    )
                seen_variants.add(key)
                validated["Variants"].append(
                    ValidatedImportRow(
                        "Variants",
                        row_number,
                        term_code,
                        variant_payload,
                        {
                            "languageCode": language,
                            "preferredNormalised": preferred_normalized,
                        },
                    )
                )
                preview["Variants"].append(
                    variant_payload.model_dump(
                        mode="json",
                        by_alias=True,
                    )
                    | {
                        "termCode": term_code,
                        "languageCode": language,
                        "preferredTermText": preferred_text,
                    }
                )
            except (ValueError, ValidationError) as exc:
                issues.append(self._issue("Variants", row_number, exc))

        for row_number, row in raw_rows["Exceptions"]:
            try:
                term_code = str(
                    _text(row["term_code"], required=True)
                ).upper()
                if term_code not in term_definitions:
                    raise ValueError("Referenced term does not exist.")
                exception_payload = GlossaryExceptionCreate(
                    glossary_term_id=UUID(int=0),
                    scope_type=GlossaryExceptionScopeType(
                        _required_text(row["scope_type"])
                    ),
                    department_id=_uuid(row["department_id"]),
                    document_id=_uuid(row["document_id"]),
                    document_revision_id=_uuid(
                        row["document_revision_id"]
                    ),
                    document_file_id=_uuid(row["document_file_id"]),
                    section_definition_id=_uuid(
                        row["section_definition_id"]
                    ),
                    language_code=(
                        GlossaryLanguageCode(language_code)
                        if (
                            language_code := _text(
                                row["language_code"]
                            )
                        )
                        else None
                    ),
                    exception_type=GlossaryExceptionType(
                        _required_text(row["exception_type"])
                    ),
                    reason=_required_text(row["reason"]),
                    effective_from=_date(row["effective_from"]),
                    effective_to=_date(row["effective_to"]),
                    is_active=_bool(row["is_active"], default=True),
                    approved_by=_uuid(row["approved_by"]),
                )
                validated["Exceptions"].append(
                    ValidatedImportRow(
                        "Exceptions",
                        row_number,
                        term_code,
                        exception_payload,
                        {},
                    )
                )
                preview["Exceptions"].append(
                    exception_payload.model_dump(
                        mode="json",
                        by_alias=True,
                    )
                    | {"termCode": term_code}
                )
            except (ValueError, ValidationError) as exc:
                issues.append(self._issue("Exceptions", row_number, exc))

        total_rows = sum(len(rows) for rows in raw_rows.values())
        invalid_row_keys = {
            (issue.sheet, issue.row_number) for issue in issues
        }
        invalid_counts = Counter(
            sheet for sheet, _ in invalid_row_keys
        )
        summaries = [
            GlossaryImportSheetSummary(
                sheet=name,
                total_rows=len(raw_rows[name]),
                valid_rows=(
                    len(raw_rows[name]) - invalid_counts[name]
                ),
                invalid_rows=invalid_counts[name],
            )
            for name in SHEET_HEADERS
        ]
        response = GlossaryImportPreviewResponse(
            valid=not issues,
            total_rows=total_rows,
            valid_rows=total_rows - len(invalid_row_keys),
            invalid_rows=len(invalid_row_keys),
            sheets=summaries,
            issues=issues[:5000],
            preview={
                name: [
                    {
                        key: _json_value(value)
                        for key, value in item.items()
                    }
                    for item in rows[:100]
                ]
                for name, rows in preview.items()
            },
            warnings=(
                ["Issue list is limited to 5,000 entries."]
                if len(issues) > 5000
                else []
            ),
        )
        return GlossaryWorkbookValidation(validated, response)

    def _load(self, *, filename: str | None, content: bytes):
        if not filename or Path(filename).suffix.casefold() != ".xlsx":
            raise glossary_error(
                "Glossary import must be an XLSX workbook.",
                field="file",
            )
        if not content or len(content) > self.maximum_bytes:
            raise glossary_error(
                "Glossary workbook is empty or exceeds the size limit.",
                field="file",
            )
        stream = BytesIO(content)
        if not is_zipfile(stream):
            raise glossary_error(
                "Glossary workbook signature is invalid.",
                field="file",
            )
        stream.seek(0)
        try:
            workbook = load_workbook(
                stream,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise glossary_error(
                "Glossary workbook could not be read.",
                field="file",
            ) from exc
        missing = set(SHEET_HEADERS) - set(workbook.sheetnames)
        if missing:
            workbook.close()
            raise glossary_error(
                "Workbook is missing sheets: "
                + ", ".join(sorted(missing)),
                field="file",
            )
        return workbook

    def _read_rows(
        self,
        workbook,
    ) -> dict[str, list[tuple[int, dict[str, object]]]]:
        result: dict[str, list[tuple[int, dict[str, object]]]] = {}
        total = 0
        try:
            for name, expected in SHEET_HEADERS.items():
                sheet = workbook[name]
                rows = sheet.iter_rows()
                header_cells = next(rows, ())
                headers = tuple(
                    str(cell.value or "").strip() for cell in header_cells
                )
                if headers[: len(expected)] != expected:
                    raise glossary_error(
                        f"Sheet '{name}' headers do not match the template.",
                        field="file",
                    )
                items: list[tuple[int, dict[str, object]]] = []
                for row_number, cells in enumerate(rows, start=2):
                    if any(cell.data_type == "f" for cell in cells):
                        raise glossary_error(
                            "Formula cells are not allowed in glossary "
                            f"imports ({name} row {row_number}).",
                            field="file",
                        )
                    values = [cell.value for cell in cells[: len(expected)]]
                    if not any(value not in (None, "") for value in values):
                        continue
                    items.append(
                        (
                            row_number,
                            dict(zip(expected, values, strict=True)),
                        )
                    )
                    total += 1
                    if total > self.maximum_rows:
                        raise glossary_error(
                            "Glossary workbook exceeds the configured row "
                            "limit.",
                            field="file",
                        )
                result[name] = items
        finally:
            workbook.close()
        return result

    @staticmethod
    def _issue(
        sheet: str,
        row_number: int,
        exc: ValueError | ValidationError,
    ) -> GlossaryImportIssue:
        if isinstance(exc, ValidationError):
            first = exc.errors()[0]
            field = ".".join(str(item) for item in first["loc"]) or None
            message = str(first["msg"])
        else:
            field = None
            message = str(exc)
        code = (
            "DUPLICATE"
            if "duplicate" in message.casefold()
            else "INVALID_ROW"
        )
        return GlossaryImportIssue(
            sheet=sheet,
            row_number=row_number,
            code=code,
            field=field,
            message=message,
        )
