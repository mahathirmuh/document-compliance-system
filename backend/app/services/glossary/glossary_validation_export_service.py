"""Authenticated, bounded glossary validation JSON/XLSX exports."""

from __future__ import annotations

import json
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.models.user import User
from app.schemas.glossary_validation import (
    GlossaryFindingListResponse,
    GlossaryMatchListResponse,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.glossary.base import GlossaryServiceBase, glossary_error
from app.services.glossary.glossary_import_service import XLSX_CONTENT_TYPE
from app.services.glossary.glossary_summary_service import (
    GlossarySummaryService,
)


class GlossaryValidationExportService(GlossaryServiceBase):
    """Export match snippets and findings, never full document content."""

    def __init__(
        self,
        session: AsyncSession,
        user: User,
        metadata: RequestMetadata,
        *,
        maximum_rows: int = 100_000,
    ) -> None:
        super().__init__(session, user, metadata)
        self.maximum_rows = maximum_rows
        self.summary_service = GlossarySummaryService(
            session,
            user,
            metadata,
        )

    async def export(
        self,
        run_id: UUID,
        *,
        export_format: str,
    ) -> tuple[bytes, str, str]:
        run = await self.summary_service.run(run_id)
        summary = await self.summary_service.summary(run_id)
        matches = await self.summary_service.list_matches(
            run_id,
            page=1,
            page_size=self.maximum_rows,
        )
        findings = await self.summary_service.list_findings(
            run_id,
            page=1,
            page_size=self.maximum_rows,
        )
        if (
            matches.total_items > self.maximum_rows
            or findings.total_items > self.maximum_rows
        ):
            raise glossary_error(
                "Glossary validation export exceeds the row limit.",
                field="format",
            )
        normalized = export_format.casefold()
        if normalized == "json":
            content = json.dumps(
                {
                    "run": run.model_dump(mode="json", by_alias=True),
                    "summary": summary.model_dump(
                        mode="json",
                        by_alias=True,
                    ),
                    "matches": [
                        item.model_dump(mode="json", by_alias=True)
                        for item in matches.items
                    ],
                    "findings": [
                        item.model_dump(mode="json", by_alias=True)
                        for item in findings.items
                    ],
                    "disclaimer": (
                        "Glossary matches are review signals and do not prove "
                        "translation correctness."
                    ),
                },
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8")
            media_type = "application/json"
            filename = f"glossary_validation_{run.id}.json"
        elif normalized == "xlsx":
            content = self._xlsx(matches, findings, summary)
            media_type = XLSX_CONTENT_TYPE
            filename = f"glossary_validation_{run.id}.xlsx"
        else:
            raise glossary_error(
                "Glossary validation export format must be xlsx or json.",
                field="format",
            )
        await self.audit(
            action=AuditAction.EXPORT_GLOSSARY_VALIDATION,
            entity_type="GlossaryValidationRun",
            entity_id=run.id,
            description="Glossary validation exported.",
            new_values={
                "format": normalized,
                "matchCount": matches.total_items,
                "findingCount": findings.total_items,
            },
        )
        await self.session.commit()
        return content, filename, media_type

    @staticmethod
    def _xlsx(
        matches: GlossaryMatchListResponse,
        findings: GlossaryFindingListResponse,
        summary,
    ) -> bytes:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        for key, value in summary.model_dump(
            mode="json",
            by_alias=True,
        ).items():
            summary_sheet.append([key, str(value)])
        match_sheet = workbook.create_sheet("Matches")
        match_sheet.append(
            (
                "term_code",
                "concept_name",
                "language_code",
                "matched_text",
                "match_type",
                "is_preferred",
                "is_forbidden",
                "source_reference",
                "exception_id",
            )
        )
        for match in matches.items:
            match_sheet.append(
                (
                    match.term_code,
                    match.concept_name,
                    match.language_code.value,
                    match.matched_text,
                    match.match_type.value,
                    match.is_preferred,
                    match.is_forbidden,
                    match.source_reference,
                    str(match.exception_id) if match.exception_id else None,
                )
            )
        finding_sheet = workbook.create_sheet("Findings")
        finding_sheet.append(
            (
                "finding_code",
                "severity",
                "status",
                "title",
                "language_code",
                "source_reference",
                "is_repeat",
            )
        )
        for finding in findings.items:
            finding_sheet.append(
                (
                    finding.finding_code,
                    finding.severity,
                    finding.status,
                    finding.title,
                    finding.language_code.value
                    if finding.language_code
                    else None,
                    finding.source_reference,
                    finding.is_repeat,
                )
            )
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1F4E78",
                )
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
