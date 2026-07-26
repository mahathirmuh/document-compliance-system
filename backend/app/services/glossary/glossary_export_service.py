"""Scoped glossary JSON/XLSX export with spreadsheet safety."""

from __future__ import annotations

import json
from datetime import date, datetime
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.models.glossary_exception import GlossaryException
from app.models.glossary_profile import GlossaryProfile
from app.models.glossary_term import GlossaryTerm
from app.models.user import User
from app.repositories.glossary_exception_repository import (
    GlossaryExceptionRepository,
)
from app.repositories.glossary_profile_repository import (
    GlossaryProfileRepository,
)
from app.repositories.glossary_term_repository import GlossaryTermRepository
from app.services.auth.auth_service import RequestMetadata
from app.services.glossary.base import GlossaryServiceBase, glossary_error
from app.services.glossary.glossary_import_service import (
    SHEET_HEADERS,
    XLSX_CONTENT_TYPE,
)

JSON_CONTENT_TYPE = "application/json"


def _excel_safe(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    return value


def _enum_value(value: object) -> object:
    return getattr(value, "value", value)


class GlossaryExportService(GlossaryServiceBase):
    """Export only profiles and exceptions visible to the current user."""

    def __init__(
        self,
        session: AsyncSession,
        user: User,
        metadata: RequestMetadata,
        *,
        maximum_rows: int = 100_000,
    ) -> None:
        super().__init__(session, user, metadata)
        self.maximum_rows = maximum_rows
        self.profiles = GlossaryProfileRepository(session)
        self.terms = GlossaryTermRepository(session)
        self.exceptions = GlossaryExceptionRepository(session)

    async def export(
        self,
        *,
        export_format: str,
        department_id: UUID | None = None,
        profile_ids: list[UUID] | None = None,
        include_inactive: bool = False,
    ) -> tuple[bytes, str, str]:
        if profile_ids:
            profiles = await self.profiles.list_by_ids(
                profile_ids,
                department_ids=self.department_ids,
                active_only=not include_inactive,
            )
            if len(profiles) != len(profile_ids):
                raise glossary_error(
                    "One or more profiles are outside your scope.",
                    field="profileIds",
                    status_code=404,
                )
        else:
            profiles, _ = await self.profiles.list_page(
                department_ids=self.department_ids,
                department_id=department_id,
                is_active=None if include_inactive else True,
                page=1,
                page_size=self.maximum_rows,
            )
        terms = await self.terms.list_for_profiles(
            [profile.id for profile in profiles],
            active_only=not include_inactive,
        )
        term_ids = [term.id for term in terms]
        all_exceptions = await self.exceptions.list_for_terms(
            term_ids,
            department_ids=self.department_ids,
            include_inactive=include_inactive,
        )
        exceptions = list(all_exceptions)
        row_count = (
            len(profiles)
            + len(terms)
            + sum(len(term.translations) for term in terms)
            + sum(
                len(translation.variants)
                for term in terms
                for translation in term.translations
            )
            + len(exceptions)
        )
        if row_count > self.maximum_rows:
            raise glossary_error(
                "Glossary export exceeds the configured row limit.",
                field="format",
            )
        normalized_format = export_format.casefold()
        if normalized_format == "json":
            content = self._json(profiles, terms, exceptions)
            filename = "glossary_export.json"
            media_type = JSON_CONTENT_TYPE
        elif normalized_format == "xlsx":
            content = self._xlsx(profiles, terms, exceptions)
            filename = "glossary_export.xlsx"
            media_type = XLSX_CONTENT_TYPE
        else:
            raise glossary_error(
                "Glossary export format must be xlsx or json.",
                field="format",
            )
        await self.audit(
            action=AuditAction.EXPORT_GLOSSARY,
            entity_type="Glossary",
            entity_id=None,
            description="Glossary exported.",
            new_values={
                "format": normalized_format,
                "profileCount": len(profiles),
                "termCount": len(terms),
                "exceptionCount": len(exceptions),
            },
        )
        await self.session.commit()
        return content, filename, media_type

    @staticmethod
    def _json(
        profiles: list[GlossaryProfile],
        terms: list[GlossaryTerm],
        exceptions: list[GlossaryException],
    ) -> bytes:
        payload = {
            "schemaVersion": "1.0",
            "generatedAt": datetime.now().astimezone().isoformat(),
            "profiles": [
                {
                    "id": str(item.id),
                    "code": item.code,
                    "name": item.name,
                    "description": item.description,
                    "scopeType": item.scope_type.value,
                    "departmentId": (
                        str(item.department_id)
                        if item.department_id
                        else None
                    ),
                    "documentTypeId": (
                        str(item.document_type_id)
                        if item.document_type_id
                        else None
                    ),
                    "isDefault": item.is_default,
                    "isActive": item.is_active,
                    "version": item.version,
                }
                for item in profiles
            ],
            "terms": [
                {
                    "id": str(term.id),
                    "profileId": str(term.glossary_profile_id),
                    "profileCode": term.profile.code,
                    "termCode": term.term_code,
                    "conceptName": term.concept_name,
                    "description": term.description,
                    "termType": term.term_type.value,
                    "severity": term.severity.value,
                    "isCaseSensitive": term.is_case_sensitive,
                    "matchWholeWord": term.match_whole_word,
                    "allowInflection": term.allow_inflection,
                    "isRegex": term.is_regex,
                    "isActive": term.is_active,
                    "notes": term.notes,
                    "translations": [
                        {
                            "id": str(translation.id),
                            "languageCode": translation.language_code.value,
                            "termText": translation.term_text,
                            "isPreferred": translation.is_preferred,
                            "isForbidden": translation.is_forbidden,
                            "isRequired": translation.is_required,
                            "priority": translation.priority,
                            "usageNote": translation.usage_note,
                            "exampleText": translation.example_text,
                            "isActive": translation.is_active,
                            "variants": [
                                {
                                    "id": str(variant.id),
                                    "variantText": variant.variant_text,
                                    "variantType": variant.variant_type.value,
                                    "isAllowed": variant.is_allowed,
                                    "isActive": variant.is_active,
                                }
                                for variant in translation.variants
                            ],
                        }
                        for translation in term.translations
                    ],
                }
                for term in terms
            ],
            "exceptions": [
                {
                    "id": str(item.id),
                    "termId": str(item.glossary_term_id),
                    "termCode": item.term.term_code,
                    "scopeType": item.scope_type.value,
                    "departmentId": (
                        str(item.department_id)
                        if item.department_id
                        else None
                    ),
                    "documentId": (
                        str(item.document_id) if item.document_id else None
                    ),
                    "documentRevisionId": (
                        str(item.document_revision_id)
                        if item.document_revision_id
                        else None
                    ),
                    "documentFileId": (
                        str(item.document_file_id)
                        if item.document_file_id
                        else None
                    ),
                    "sectionDefinitionId": (
                        str(item.section_definition_id)
                        if item.section_definition_id
                        else None
                    ),
                    "languageCode": (
                        item.language_code.value
                        if item.language_code
                        else None
                    ),
                    "exceptionType": item.exception_type.value,
                    "reason": item.reason,
                    "effectiveFrom": (
                        item.effective_from.isoformat()
                        if item.effective_from
                        else None
                    ),
                    "effectiveTo": (
                        item.effective_to.isoformat()
                        if item.effective_to
                        else None
                    ),
                    "isActive": item.is_active,
                    "approvedBy": (
                        str(item.approved_by) if item.approved_by else None
                    ),
                }
                for item in exceptions
            ],
        }
        return json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ).encode("utf-8")

    @staticmethod
    def _xlsx(
        profiles: list[GlossaryProfile],
        terms: list[GlossaryTerm],
        exceptions: list[GlossaryException],
    ) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = {
            name: workbook.create_sheet(name) for name in SHEET_HEADERS
        }
        for name, headers in SHEET_HEADERS.items():
            sheet = sheets[name]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1F4E78",
                )
            sheet.freeze_panes = "A2"
        for item in profiles:
            GlossaryExportService._append(
                sheets["Profiles"],
                (
                    item.code,
                    item.name,
                    item.description,
                    item.scope_type,
                    item.department_id,
                    item.document_type_id,
                    item.is_default,
                    item.is_active,
                ),
            )
        for term in terms:
            GlossaryExportService._append(
                sheets["Terms"],
                (
                    term.profile.code,
                    term.term_code,
                    term.concept_name,
                    term.description,
                    term.term_type,
                    term.severity,
                    term.is_case_sensitive,
                    term.match_whole_word,
                    term.allow_inflection,
                    term.is_regex,
                    term.is_active,
                    term.notes,
                ),
            )
            for translation in term.translations:
                GlossaryExportService._append(
                    sheets["Translations"],
                    (
                        term.term_code,
                        translation.language_code,
                        translation.term_text,
                        translation.is_preferred,
                        translation.is_forbidden,
                        translation.is_required,
                        translation.priority,
                        translation.usage_note,
                        translation.example_text,
                        translation.is_active,
                    ),
                )
                for variant in translation.variants:
                    GlossaryExportService._append(
                        sheets["Variants"],
                        (
                            term.term_code,
                            translation.language_code,
                            translation.term_text,
                            variant.variant_text,
                            variant.variant_type,
                            variant.is_allowed,
                            variant.is_active,
                        ),
                    )
        for exception in exceptions:
            GlossaryExportService._append(
                sheets["Exceptions"],
                (
                    exception.term.term_code,
                    exception.scope_type,
                    exception.department_id,
                    exception.document_id,
                    exception.document_revision_id,
                    exception.document_file_id,
                    exception.section_definition_id,
                    exception.language_code,
                    exception.exception_type,
                    exception.reason,
                    exception.effective_from,
                    exception.effective_to,
                    exception.is_active,
                    exception.approved_by,
                ),
            )
        for sheet in sheets.values():
            sheet.auto_filter.ref = sheet.dimensions
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _append(sheet, values: tuple[Any, ...]) -> None:
        sheet.append(
            [
                _excel_safe(_enum_value(value))
                for value in values
            ]
        )
