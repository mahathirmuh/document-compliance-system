"""Filtered, department-scoped Document Register XLSX export."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.core.config import Settings
from app.models.document import Document
from app.models.user import User
from app.repositories.document_repository import DocumentRepository
from app.schemas.document import DocumentFilter
from app.services.auth.auth_service import RequestMetadata
from app.services.documents.base import DocumentServiceBase
from app.services.documents.date_filter import created_at_utc_bounds
from app.services.documents.xlsx_safety import excel_safe
from app.utils.datetime import ensure_utc, utc_now

DOCUMENT_EXPORT_HEADERS = (
    "Company Code",
    "Department Code",
    "Department Name",
    "Section Code",
    "Section Name",
    "Document Type Code",
    "Document Type Name",
    "Document Number",
    "Base Document Code",
    "Document Title",
    "Current Revision",
    "Full Document Code",
    "Document Status",
    "Validation Rule",
    "Issue Date",
    "Effective Date",
    "Review Date",
    "Expiry Date",
    "Owner Department",
    "Document Owner",
    "SharePoint URL",
    "External Reference",
    "Remarks",
    "Archived",
    "Created At",
    "Updated At",
)

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _excel_datetime(
    value: datetime,
    timezone: ZoneInfo,
) -> datetime:
    """Convert an aware database timestamp to Excel-compatible local time."""
    return (
        ensure_utc(value)
        .astimezone(timezone)
        .replace(tzinfo=None)
    )


class DocumentExportService(DocumentServiceBase):
    """Export the same filters and department scope as the register list."""

    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        user: User,
        metadata: RequestMetadata,
    ) -> None:
        super().__init__(session, user, metadata)
        self.settings = settings
        self.documents = DocumentRepository(session)

    async def export(
        self,
        filters: DocumentFilter,
    ) -> tuple[bytes, str]:
        values = filters.model_dump(by_alias=False)
        created_from = values.pop("created_from")
        created_to = values.pop("created_to")
        (
            values["created_from_utc"],
            values["created_to_utc_exclusive"],
        ) = created_at_utc_bounds(
            created_from,
            created_to,
            self.settings.application_timezone,
        )
        values["page"] = 1
        values["page_size"] = self.settings.document_register_export_max_rows
        items, total = await self.documents.list(
            **values,
            scope_all_departments=self.policy.view_all_departments,
            scope_department_id=self.policy.scope_department_id,
        )
        workbook = Workbook(write_only=False)
        sheet = workbook.active
        sheet.title = "Document Register"
        sheet.append(DOCUMENT_EXPORT_HEADERS)
        application_timezone = ZoneInfo(
            self.settings.application_timezone
        )
        for document in items:
            row = self._row(document, application_timezone)
            sheet.append([excel_safe(value) for value in row])
            row_number = sheet.max_row
            sharepoint_cell = sheet.cell(row=row_number, column=21)
            if (
                isinstance(sharepoint_cell.value, str)
                and sharepoint_cell.value.startswith(("http://", "https://"))
            ):
                sharepoint_cell.hyperlink = sharepoint_cell.value
                sharepoint_cell.style = "Hyperlink"
            for column in (15, 16, 17, 18):
                sheet.cell(row=row_number, column=column).number_format = (
                    "yyyy-mm-dd"
                )
            for column in (25, 26):
                sheet.cell(row=row_number, column=column).number_format = (
                    "yyyy-mm-dd hh:mm:ss"
                )
        self._format_sheet(sheet)
        metadata = workbook.create_sheet("_metadata")
        metadata.sheet_state = "hidden"
        generated = utc_now().astimezone(
            ZoneInfo(self.settings.application_timezone)
        )
        metadata.append(("generated_at", generated.isoformat()))
        metadata.append(("available_rows", total))
        metadata.append(("exported_rows", len(items)))
        metadata.append(("truncated", total > len(items)))
        output = BytesIO()
        workbook.save(output)

        await self.audit(
            action=AuditAction.EXPORT_DOCUMENT_REGISTER,
            entity_type="document_register_export",
            entity_id=None,
            description="Exported Document Register to XLSX.",
            new_values={
                "availableRows": total,
                "exportedRows": len(items),
                "truncated": total > len(items),
                "filters": filters.model_dump(mode="json", by_alias=True),
            },
        )
        await self.session.commit()
        filename = (
            "document_register_"
            f"{generated.strftime('%Y-%m-%d_%H-%M')}.xlsx"
        )
        return output.getvalue(), filename

    @staticmethod
    def _row(
        document: Document,
        application_timezone: ZoneInfo,
    ) -> tuple[Any, ...]:
        revision = document.current_revision
        return (
            document.company_code,
            document.department.code,
            document.department.name,
            document.section.code if document.section is not None else None,
            document.section.name if document.section is not None else None,
            document.document_type.code,
            document.document_type.name,
            document.document_number,
            document.base_document_code,
            document.title,
            revision.revision_code if revision is not None else None,
            revision.full_document_code if revision is not None else None,
            (
                revision.document_status.code
                if revision is not None
                else None
            ),
            (
                revision.validation_rule.code
                if revision is not None
                and revision.validation_rule is not None
                else None
            ),
            revision.issue_date if revision is not None else None,
            revision.effective_date if revision is not None else None,
            revision.review_date if revision is not None else None,
            revision.expiry_date if revision is not None else None,
            (
                document.owner_department.code
                if document.owner_department is not None
                else None
            ),
            document.document_owner_name,
            revision.sharepoint_url if revision is not None else None,
            revision.external_reference if revision is not None else None,
            revision.remarks if revision is not None else None,
            document.is_archived,
            _excel_datetime(
                document.created_at,
                application_timezone,
            ),
            _excel_datetime(
                document.updated_at,
                application_timezone,
            ),
        )

    @staticmethod
    def _format_sheet(sheet: Any) -> None:
        fill = PatternFill("solid", fgColor="1D4ED8")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(DOCUMENT_EXPORT_HEADERS))}"
            f"{max(sheet.max_row, 1)}"
        )
        for column in range(1, len(DOCUMENT_EXPORT_HEADERS) + 1):
            maximum = 12
            for cells in sheet.iter_cols(
                min_col=column,
                max_col=column,
                min_row=1,
                max_row=sheet.max_row,
            ):
                for cell in cells:
                    value = cell.value
                    if isinstance(value, (date, datetime)):
                        length = 19
                    else:
                        length = len(str(value or ""))
                    maximum = max(maximum, length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = min(
                maximum,
                60,
            )
