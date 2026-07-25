"""Safe XLSX template, preview, and confirmation for Document Register."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.authorization import AuditAction
from app.core.config import Settings
from app.core.exceptions import ApplicationError
from app.models.department import Department
from app.models.document import Document
from app.models.document_revision import DocumentRevision
from app.models.document_status import DocumentStatus
from app.models.document_type import DocumentType
from app.models.section import Section
from app.models.user import User
from app.models.validation_rule import ValidationRule
from app.repositories.document_repository import DocumentRepository
from app.repositories.document_revision_repository import (
    DocumentRevisionRepository,
)
from app.schemas.common import ErrorDetail
from app.schemas.document_import import (
    DocumentImportMode,
    DocumentImportPreviewResponse,
    DocumentImportPreviewRow,
    DocumentImportResultResponse,
    DocumentImportRowStatus,
)
from app.schemas.document_revision import DocumentRevisionCreate
from app.services.auth.auth_service import RequestMetadata
from app.services.documents.base import DocumentServiceBase
from app.services.documents.document_code_service import (
    DocumentCodeError,
    DocumentCodeService,
)
from app.services.documents.xlsx_safety import excel_safe

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DOCUMENT_IMPORT_HEADERS = (
    "company_code",
    "department_code",
    "section_code",
    "document_type_code",
    "document_number",
    "document_title",
    "description",
    "revision",
    "document_status_code",
    "validation_rule_code",
    "issue_date",
    "effective_date",
    "review_date",
    "expiry_date",
    "owner_department_code",
    "document_owner_name",
    "sharepoint_url",
    "external_reference",
    "remarks",
)

REQUIRED_IMPORT_HEADERS = frozenset(
    {
        "company_code",
        "department_code",
        "document_type_code",
        "document_number",
        "document_title",
        "revision",
    }
)

IMPORT_QUERY_BATCH_SIZE = 500
IMPORT_AUDIT_CHUNK_SIZE = 100


@dataclass(slots=True)
class ParsedDocumentRow:
    row_number: int
    raw: dict[str, Any]
    status: DocumentImportRowStatus
    operation: str | None = None
    base_document_code: str | None = None
    revision_code: str | None = None
    department: Department | None = None
    section: Section | None = None
    document_type: DocumentType | None = None
    status_entity: DocumentStatus | None = None
    validation_rule: ValidationRule | None = None
    owner_department: Department | None = None
    existing_document: Document | None = None
    existing_revision: DocumentRevision | None = None
    revision_payload: DocumentRevisionCreate | None = None
    normalized: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    in_file_duplicate: bool = False

    def preview(self) -> DocumentImportPreviewRow:
        return DocumentImportPreviewRow(
            row_number=self.row_number,
            status=self.status,
            base_document_code=self.base_document_code,
            revision_code=self.revision_code,
            title=self.normalized.get("title"),
            department_code=self.normalized.get("departmentCode"),
            document_type_code=self.normalized.get("documentTypeCode"),
            data=_json_safe(self.normalized or self.raw),
            errors=self.errors,
            warnings=self.warnings,
        )


@dataclass(slots=True)
class ImportLookup:
    departments_by_code: dict[str, Department]
    sections_by_department_code: dict[tuple[UUID, str], Section]
    document_types_by_code: dict[str, DocumentType]
    statuses_by_code: dict[str, DocumentStatus]
    initial_status: DocumentStatus | None
    rules_by_code: dict[str, ValidationRule]
    rules_by_id: dict[UUID, ValidationRule]
    default_rules_by_type: dict[UUID | None, ValidationRule]
    documents_by_base: dict[str, Document] = field(default_factory=dict)
    revisions_by_key: dict[
        tuple[UUID, str],
        DocumentRevision,
    ] = field(default_factory=dict)


def _json_safe(values: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in values.items():
        if isinstance(value, (date, datetime)):
            result[key] = value.isoformat()
        elif value is None or isinstance(value, (str, int, float, bool)):
            result[key] = value
        else:
            result[key] = str(value)
    return result


def chunk_import_audit_changes(
    changes: list[dict[str, Any]],
    *,
    chunk_size: int = IMPORT_AUDIT_CHUNK_SIZE,
) -> list[list[dict[str, Any]]]:
    """Split import trace changes without dropping any entries."""
    if chunk_size < 1:
        raise ValueError("chunk_size must be at least 1.")
    return [
        changes[offset : offset + chunk_size]
        for offset in range(0, len(changes), chunk_size)
    ]


def _text(
    value: Any,
    *,
    required: bool = False,
    uppercase: bool = False,
    field_name: str,
) -> str | None:
    if value is None:
        normalized = ""
    else:
        normalized = str(value).strip()
    if normalized.startswith("="):
        raise ValueError(f"{field_name}: formulas are not accepted.")
    if required and not normalized:
        raise ValueError(f"{field_name}: value is required.")
    if not normalized:
        return None
    return normalized.upper() if uppercase else normalized


def _date_value(value: Any, *, field_name: str) -> date | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ValueError(
            f"{field_name}: date must use YYYY-MM-DD."
        ) from exc


def _import_error(message: str) -> ApplicationError:
    return ApplicationError(
        "Document Register import failed.",
        status_code=400,
        errors=[ErrorDetail(field="file", message=message)],
    )


class DocumentImportService(DocumentServiceBase):
    """Revalidate XLSX bytes and commit document/revision rows atomically."""

    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        user: User,
        metadata: RequestMetadata,
    ) -> None:
        super().__init__(session, user, metadata)
        self.settings = settings
        self.documents = DocumentRepository(session)
        self.revisions = DocumentRevisionRepository(session)
        self.codes = DocumentCodeService()

    async def template(self) -> tuple[bytes, str]:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Document Register"
        sheet.append(DOCUMENT_IMPORT_HEADERS)
        for cell in sheet[1]:
            header = str(cell.value)
            required = header in REQUIRED_IMPORT_HEADERS
            cell.comment = Comment(
                (
                    "Required field."
                    if required
                    else "Optional field. See Reference sheet for codes."
                ),
                "Document Compliance",
            )
        self._format_sheet(sheet, len(DOCUMENT_IMPORT_HEADERS))

        reference = workbook.create_sheet("Reference")
        reference.append(("entity", "parent_code", "code", "name"))
        queries = (
            (
                "Department",
                None,
                await self.session.scalars(
                    select(Department)
                    .where(
                        Department.deleted_at.is_(None),
                        Department.is_active.is_(True),
                    )
                    .order_by(Department.code)
                ),
            ),
            (
                "Section",
                "department",
                await self.session.scalars(
                    select(Section)
                    .options(selectinload(Section.department))
                    .where(
                        Section.deleted_at.is_(None),
                        Section.is_active.is_(True),
                    )
                    .order_by(Section.code)
                ),
            ),
            (
                "Document Type",
                None,
                await self.session.scalars(
                    select(DocumentType)
                    .where(
                        DocumentType.deleted_at.is_(None),
                        DocumentType.is_active.is_(True),
                    )
                    .order_by(DocumentType.code)
                ),
            ),
            (
                "Document Status",
                None,
                await self.session.scalars(
                    select(DocumentStatus)
                    .where(
                        DocumentStatus.deleted_at.is_(None),
                        DocumentStatus.is_active.is_(True),
                    )
                    .order_by(DocumentStatus.display_order)
                ),
            ),
            (
                "Validation Rule",
                "document_type",
                await self.session.scalars(
                    select(ValidationRule)
                    .options(selectinload(ValidationRule.document_type))
                    .where(
                        ValidationRule.deleted_at.is_(None),
                        ValidationRule.is_active.is_(True),
                    )
                    .order_by(ValidationRule.code)
                ),
            ),
        )
        for entity_name, parent_kind, scalar_result in queries:
            for entity in scalar_result.all():
                parent = None
                if parent_kind == "department":
                    parent = (
                        entity.department.code
                        if entity.department is not None
                        else None
                    )
                elif parent_kind == "document_type":
                    if entity.document_type is not None:
                        parent = entity.document_type.code
                    else:
                        parent = "GLOBAL"
                reference.append(
                    tuple(
                        excel_safe(value)
                        for value in (
                            entity_name,
                            parent,
                            entity.code,
                            entity.name,
                        )
                    )
                )
        self._format_sheet(reference, 4)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), "document_register_template.xlsx"

    async def preview(
        self,
        *,
        filename: str | None,
        content: bytes,
    ) -> DocumentImportPreviewResponse:
        rows = await self._analyze(
            filename=filename,
            content=content,
            for_update=False,
        )
        return self._preview_response(rows)

    async def confirm(
        self,
        *,
        mode: DocumentImportMode,
        filename: str | None,
        content: bytes,
    ) -> DocumentImportResultResponse:
        # Preview output is intentionally not trusted; bytes are parsed again.
        rows = await self._analyze(
            filename=filename,
            content=content,
            for_update=True,
        )
        documents_created = 0
        revisions_added = 0
        metadata_updated = 0
        duplicates_skipped = 0
        invalid_skipped = 0
        failed = 0
        audit_changes: list[dict[str, Any]] = []

        documents_by_base = {
            row.base_document_code: row.existing_document
            for row in rows
            if (
                row.base_document_code is not None
                and row.existing_document is not None
            )
        }
        try:
            for row in rows:
                if row.status is DocumentImportRowStatus.INVALID:
                    invalid_skipped += 1
                    continue
                if row.in_file_duplicate:
                    duplicates_skipped += 1
                    continue
                if (
                    row.base_document_code is None
                    or row.revision_code is None
                    or row.department is None
                    or row.document_type is None
                    or row.status_entity is None
                    or row.revision_payload is None
                ):
                    failed += 1
                    continue
                document = documents_by_base.get(row.base_document_code)
                document_created = False
                metadata_changed = False
                metadata_old: dict[str, Any] | None = None
                metadata_new: dict[str, Any] | None = None
                if document is not None:
                    self.policy.ensure_document_access(document)
                    existing_revision = row.existing_revision
                    if existing_revision is not None:
                        if mode is DocumentImportMode.UPSERT_METADATA:
                            metadata_old = self._metadata_values(document)
                            metadata_changed = self._apply_metadata(
                                document,
                                row,
                            )
                            if metadata_changed:
                                metadata_updated += 1
                                document.updated_by = self.user.id
                                metadata_new = self._metadata_values(document)
                                audit_changes.append(
                                    {
                                        "documentId": str(document.id),
                                        "baseDocumentCode": (
                                            document.base_document_code
                                        ),
                                        "revisionId": str(
                                            existing_revision.id
                                        ),
                                        "operation": "UPSERT_METADATA",
                                        "oldValues": metadata_old,
                                        "newValues": metadata_new,
                                    }
                                )
                        duplicates_skipped += 1
                        continue
                    if mode is DocumentImportMode.CREATE_ONLY:
                        duplicates_skipped += 1
                        continue
                    if mode is DocumentImportMode.UPSERT_METADATA:
                        metadata_old = self._metadata_values(document)
                        metadata_changed = self._apply_metadata(
                            document,
                            row,
                        )
                        if metadata_changed:
                            metadata_updated += 1
                            document.updated_by = self.user.id
                            metadata_new = self._metadata_values(document)
                else:
                    self.policy.ensure_create_department(row.department.id)
                    document = Document(
                        company_code=row.normalized["companyCode"],
                        department_id=row.department.id,
                        section_id=(
                            row.section.id if row.section is not None else None
                        ),
                        document_type_id=row.document_type.id,
                        document_number=row.normalized["documentNumber"],
                        base_document_code=row.base_document_code,
                        title=row.normalized["title"],
                        description=row.normalized.get("description"),
                        owner_department_id=(
                            row.owner_department.id
                            if row.owner_department is not None
                            else row.department.id
                        ),
                        document_owner_name=row.normalized.get(
                            "documentOwnerName"
                        ),
                        created_by=self.user.id,
                        updated_by=self.user.id,
                    )
                    await self.documents.create(document)
                    documents_by_base[row.base_document_code] = document
                    documents_created += 1
                    document_created = True

                old_current_revision_id = document.current_revision_id
                revision = DocumentRevision(
                    document_id=document.id,
                    revision_code=row.revision_code,
                    revision_number=self.codes.revision_number(
                        row.revision_code
                    ),
                    full_document_code=self.codes.generate_full_document_code(
                        document.base_document_code,
                        row.revision_code,
                    ),
                    document_status_id=row.status_entity.id,
                    validation_rule_id=(
                        row.validation_rule.id
                        if row.validation_rule is not None
                        else None
                    ),
                    issue_date=row.revision_payload.issue_date,
                    effective_date=row.revision_payload.effective_date,
                    review_date=row.revision_payload.review_date,
                    expiry_date=row.revision_payload.expiry_date,
                    sharepoint_url=row.revision_payload.sharepoint_url,
                    external_reference=(
                        row.revision_payload.external_reference
                    ),
                    remarks=row.revision_payload.remarks,
                    is_current=False,
                    created_by=self.user.id,
                    updated_by=self.user.id,
                )
                await self.revisions.create(revision)
                await self.revisions.set_current(revision)
                document.current_revision_id = revision.id
                document.updated_by = self.user.id
                await self.session.flush()
                revisions_added += 1
                if document_created:
                    operation = "CREATE_DOCUMENT_AND_REVISION"
                    old_values = None
                    new_values = {
                        **self._metadata_values(document),
                        "currentRevisionId": str(revision.id),
                        "revisionCode": revision.revision_code,
                        "documentStatusId": str(
                            revision.document_status_id
                        ),
                    }
                elif metadata_changed:
                    operation = "UPSERT_METADATA_AND_ADD_REVISION"
                    old_values = {
                        **(metadata_old or {}),
                        "currentRevisionId": (
                            str(old_current_revision_id)
                            if old_current_revision_id is not None
                            else None
                        ),
                    }
                    new_values = {
                        **(metadata_new or {}),
                        "currentRevisionId": str(revision.id),
                    }
                else:
                    operation = "ADD_REVISION"
                    old_values = {
                        "currentRevisionId": (
                            str(old_current_revision_id)
                            if old_current_revision_id is not None
                            else None
                        )
                    }
                    new_values = {
                        "currentRevisionId": str(revision.id),
                    }
                audit_changes.append(
                    {
                        "documentId": str(document.id),
                        "baseDocumentCode": document.base_document_code,
                        "revisionId": str(revision.id),
                        "operation": operation,
                        "oldValues": old_values,
                        "newValues": new_values,
                    }
                )

            result = DocumentImportResultResponse(
                mode=mode,
                total_rows=len(rows),
                documents_created=documents_created,
                revisions_added=revisions_added,
                metadata_updated=metadata_updated,
                duplicates_skipped=duplicates_skipped,
                invalid_skipped=invalid_skipped,
                failed=failed,
            )
            trace_id = str(uuid4())
            audit_chunks = chunk_import_audit_changes(audit_changes)
            await self.audit(
                action=AuditAction.IMPORT_DOCUMENT_REGISTER,
                entity_type="document_register_import",
                entity_id=None,
                description=(
                    f"Imported Document Register using mode {mode.value}."
                ),
                new_values={
                    **result.model_dump(mode="json", by_alias=True),
                    "traceId": trace_id,
                    "traceChangeCount": len(audit_changes),
                    "traceChunkCount": len(audit_chunks),
                    "traceChunkSize": IMPORT_AUDIT_CHUNK_SIZE,
                },
            )
            for chunk_index, chunk in enumerate(audit_chunks, start=1):
                await self.audit(
                    action=AuditAction.IMPORT_DOCUMENT_REGISTER,
                    entity_type="document_register_import_chunk",
                    entity_id=None,
                    description=(
                        "Document Register import trace "
                        f"{chunk_index}/{len(audit_chunks)}."
                    ),
                    new_values={
                        "traceId": trace_id,
                        "mode": mode.value,
                        "chunkIndex": chunk_index,
                        "chunkCount": len(audit_chunks),
                        "changeCount": len(chunk),
                        "changes": chunk,
                    },
                )
            await self.session.commit()
            return result
        except IntegrityError as exc:
            await self.session.rollback()
            raise ApplicationError(
                "Document Register import could not be committed.",
                status_code=409,
                errors=[
                    ErrorDetail(
                        field="file",
                        message=(
                            "Imported document codes conflict with data "
                            "changed by another request. Preview again."
                        ),
                    )
                ],
            ) from exc

    async def _build_lookup(
        self,
        raw_rows: list[tuple[int, dict[str, Any]]],
        *,
        for_update: bool,
    ) -> ImportLookup:
        """Preload import references and lock affected rows in stable order."""
        departments = list(
            (
                await self.session.scalars(
                    select(Department).where(
                        Department.deleted_at.is_(None)
                    )
                )
            ).all()
        )
        sections = list(
            (
                await self.session.scalars(
                    select(Section).where(Section.deleted_at.is_(None))
                )
            ).all()
        )
        document_types = list(
            (
                await self.session.scalars(
                    select(DocumentType).where(
                        DocumentType.deleted_at.is_(None)
                    )
                )
            ).all()
        )
        statuses = list(
            (
                await self.session.scalars(
                    select(DocumentStatus).where(
                        DocumentStatus.deleted_at.is_(None)
                    )
                )
            ).all()
        )
        rules = list(
            (
                await self.session.scalars(
                    select(ValidationRule).where(
                        ValidationRule.deleted_at.is_(None)
                    )
                )
            ).all()
        )
        default_rules: dict[UUID | None, ValidationRule] = {}
        for rule in sorted(rules, key=lambda item: (item.code, str(item.id))):
            if rule.is_default:
                default_rules.setdefault(rule.document_type_id, rule)
        lookup = ImportLookup(
            departments_by_code={
                entity.code: entity for entity in departments
            },
            sections_by_department_code={
                (entity.department_id, entity.code): entity
                for entity in sections
            },
            document_types_by_code={
                entity.code: entity for entity in document_types
            },
            statuses_by_code={
                entity.code: entity for entity in statuses
            },
            initial_status=next(
                (entity for entity in statuses if entity.is_initial),
                None,
            ),
            rules_by_code={entity.code: entity for entity in rules},
            rules_by_id={entity.id: entity for entity in rules},
            default_rules_by_type=default_rules,
        )

        identity_keys: set[tuple[str, str]] = set()
        for _, raw in raw_rows:
            try:
                identity_keys.add(self._identity_from_raw(raw, lookup))
            except (ValueError, DocumentCodeError):
                continue

        base_codes = sorted({key[0] for key in identity_keys})
        documents: list[Document] = []
        for offset in range(
            0,
            len(base_codes),
            IMPORT_QUERY_BATCH_SIZE,
        ):
            documents.extend(
                await self.documents.list_by_base_codes(
                    base_codes[offset : offset + IMPORT_QUERY_BATCH_SIZE]
                )
            )
        if for_update and documents:
            document_ids = sorted(
                {document.id for document in documents},
                key=str,
            )
            locked: list[Document] = []
            for offset in range(
                0,
                len(document_ids),
                IMPORT_QUERY_BATCH_SIZE,
            ):
                locked.extend(
                    await self.documents.lock_by_ids(
                        document_ids[
                            offset : offset + IMPORT_QUERY_BATCH_SIZE
                        ]
                    )
                )
            documents = locked
        lookup.documents_by_base = {
            document.base_document_code: document for document in documents
        }

        revision_keys = sorted(
            {
                (
                    lookup.documents_by_base[base_code].id,
                    revision_code,
                )
                for base_code, revision_code in identity_keys
                if base_code in lookup.documents_by_base
            },
            key=lambda item: (str(item[0]), item[1]),
        )
        revisions: list[DocumentRevision] = []
        for offset in range(
            0,
            len(revision_keys),
            IMPORT_QUERY_BATCH_SIZE,
        ):
            revisions.extend(
                await self.revisions.list_by_document_codes(
                    revision_keys[
                        offset : offset + IMPORT_QUERY_BATCH_SIZE
                    ],
                    for_update=for_update,
                )
            )
        lookup.revisions_by_key = {
            (revision.document_id, revision.revision_code): revision
            for revision in revisions
        }
        return lookup

    def _identity_from_raw(
        self,
        raw: dict[str, Any],
        lookup: ImportLookup,
    ) -> tuple[str, str]:
        company_code = self.codes.normalize_component(
            str(
                _text(
                    raw["company_code"],
                    required=True,
                    uppercase=True,
                    field_name="company_code",
                )
            ),
            field="companyCode",
        )
        department_code = str(
            _text(
                raw["department_code"],
                required=True,
                uppercase=True,
                field_name="department_code",
            )
        )
        document_type_code = str(
            _text(
                raw["document_type_code"],
                required=True,
                uppercase=True,
                field_name="document_type_code",
            )
        )
        department = lookup.departments_by_code.get(department_code)
        document_type = lookup.document_types_by_code.get(
            document_type_code
        )
        if department is None or document_type is None:
            raise ValueError("Import identity references unknown master data.")
        section_code = _text(
            raw["section_code"],
            uppercase=True,
            field_name="section_code",
        )
        section = (
            lookup.sections_by_department_code.get(
                (department.id, section_code)
            )
            if section_code
            else None
        )
        if document_type.requires_section and section is None:
            raise ValueError("Import identity requires a valid section.")
        if not document_type.requires_section and section_code:
            raise ValueError("Import identity must not include a section.")
        document_number = self.codes.normalize_document_number(
            str(
                _text(
                    raw["document_number"],
                    required=True,
                    uppercase=True,
                    field_name="document_number",
                )
            ),
            maximum=self.settings.document_number_max_length,
        )
        revision_code = self.codes.normalize_revision_code(
            str(
                _text(
                    raw["revision"],
                    required=True,
                    field_name="revision",
                )
            )
        )
        base_code = self.codes.generate_base_document_code(
            company_code=company_code,
            department_code=department.code,
            section_code=section.code if section is not None else None,
            document_type_code=document_type.code,
            document_number=document_number,
            requires_section=document_type.requires_section,
        )
        return base_code, revision_code

    @staticmethod
    def _resolve_import_rule(
        document_type: DocumentType,
        supplied_rule: ValidationRule | None,
        lookup: ImportLookup,
    ) -> ValidationRule | None:
        rule = supplied_rule
        if rule is None:
            if document_type.default_validation_rule_id is not None:
                rule = lookup.rules_by_id.get(
                    document_type.default_validation_rule_id
                )
            if rule is None:
                rule = lookup.default_rules_by_type.get(document_type.id)
            if rule is None:
                rule = lookup.default_rules_by_type.get(None)
        if rule is not None:
            if not rule.is_active:
                raise ValueError(
                    "validation_rule_code: rule must be active."
                )
            if rule.document_type_id not in (None, document_type.id):
                raise ValueError(
                    "validation_rule_code: rule does not apply to this "
                    "document type."
                )
        return rule

    async def _analyze(
        self,
        *,
        filename: str | None,
        content: bytes,
        for_update: bool,
    ) -> list[ParsedDocumentRow]:
        raw_rows = self._read_workbook(filename, content)
        lookup = await self._build_lookup(
            raw_rows,
            for_update=for_update,
        )
        rows: list[ParsedDocumentRow] = []
        seen_keys: set[tuple[str, str]] = set()
        seen_bases: set[str] = set()
        for row_number, raw in raw_rows:
            row = await self._parse_row(
                row_number,
                raw,
                lookup=lookup,
            )
            if row.base_document_code and row.revision_code:
                key = (row.base_document_code, row.revision_code)
                if key in seen_keys:
                    row.status = DocumentImportRowStatus.DUPLICATE
                    row.errors.append(
                        "Duplicate document revision in uploaded workbook."
                    )
                    row.in_file_duplicate = True
                else:
                    seen_keys.add(key)
                    if (
                        row.existing_document is None
                        and row.base_document_code in seen_bases
                        and row.status
                        is DocumentImportRowStatus.VALID_CREATE
                    ):
                        row.operation = "add_revision"
                        row.status = (
                            DocumentImportRowStatus.VALID_ADD_REVISION
                        )
                    seen_bases.add(row.base_document_code)
            rows.append(row)
        return rows

    async def _parse_row(
        self,
        row_number: int,
        raw: dict[str, Any],
        *,
        lookup: ImportLookup,
    ) -> ParsedDocumentRow:
        row = ParsedDocumentRow(
            row_number=row_number,
            raw=raw,
            status=DocumentImportRowStatus.INVALID,
        )
        try:
            company_code = self.codes.normalize_component(
                str(
                    _text(
                        raw["company_code"],
                        required=True,
                        uppercase=True,
                        field_name="company_code",
                    )
                ),
                field="companyCode",
            )
            department_code = str(
                _text(
                    raw["department_code"],
                    required=True,
                    uppercase=True,
                    field_name="department_code",
                )
            )
            document_type_code = str(
                _text(
                    raw["document_type_code"],
                    required=True,
                    uppercase=True,
                    field_name="document_type_code",
                )
            )
            document_number = self.codes.normalize_document_number(
                str(
                    _text(
                        raw["document_number"],
                        required=True,
                        uppercase=True,
                        field_name="document_number",
                    )
                ),
                maximum=self.settings.document_number_max_length,
            )
            title = str(
                _text(
                    raw["document_title"],
                    required=True,
                    field_name="document_title",
                )
            )
            if len(title) > self.settings.document_title_max_length:
                raise ValueError(
                    "document_title: exceeds configured maximum length."
                )
            revision_code = self.codes.normalize_revision_code(
                str(
                    _text(
                        raw["revision"],
                        required=True,
                        field_name="revision",
                    )
                )
            )
            department = lookup.departments_by_code.get(department_code)
            if department is None:
                raise ValueError(
                    f'department_code: "{department_code}" was not found.'
                )
            document_type = lookup.document_types_by_code.get(
                document_type_code
            )
            if document_type is None:
                raise ValueError(
                    f'document_type_code: "{document_type_code}" was not found.'
                )
            section_code = _text(
                raw["section_code"],
                uppercase=True,
                field_name="section_code",
            )
            section: Section | None = None
            if document_type.requires_section:
                if not section_code:
                    raise ValueError(
                        "section_code: required for this document type."
                    )
                section = lookup.sections_by_department_code.get(
                    (department.id, section_code)
                )
                if section is None:
                    raise ValueError(
                        "section_code: section was not found in the selected "
                        "department."
                    )
            elif section_code:
                raise ValueError(
                    "section_code: must be empty for this document type."
                )
            base_code = self.codes.generate_base_document_code(
                company_code=company_code,
                department_code=department.code,
                section_code=section.code if section is not None else None,
                document_type_code=document_type.code,
                document_number=document_number,
                requires_section=document_type.requires_section,
            )
            existing_document = lookup.documents_by_base.get(base_code)
            if existing_document is None:
                self.policy.ensure_create_department(department.id)
                if not department.is_active:
                    raise ValueError("department_code: department is inactive.")
                if section is not None and not section.is_active:
                    raise ValueError("section_code: section is inactive.")
                if not document_type.is_active:
                    raise ValueError(
                        "document_type_code: document type is inactive."
                    )
            else:
                self.policy.ensure_document_access(existing_document)
                if existing_document.is_archived:
                    raise ValueError(
                        "base_document_code: archived documents are "
                        "read-only until restored."
                    )

            status_code = _text(
                raw["document_status_code"],
                uppercase=True,
                field_name="document_status_code",
            )
            status = (
                lookup.statuses_by_code.get(status_code)
                if status_code
                else lookup.initial_status
            )
            if status is None or not status.is_active:
                raise ValueError(
                    "document_status_code: active status was not found."
                )
            rule_code = _text(
                raw["validation_rule_code"],
                uppercase=True,
                field_name="validation_rule_code",
            )
            supplied_rule = (
                lookup.rules_by_code.get(rule_code)
                if rule_code
                else None
            )
            if rule_code and supplied_rule is None:
                raise ValueError(
                    "validation_rule_code: rule was not found."
                )
            rule = self._resolve_import_rule(
                document_type,
                supplied_rule,
                lookup,
            )
            owner_code = _text(
                raw["owner_department_code"],
                uppercase=True,
                field_name="owner_department_code",
            )
            owner_department = (
                lookup.departments_by_code.get(owner_code)
                if owner_code
                else department
            )
            if owner_department is None:
                raise ValueError(
                    "owner_department_code: department was not found."
                )
            if (
                not owner_department.is_active
                and (
                    existing_document is None
                    or owner_department.id
                    != existing_document.owner_department_id
                )
            ):
                raise ValueError(
                    "owner_department_code: owner department is inactive."
                )
            revision_payload = DocumentRevisionCreate(
                revision_code=revision_code,
                document_status_id=status.id,
                validation_rule_id=rule.id if rule is not None else None,
                issue_date=_date_value(
                    raw["issue_date"],
                    field_name="issue_date",
                ),
                effective_date=_date_value(
                    raw["effective_date"],
                    field_name="effective_date",
                ),
                review_date=_date_value(
                    raw["review_date"],
                    field_name="review_date",
                ),
                expiry_date=_date_value(
                    raw["expiry_date"],
                    field_name="expiry_date",
                ),
                sharepoint_url=_text(
                    raw["sharepoint_url"],
                    field_name="sharepoint_url",
                ),
                external_reference=_text(
                    raw["external_reference"],
                    field_name="external_reference",
                ),
                remarks=_text(raw["remarks"], field_name="remarks"),
                set_as_current=True,
            )
            existing_revision = (
                lookup.revisions_by_key.get(
                    (existing_document.id, revision_code)
                )
                if existing_document is not None
                else None
            )
            row.base_document_code = base_code
            row.revision_code = revision_code
            row.department = department
            row.section = section
            row.document_type = document_type
            row.status_entity = status
            row.validation_rule = rule
            row.owner_department = owner_department
            row.existing_document = existing_document
            row.existing_revision = existing_revision
            row.revision_payload = revision_payload
            row.normalized = {
                "companyCode": company_code,
                "departmentCode": department.code,
                "sectionCode": section.code if section is not None else None,
                "documentTypeCode": document_type.code,
                "documentNumber": document_number,
                "baseDocumentCode": base_code,
                "title": title,
                "description": _text(
                    raw["description"],
                    field_name="description",
                ),
                "revisionCode": revision_code,
                "documentStatusCode": status.code,
                "validationRuleCode": rule.code if rule is not None else None,
                "issueDate": revision_payload.issue_date,
                "effectiveDate": revision_payload.effective_date,
                "reviewDate": revision_payload.review_date,
                "expiryDate": revision_payload.expiry_date,
                "ownerDepartmentCode": owner_department.code,
                "documentOwnerName": self._owner_name(
                    raw["document_owner_name"]
                ),
                "sharepointUrl": revision_payload.sharepoint_url,
                "externalReference": (
                    revision_payload.external_reference
                ),
                "remarks": revision_payload.remarks,
            }
            if existing_revision is not None:
                row.status = DocumentImportRowStatus.DUPLICATE
                row.operation = "duplicate"
                row.errors.append(
                    "Document and revision already exist in the database."
                )
            elif existing_document is not None:
                row.status = DocumentImportRowStatus.VALID_ADD_REVISION
                row.operation = "add_revision"
                if (
                    existing_document.title != title
                    or existing_document.description
                    != row.normalized["description"]
                ):
                    row.warnings.append(
                        "Existing metadata differs; it is changed only in "
                        "UPSERT_METADATA mode."
                    )
                    row.status = DocumentImportRowStatus.WARNING
            else:
                row.status = DocumentImportRowStatus.VALID_CREATE
                row.operation = "create"
            return row
        except (ValueError, DocumentCodeError, ApplicationError) as exc:
            if isinstance(exc, ApplicationError):
                row.errors.extend(
                    detail.message
                    for detail in (exc.errors or [])
                )
                if not exc.errors:
                    row.errors.append(exc.message)
            else:
                row.errors.append(str(exc))
            row.status = DocumentImportRowStatus.INVALID
            return row

    def _read_workbook(
        self,
        filename: str | None,
        content: bytes,
    ) -> list[tuple[int, dict[str, Any]]]:
        if not filename or Path(filename).suffix.lower() != ".xlsx":
            raise _import_error("Only .xlsx files are accepted.")
        max_bytes = self.settings.document_import_max_file_size_mb * 1024 * 1024
        if len(content) > max_bytes:
            raise _import_error(
                "File exceeds the configured maximum size of "
                f"{self.settings.document_import_max_file_size_mb} MB."
            )
        if (
            not content
            or not content.startswith(b"PK")
            or not is_zipfile(BytesIO(content))
        ):
            raise _import_error("The uploaded file is not a valid XLSX file.")
        self._validate_zip_safety(content, max_bytes=max_bytes)
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise _import_error("The XLSX workbook could not be read.") from exc
        sheet = workbook["Document Register"] if "Document Register" in workbook.sheetnames else workbook.active
        header_row = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
        actual = tuple(
            str(value).strip() if value is not None else ""
            for value in header_row
        )
        if actual != DOCUMENT_IMPORT_HEADERS:
            workbook.close()
            raise _import_error(
                "Invalid header. Expected exactly: "
                + ", ".join(DOCUMENT_IMPORT_HEADERS)
            )
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if all(
                value is None or str(value).strip() == ""
                for value in values
            ):
                continue
            if len(rows) >= self.settings.document_register_import_max_rows:
                workbook.close()
                raise _import_error(
                    "Workbook exceeds the configured import row limit of "
                    f"{self.settings.document_register_import_max_rows}."
                )
            rows.append(
                (
                    row_number,
                    dict(zip(DOCUMENT_IMPORT_HEADERS, values, strict=False)),
                )
            )
        workbook.close()
        return rows

    @staticmethod
    def _owner_name(value: Any) -> str | None:
        owner_name = _text(
            value,
            field_name="document_owner_name",
        )
        if owner_name is not None and len(owner_name) > 150:
            raise ValueError(
                "document_owner_name: must contain at most 150 characters."
            )
        return owner_name

    @staticmethod
    def _validate_zip_safety(content: bytes, *, max_bytes: int) -> None:
        """Reject XLSX archives whose expanded form is unsafe to inspect."""
        expanded_limit = max(
            64 * 1024 * 1024,
            min(max_bytes * 10, 256 * 1024 * 1024),
        )
        total_expanded = 0
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
                if len(entries) > 2048:
                    raise _import_error(
                        "The XLSX archive contains too many entries."
                    )
                for entry in entries:
                    if entry.flag_bits & 0x1:
                        raise _import_error(
                            "Encrypted XLSX archives are not accepted."
                        )
                    total_expanded += entry.file_size
                    if (
                        entry.file_size > expanded_limit
                        or total_expanded > expanded_limit
                    ):
                        raise _import_error(
                            "The XLSX archive expands beyond the safe limit."
                        )
                    if (
                        entry.file_size > 1024 * 1024
                        and (
                            entry.compress_size == 0
                            or entry.file_size
                            > entry.compress_size * 200
                        )
                    ):
                        raise _import_error(
                            "The XLSX archive has an unsafe compression ratio."
                        )
        except BadZipFile as exc:
            raise _import_error(
                "The uploaded file is not a valid XLSX file."
            ) from exc

    @staticmethod
    def _preview_response(
        rows: list[ParsedDocumentRow],
    ) -> DocumentImportPreviewResponse:
        return DocumentImportPreviewResponse(
            total_rows=len(rows),
            valid_create_rows=sum(
                row.status is DocumentImportRowStatus.VALID_CREATE
                for row in rows
            ),
            valid_add_revision_rows=sum(
                row.status is DocumentImportRowStatus.VALID_ADD_REVISION
                for row in rows
            ),
            warning_rows=sum(
                row.status is DocumentImportRowStatus.WARNING
                for row in rows
            ),
            duplicate_rows=sum(
                row.status is DocumentImportRowStatus.DUPLICATE
                for row in rows
            ),
            invalid_rows=sum(
                row.status is DocumentImportRowStatus.INVALID
                for row in rows
            ),
            rows=[row.preview() for row in rows],
            warnings=(
                [
                    (
                        "Confirmation re-reads the workbook and never changes "
                        "an existing revision."
                    )
                ]
                if rows
                else []
            ),
        )

    @staticmethod
    def _apply_metadata(
        document: Document,
        row: ParsedDocumentRow,
    ) -> bool:
        values = {
            "title": row.normalized["title"],
            "description": row.normalized.get("description"),
            "owner_department_id": (
                row.owner_department.id
                if row.owner_department is not None
                else document.department_id
            ),
            "document_owner_name": row.normalized.get(
                "documentOwnerName"
            ),
        }
        changed = any(
            getattr(document, field_name) != value
            for field_name, value in values.items()
        )
        if changed:
            for field_name, value in values.items():
                setattr(document, field_name, value)
        return changed

    @staticmethod
    def _metadata_values(document: Document) -> dict[str, Any]:
        return {
            "title": document.title,
            "description": document.description,
            "ownerDepartmentId": (
                str(document.owner_department_id)
                if document.owner_department_id is not None
                else None
            ),
            "documentOwnerName": document.document_owner_name,
        }

    @staticmethod
    def _format_sheet(sheet: Any, column_count: int) -> None:
        fill = PatternFill("solid", fgColor="1D4ED8")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(column_count)}{max(1, sheet.max_row)}"
        )
        for column in range(1, column_count + 1):
            maximum = 12
            for cells in sheet.iter_cols(
                min_col=column,
                max_col=column,
                min_row=1,
                max_row=sheet.max_row,
            ):
                for cell in cells:
                    maximum = max(maximum, len(str(cell.value or "")) + 2)
            sheet.column_dimensions[get_column_letter(column)].width = min(
                maximum,
                60,
            )
