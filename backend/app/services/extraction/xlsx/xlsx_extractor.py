"""Streaming XLSX extractor that never calculates workbook formulas."""

from datetime import date, datetime
from pathlib import Path
from typing import cast
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import JsonValue

from app.schemas.extraction import (
    ExtractedDocumentData,
    ExtractionResultStatus,
)
from app.services.extraction.base_extractor import (
    BaseDocumentExtractor,
    ExtractionError,
    ExtractionResourceLimitError,
)
from app.services.extraction.xlsx.xlsx_merge_extractor import read_merged_ranges
from app.services.extraction.xlsx.xlsx_sheet_extractor import (
    XLSXWorkbookCounters,
    extract_xlsx_sheet,
)
from app.services.extraction.xlsx.xlsx_worksheet_metadata import (
    read_worksheet_package_metadata,
)


class XLSXExtractor(BaseDocumentExtractor):
    """Extract worksheet/cell/formula/merge metadata using read-only workbooks."""

    extractor_version = "1.0.0"

    def supports(self, extension: str) -> bool:
        return self.normalize_extension(extension) == "xlsx"

    async def inspect(self, file_path: Path) -> dict[str, JsonValue]:
        file_size = self.validate_source_path(file_path)
        package_metadata, _ = _inspect_xlsx_package(file_path)
        workbook = self._open_workbook(file_path, data_only=False)
        try:
            return {
                "fileSize": file_size,
                "worksheetCount": len(workbook.worksheets),
                "worksheets": [
                    {
                        "name": worksheet.title,
                        "state": worksheet.sheet_state,
                        "maxRow": worksheet.max_row,
                        "maxColumn": worksheet.max_column,
                    }
                    for worksheet in workbook.worksheets
                ],
                "workbookProperties": _workbook_properties(workbook),
                **package_metadata,
            }
        finally:
            workbook.close()

    async def extract(
        self,
        file_path: Path,
        context: dict[str, object],
    ) -> ExtractedDocumentData:
        extraction_context = self.resolve_context(context)
        file_size = self.validate_source_path(file_path, extraction_context)
        package_metadata, package_warnings = _inspect_xlsx_package(file_path)
        workbook = self._open_workbook(file_path, data_only=False)
        cached_workbook = None
        try:
            if len(workbook.worksheets) > extraction_context.xlsx_max_worksheets:
                raise ExtractionResourceLimitError(
                    "XLSX_TOO_MANY_SHEETS",
                    "The workbook exceeds the configured worksheet limit.",
                    details={
                        "maximumWorksheets": (
                            extraction_context.xlsx_max_worksheets
                        ),
                        "actualWorksheets": len(workbook.worksheets),
                    },
                )

            cached_workbook = self._open_workbook(file_path, data_only=True)
            worksheet_paths = {
                str(worksheet.title): str(worksheet._worksheet_path)
                for worksheet in workbook.worksheets
            }
            merged_ranges_by_sheet = read_merged_ranges(
                file_path,
                worksheet_paths,
            )
            package_metadata_by_sheet = read_worksheet_package_metadata(
                file_path,
                worksheet_paths,
            )
            containers = []
            counters = XLSXWorkbookCounters()
            total_sheets = len(workbook.worksheets)
            for sheet_index, worksheet in enumerate(
                workbook.worksheets,
                start=1,
            ):
                await self.checkpoint(
                    extraction_context,
                    10 + int(65 * (sheet_index - 1) / max(1, total_sheets)),
                    f"Reading worksheet {worksheet.title}",
                )
                cached_worksheet = cached_workbook[worksheet.title]
                result = extract_xlsx_sheet(
                    worksheet,
                    cached_worksheet,
                    sheet_index=sheet_index,
                    merged_ranges=merged_ranges_by_sheet.get(
                        str(worksheet.title),
                        (),
                    ),
                    counters=counters,
                    maximum_rows=extraction_context.xlsx_max_rows_per_sheet,
                    maximum_cells=(
                        extraction_context.xlsx_max_cells_per_workbook
                    ),
                    maximum_formulas=extraction_context.xlsx_max_formulas,
                    package_metadata=package_metadata_by_sheet[
                        str(worksheet.title)
                    ],
                )
                containers.append(result.container)
                counters = result.counters

            if counters.cells == 0:
                raise ExtractionError(
                    "XLSX_EMPTY",
                    "The XLSX does not contain any non-empty cells.",
                )

            warnings = list(package_warnings)
            if counters.formulas_without_cached_values:
                warnings.append(
                    f"{counters.formulas_without_cached_values} formula cells "
                    "do not contain cached values; formulas were not executed."
                )
            await self.checkpoint(
                extraction_context,
                75,
                "XLSX extraction completed",
            )
            metadata: dict[str, JsonValue] = {
                "fileSize": file_size,
                "totalWorksheets": total_sheets,
                "totalCells": counters.cells,
                "totalFormulas": counters.formulas,
                "formulasWithoutCachedValues": (
                    counters.formulas_without_cached_values
                ),
                "workbookProperties": _workbook_properties(workbook),
                "definedNames": cast(
                    JsonValue,
                    _defined_names(workbook),
                ),
                "calculationPerformed": False,
                **package_metadata,
            }
            status = (
                ExtractionResultStatus.PARTIALLY_COMPLETED
                if package_warnings
                else ExtractionResultStatus.COMPLETED
            )
            return ExtractedDocumentData(
                extractor_type="XLSX",
                extractor_version=self.extractor_version,
                status=status,
                metadata=metadata,
                containers=containers,
                warnings=warnings,
                requires_ocr=False,
                has_selectable_text=True,
            )
        except ExtractionError:
            raise
        except (AttributeError, KeyError, RuntimeError, TypeError, ValueError) as exc:
            raise ExtractionError(
                "XLSX_EXTRACTION_FAILED",
                "The XLSX content could not be extracted.",
            ) from exc
        finally:
            workbook.close()
            if cached_workbook is not None:
                cached_workbook.close()

    @staticmethod
    def _open_workbook(file_path: Path, *, data_only: bool):
        try:
            return load_workbook(
                file_path,
                read_only=True,
                data_only=data_only,
                keep_links=False,
            )
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
            raise ExtractionError(
                "XLSX_CORRUPT",
                "The XLSX is corrupt or is not a valid XLSX workbook.",
            ) from exc


def _inspect_xlsx_package(
    file_path: Path,
) -> tuple[dict[str, JsonValue], list[str]]:
    try:
        with ZipFile(file_path) as archive:
            names = archive.namelist()
            _reject_unsafe_xml_declarations(archive, names)
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise ExtractionError(
                    "XLSX_EXTRACTION_FAILED",
                    "Macro-enabled spreadsheet content is not supported.",
                )
            external_links = sum(
                name.startswith("xl/externalLinks/")
                and name.endswith(".xml")
                for name in names
            )
            embedded_objects = sum(
                name.startswith(("xl/embeddings/", "xl/activeX/"))
                for name in names
            )
            warnings = []
            if embedded_objects:
                warnings.append(
                    "Embedded spreadsheet objects were not extracted."
                )
            return {
                "externalLinkCount": external_links,
                "externalLinksFollowed": False,
                "embeddedObjectCount": embedded_objects,
            }, warnings
    except ExtractionError:
        raise
    except (BadZipFile, OSError, RuntimeError) as exc:
        raise ExtractionError(
            "XLSX_CORRUPT",
            "The XLSX is corrupt or is not a valid XLSX workbook.",
        ) from exc


def _reject_unsafe_xml_declarations(
    archive: ZipFile,
    names: list[str],
) -> None:
    """Reject OOXML DTD/entity declarations before openpyxl parses XML."""
    forbidden_markers = (b"<!doctype", b"<!entity")
    for name in names:
        if not name.lower().endswith((".xml", ".rels")):
            continue
        overlap = b""
        with archive.open(name) as source:
            while chunk := source.read(64 * 1024):
                search_buffer = (overlap + chunk).lower()
                if any(
                    marker in search_buffer for marker in forbidden_markers
                ):
                    raise ExtractionError(
                        "XLSX_CORRUPT",
                        "The XLSX contains unsafe XML declarations.",
                    )
                overlap = search_buffer[-32:]


def _workbook_properties(workbook) -> dict[str, JsonValue]:
    properties = workbook.properties
    names = (
        "title",
        "subject",
        "creator",
        "keywords",
        "description",
        "lastModifiedBy",
        "category",
        "contentStatus",
        "identifier",
        "language",
        "version",
        "created",
        "modified",
    )
    result: dict[str, JsonValue] = {}
    for name in names:
        value = getattr(properties, name, None)
        if value in (None, ""):
            continue
        if isinstance(value, (datetime, date)):
            result[name] = value.isoformat()
        elif isinstance(value, (str, int, float, bool)):
            result[name] = value
        else:
            result[name] = str(value)
    return result


def _defined_names(workbook) -> list[dict[str, JsonValue]]:
    result: list[dict[str, JsonValue]] = []
    try:
        values = workbook.defined_names.values()
    except AttributeError:
        return result
    for item in values:
        result.append(
            {
                "name": str(item.name),
                "localSheetId": item.localSheetId,
                "hidden": bool(item.hidden),
                "function": bool(item.function),
            }
        )
    return result
