"""Streaming worksheet extraction with per-workbook counters."""

from dataclasses import dataclass
from itertools import zip_longest
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from app.schemas.extraction import (
    ExtractedContainerData,
    ExtractedContainerType,
)
from app.services.extraction.base_extractor import ExtractionResourceLimitError
from app.services.extraction.text_normalizer import (
    count_characters,
    count_words,
    normalize_text,
)
from app.services.extraction.xlsx.xlsx_cell_extractor import extract_xlsx_cell
from app.services.extraction.xlsx.xlsx_merge_extractor import XLSXMergedRange
from app.services.extraction.xlsx.xlsx_worksheet_metadata import (
    XLSXWorksheetPackageMetadata,
)


@dataclass(frozen=True, slots=True)
class _EmptyWorksheetCell:
    coordinate: str
    row: int
    column: int
    value: None = None
    data_type: str = "n"
    number_format: str = "General"
    hyperlink: None = None


@dataclass(frozen=True, slots=True)
class XLSXWorkbookCounters:
    """Cumulative workbook work units used for resource enforcement."""

    cells: int = 0
    formulas: int = 0
    formulas_without_cached_values: int = 0


@dataclass(frozen=True, slots=True)
class ExtractedXLSXSheet:
    """One worksheet container and updated cumulative counters."""

    container: ExtractedContainerData
    counters: XLSXWorkbookCounters


def extract_xlsx_sheet(
    worksheet: Any,
    cached_worksheet: Any,
    *,
    sheet_index: int,
    merged_ranges: tuple[XLSXMergedRange, ...],
    counters: XLSXWorkbookCounters,
    maximum_rows: int,
    maximum_cells: int,
    maximum_formulas: int,
    package_metadata: XLSXWorksheetPackageMetadata,
) -> ExtractedXLSXSheet:
    """Extract non-empty cells in row order without evaluating formulas."""
    reported_dimension = package_metadata.reported_dimension
    _validate_reported_dimension(
        worksheet,
        reported_dimension=reported_dimension,
        maximum_rows=maximum_rows,
        maximum_cells=maximum_cells,
    )
    actual_dimension = _reset_and_calculate_dimensions(worksheet)
    _reset_and_calculate_dimensions(cached_worksheet)
    if worksheet.max_row > maximum_rows:
        raise ExtractionResourceLimitError(
            "XLSX_WORKBOOK_TOO_LARGE",
            "A worksheet exceeds the configured row limit.",
            details={
                "sheet": worksheet.title,
                "maximumRows": maximum_rows,
                "actualRows": worksheet.max_row,
            },
        )
    potential_area = int(worksheet.max_row) * int(worksheet.max_column)
    maximum_safe_area = max(10_000_000, maximum_cells * 20)
    if potential_area > maximum_safe_area:
        raise ExtractionResourceLimitError(
            "XLSX_WORKBOOK_TOO_LARGE",
            "The worksheet used range is too sparse or large to process safely.",
            details={
                "sheet": worksheet.title,
                "maximumSafeArea": maximum_safe_area,
                "actualUsedRangeArea": potential_area,
            },
        )

    merge_by_start = {
        merged_range.start_cell: merged_range
        for merged_range in merged_ranges
    }
    blocks = []
    block_order = 1
    cell_count = counters.cells
    formula_count = counters.formulas
    missing_cached_count = counters.formulas_without_cached_values
    emitted_merge_starts: set[str] = set()
    for row_number, (formula_row, cached_row) in enumerate(
        zip_longest(
        worksheet.iter_rows(),
        cached_worksheet.iter_rows(),
        fillvalue=(),
        ),
        start=1,
    ):
        for column_number, cell in enumerate(formula_row, start=1):
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            merge = merge_by_start.get(coordinate)
            if not hasattr(cell, "coordinate") and merge is None:
                continue
            if not hasattr(cell, "coordinate"):
                cell = _EmptyWorksheetCell(
                    coordinate=coordinate,
                    row=row_number,
                    column=column_number,
                )
            cached_cell = (
                cached_row[column_number - 1]
                if column_number <= len(cached_row)
                and hasattr(cached_row[column_number - 1], "coordinate")
                else None
            )
            row_hidden = row_number in package_metadata.hidden_rows
            column_hidden = package_metadata.is_column_hidden(column_number)
            block = extract_xlsx_cell(
                cell,
                cached_cell,
                sheet_name=str(worksheet.title),
                block_order=block_order,
                merged_range=merge,
                row_hidden=row_hidden,
                column_hidden=column_hidden,
                hyperlink_target=package_metadata.hyperlinks.get(coordinate),
            )
            if block is None:
                continue
            if merge is not None:
                emitted_merge_starts.add(merge.start_cell)

            cell_count += 1
            if cell_count > maximum_cells:
                raise ExtractionResourceLimitError(
                    "XLSX_TOO_MANY_CELLS",
                    "The workbook exceeds the configured non-empty cell limit.",
                    details={
                        "maximumCells": maximum_cells,
                        "actualCells": cell_count,
                    },
                )
            if bool(block.metadata.get("isFormula")):
                formula_count += 1
                if formula_count > maximum_formulas:
                    raise ExtractionResourceLimitError(
                        "XLSX_WORKBOOK_TOO_LARGE",
                        "The workbook exceeds the configured formula limit.",
                        details={
                            "maximumFormulas": maximum_formulas,
                            "actualFormulas": formula_count,
                        },
                    )
                if not bool(block.metadata.get("hasCachedValue")):
                    missing_cached_count += 1
            blocks.append(block)
            block_order += 1

    for merge in merged_ranges:
        if merge.start_cell in emitted_merge_starts:
            continue
        synthetic_cell = _EmptyWorksheetCell(
            coordinate=merge.start_cell,
            row=merge.minimum_row,
            column=merge.minimum_column,
        )
        block = extract_xlsx_cell(
            synthetic_cell,
            None,
            sheet_name=str(worksheet.title),
            block_order=block_order,
            merged_range=merge,
            row_hidden=merge.minimum_row in package_metadata.hidden_rows,
            column_hidden=package_metadata.is_column_hidden(
                merge.minimum_column
            ),
            hyperlink_target=package_metadata.hyperlinks.get(
                merge.start_cell
            ),
        )
        assert block is not None
        cell_count += 1
        if cell_count > maximum_cells:
            raise ExtractionResourceLimitError(
                "XLSX_TOO_MANY_CELLS",
                "The workbook exceeds the configured non-empty cell limit.",
                details={
                    "maximumCells": maximum_cells,
                    "actualCells": cell_count,
                },
            )
        blocks.append(block)
        block_order += 1

    raw_text = "\n".join(block.text for block in blocks if block.text)
    normalised_text = normalize_text(raw_text)
    metadata = {
        "sheetName": str(worksheet.title),
        "sheetState": str(worksheet.sheet_state),
        "maxRow": int(worksheet.max_row),
        "maxColumn": int(worksheet.max_column),
        "reportedUsedRange": reported_dimension,
        "actualUsedRange": actual_dimension,
        "actualCellCount": cell_count - counters.cells,
        "mergedRangeCount": len(merged_ranges),
        "freezePane": package_metadata.freeze_pane,
        "tableNames": list(package_metadata.table_names),
        "sheetProtected": package_metadata.sheet_protected,
    }
    container = ExtractedContainerData(
        container_type=ExtractedContainerType.XLSX_WORKSHEET,
        container_index=sheet_index,
        name=str(worksheet.title),
        title=str(worksheet.title),
        raw_text=raw_text,
        normalised_text=normalised_text,
        character_count=count_characters(normalised_text),
        word_count=count_words(normalised_text),
        metadata=metadata,
        blocks=blocks,
        tables=[],
    )
    return ExtractedXLSXSheet(
        container=container,
        counters=XLSXWorkbookCounters(
            cells=cell_count,
            formulas=formula_count,
            formulas_without_cached_values=missing_cached_count,
        ),
    )


def _reset_and_calculate_dimensions(worksheet: Any) -> str:
    """Work around openpyxl's empty read-only worksheet dimension bug."""
    worksheet.reset_dimensions()
    try:
        return str(worksheet.calculate_dimension(force=True))
    except UnboundLocalError:
        worksheet._max_row = 1
        worksheet._max_column = 1
        return "A1:A1"


def _validate_reported_dimension(
    worksheet: Any,
    *,
    reported_dimension: str | None,
    maximum_rows: int,
    maximum_cells: int,
) -> None:
    if not reported_dimension:
        return
    try:
        cell_range = CellRange(reported_dimension)
    except (TypeError, ValueError):
        return
    if cell_range.max_row > maximum_rows:
        raise ExtractionResourceLimitError(
            "XLSX_WORKBOOK_TOO_LARGE",
            "A worksheet exceeds the configured row limit.",
            details={
                "sheet": worksheet.title,
                "maximumRows": maximum_rows,
                "reportedRows": cell_range.max_row,
            },
        )
    reported_area = (
        (cell_range.max_row - cell_range.min_row + 1)
        * (cell_range.max_col - cell_range.min_col + 1)
    )
    maximum_safe_area = max(10_000_000, maximum_cells * 20)
    if reported_area > maximum_safe_area:
        raise ExtractionResourceLimitError(
            "XLSX_WORKBOOK_TOO_LARGE",
            "The worksheet used range is too sparse or large to process safely.",
            details={
                "sheet": worksheet.title,
                "maximumSafeArea": maximum_safe_area,
                "reportedUsedRangeArea": reported_area,
            },
        )
