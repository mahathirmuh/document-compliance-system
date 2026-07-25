"""Read merged-range declarations without resolving arbitrary XML entities."""

import re
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from app.services.extraction.base_extractor import ExtractionError

_MERGE_CELL_PATTERN = re.compile(
    rb"<mergeCell\b[^>]{0,512}?\bref\s*=\s*([\"'])([^\"']{1,100})\1",
    re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class XLSXMergedRange:
    """A bounded worksheet merged range."""

    reference: str
    start_cell: str
    end_cell: str
    minimum_row: int
    maximum_row: int
    minimum_column: int
    maximum_column: int

    @property
    def row_span(self) -> int:
        return self.maximum_row - self.minimum_row + 1

    @property
    def column_span(self) -> int:
        return self.maximum_column - self.minimum_column + 1


def read_merged_ranges(
    file_path: Path,
    worksheet_paths: dict[str, str],
) -> dict[str, tuple[XLSXMergedRange, ...]]:
    """Stream only merge-cell tags from worksheet XML ZIP entries."""
    result: dict[str, tuple[XLSXMergedRange, ...]] = {}
    try:
        with ZipFile(file_path) as archive:
            archive_names = set(archive.namelist())
            for sheet_name, raw_path in worksheet_paths.items():
                archive_path = raw_path.lstrip("/")
                if archive_path not in archive_names:
                    result[sheet_name] = ()
                    continue
                references = _stream_merge_references(archive, archive_path)
                ranges: list[XLSXMergedRange] = []
                for reference in references:
                    try:
                        cell_range = CellRange(reference)
                    except (TypeError, ValueError):
                        continue
                    ranges.append(
                        XLSXMergedRange(
                            reference=str(cell_range),
                            start_cell=(
                                f"{get_column_letter(cell_range.min_col)}"
                                f"{cell_range.min_row}"
                            ),
                            end_cell=(
                                f"{get_column_letter(cell_range.max_col)}"
                                f"{cell_range.max_row}"
                            ),
                            minimum_row=cell_range.min_row,
                            maximum_row=cell_range.max_row,
                            minimum_column=cell_range.min_col,
                            maximum_column=cell_range.max_col,
                        )
                    )
                result[sheet_name] = tuple(ranges)
    except (BadZipFile, KeyError, OSError) as exc:
        raise ExtractionError(
            "XLSX_CORRUPT",
            "The XLSX is corrupt or is not a valid XLSX workbook.",
        ) from exc
    return result


def _stream_merge_references(archive: ZipFile, path: str) -> list[str]:
    references: list[str] = []
    seen: set[str] = set()
    overlap = b""
    with archive.open(path) as source:
        while chunk := source.read(64 * 1024):
            search_buffer = overlap + chunk
            for match in _MERGE_CELL_PATTERN.finditer(search_buffer):
                try:
                    reference = match.group(2).decode("ascii").upper()
                except UnicodeDecodeError:
                    continue
                if reference not in seen:
                    seen.add(reference)
                    references.append(reference)
            overlap = search_buffer[-1024:]
    return references
