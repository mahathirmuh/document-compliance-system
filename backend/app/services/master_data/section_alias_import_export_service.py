"""Safe XLSX import/export for section definitions and aliases."""

from __future__ import annotations

from io import BytesIO
from zipfile import is_zipfile

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.core.config import Settings, get_settings
from app.models.section_alias import SectionAlias
from app.models.section_definition import SectionDefinition
from app.models.user import User
from app.repositories.audit_log import AuditLogRepository
from app.repositories.section_alias_profile_repository import (
    SectionAliasProfileRepository,
)
from app.repositories.section_alias_repository import SectionAliasRepository
from app.repositories.section_definition_repository import (
    SectionDefinitionRepository,
)
from app.schemas.master_data import ImportMode
from app.schemas.section_detection import (
    SectionAliasImportError,
    SectionAliasImportPreview,
    SectionAliasImportResult,
    SectionAliasImportRow,
    SectionDefinitionImportRow,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.documents.file_signature_service import (
    FileSignatureError,
    validate_office_open_xml,
)
from app.services.master_data.base import (
    audit_dump,
    business_error,
    not_found,
)
from app.services.master_data.section_alias_service import (
    normalise_alias,
    validate_safe_regex,
)
from app.services.master_data.section_definition_service import (
    _ensure_can_configure,
)

DEFINITION_HEADERS = (
    "canonical_code",
    "display_name",
    "description",
    "display_order",
    "is_required_default",
    "is_repeatable",
    "is_active",
)
ALIAS_HEADERS = (
    "canonical_code",
    "language_code",
    "alias_text",
    "match_type",
    "priority",
    "is_regex",
    "is_active",
)
_SHEETS = ("Section Definitions", "Section Aliases")


def _safe_export_value(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _parse_bool(value: object, *, default: bool) -> object:
    if value is None or value == "":
        return default
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "yes", "1"}:
            return True
        if normalized in {"false", "no", "0"}:
            return False
    return value


class SectionAliasImportExportService:
    """Preview-before-confirm import and formula-safe profile export."""

    def __init__(
        self,
        session: AsyncSession,
        user: User,
        metadata: RequestMetadata,
        settings: Settings | None = None,
    ) -> None:
        self.session = session
        self.user = user
        self.metadata = metadata
        self.settings = settings or get_settings()
        self.profiles = SectionAliasProfileRepository(session)
        self.definitions = SectionDefinitionRepository(session)
        self.aliases = SectionAliasRepository(session)
        self.audit_logs = AuditLogRepository(session)

    async def export_profile(self, profile_id) -> bytes:
        profile = await self.profiles.get_by_id(profile_id)
        if profile is None:
            raise not_found("Section Alias Profile")
        definitions = await self.definitions.list_for_profile(profile_id)
        workbook = Workbook()
        definition_sheet = workbook.active
        definition_sheet.title = _SHEETS[0]
        definition_sheet.append(DEFINITION_HEADERS)
        alias_sheet = workbook.create_sheet(_SHEETS[1])
        alias_sheet.append(ALIAS_HEADERS)
        for definition in definitions:
            definition_sheet.append(
                tuple(
                    _safe_export_value(value)
                    for value in (
                        definition.canonical_code,
                        definition.display_name,
                        definition.description,
                        definition.display_order,
                        definition.is_required_default,
                        definition.is_repeatable,
                        definition.is_active,
                    )
                )
            )
            for alias in definition.aliases:
                alias_sheet.append(
                    tuple(
                        _safe_export_value(value)
                        for value in (
                            definition.canonical_code,
                            alias.language_code.value,
                            alias.alias_text,
                            alias.match_type.value,
                            alias.priority,
                            alias.is_regex,
                            alias.is_active,
                        )
                    )
                )
        for sheet in (definition_sheet, alias_sheet):
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        output = BytesIO()
        workbook.save(output)
        await self.audit_logs.create(
            user_id=self.user.id,
            action=AuditAction.EXPORT_SECTION_ALIASES,
            entity_type="section_alias_profile",
            entity_id=profile.id,
            description=f"Section aliases for {profile.code} were exported.",
            old_values=None,
            new_values={
                "definitionCount": len(definitions),
                "aliasCount": sum(
                    len(definition.aliases)
                    for definition in definitions
                ),
            },
            ip_address=self.metadata.ip_address,
            user_agent=self.metadata.user_agent,
        )
        await self.session.commit()
        return output.getvalue()

    async def preview_import(
        self,
        profile_id,
        content: bytes,
    ) -> SectionAliasImportPreview:
        _ensure_can_configure(self.user)
        profile = await self.profiles.get_by_id(profile_id)
        if profile is None:
            raise not_found("Section Alias Profile")
        if not profile.is_active:
            raise business_error(
                "Section alias profile must be active.",
                field="profileId",
            )
        if not content or not is_zipfile(BytesIO(content)):
            raise business_error(
                "Section alias import must be a valid XLSX workbook.",
                field="file",
            )
        try:
            validate_office_open_xml(
                BytesIO(content),
                "xlsx",
                max_uncompressed_size=(
                    self.settings.ooxml_max_uncompressed_size_bytes
                ),
            )
        except FileSignatureError as exc:
            raise business_error(
                "Section alias import must be a valid XLSX workbook.",
                field="file",
            ) from exc
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            # OpenPyXL exposes parser-specific exception types depending on
            # the optional XML backend. Keep all parser failures client-safe.
            raise business_error(
                "Section alias import must be a valid XLSX workbook.",
                field="file",
            ) from exc
        try:
            missing = [name for name in _SHEETS if name not in workbook]
            if missing:
                raise business_error(
                    f"Missing required sheets: {', '.join(missing)}.",
                    field="file",
                )
            definitions, definition_errors = self._parse_definitions(
                workbook[_SHEETS[0]],
                max_rows=self.settings.master_data_import_max_rows,
            )
            definition_count = len(definitions) + len(definition_errors)
            aliases, alias_errors = self._parse_aliases(
                workbook[_SHEETS[1]],
                max_rows=(
                    self.settings.master_data_import_max_rows
                    - definition_count
                ),
            )
        finally:
            workbook.close()
        errors = [*definition_errors, *alias_errors]
        return SectionAliasImportPreview(
            profile_id=profile_id,
            definitions=definitions,
            aliases=aliases,
            errors=errors,
            valid=not errors,
        )

    def _headers(self, sheet, expected) -> list[str]:
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(value or "").strip() for value in first]
        if tuple(headers) != expected:
            raise business_error(
                f"{sheet.title} headers must be: {', '.join(expected)}.",
                field="file",
            )
        return headers

    def _parse_definitions(
        self,
        sheet,
        *,
        max_rows: int,
    ) -> tuple[
        list[SectionDefinitionImportRow],
        list[SectionAliasImportError],
    ]:
        headers = self._headers(sheet, DEFINITION_HEADERS)
        rows: list[SectionDefinitionImportRow] = []
        errors: list[SectionAliasImportError] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            if len(rows) + len(errors) >= max_rows:
                raise business_error(
                    "Workbook exceeds the configured import row limit of "
                    f"{self.settings.master_data_import_max_rows}.",
                    field="file",
                )
            raw = dict(zip(headers, values, strict=True))
            if any(
                isinstance(value, str) and value.startswith("=")
                for value in raw.values()
            ):
                errors.append(
                    SectionAliasImportError(
                        sheet=sheet.title,
                        row_number=row_number,
                        message="Formula cells are not allowed.",
                    )
                )
                continue
            raw["row_number"] = row_number
            raw["display_order"] = raw["display_order"] or 0
            for key, default in (
                ("is_required_default", False),
                ("is_repeatable", False),
                ("is_active", True),
            ):
                raw[key] = _parse_bool(raw[key], default=default)
            try:
                rows.append(SectionDefinitionImportRow.model_validate(raw))
            except ValidationError as exc:
                errors.append(
                    SectionAliasImportError(
                        sheet=sheet.title,
                        row_number=row_number,
                        message=str(exc.errors()[0]["msg"]),
                    )
                )
        return rows, errors

    def _parse_aliases(
        self,
        sheet,
        *,
        max_rows: int,
    ) -> tuple[list[SectionAliasImportRow], list[SectionAliasImportError]]:
        headers = self._headers(sheet, ALIAS_HEADERS)
        rows: list[SectionAliasImportRow] = []
        errors: list[SectionAliasImportError] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value not in (None, "") for value in values):
                continue
            if len(rows) + len(errors) >= max_rows:
                raise business_error(
                    "Workbook exceeds the configured import row limit of "
                    f"{self.settings.master_data_import_max_rows}.",
                    field="file",
                )
            raw = dict(zip(headers, values, strict=True))
            if any(
                isinstance(value, str) and value.startswith("=")
                for value in raw.values()
            ):
                errors.append(
                    SectionAliasImportError(
                        sheet=sheet.title,
                        row_number=row_number,
                        message="Formula cells are not allowed.",
                    )
                )
                continue
            raw["row_number"] = row_number
            raw["priority"] = raw["priority"] or 0
            raw["is_regex"] = _parse_bool(
                raw["is_regex"],
                default=False,
            )
            raw["is_active"] = _parse_bool(
                raw["is_active"],
                default=True,
            )
            try:
                parsed = SectionAliasImportRow.model_validate(raw)
                if parsed.is_regex:
                    validate_safe_regex(parsed.alias_text, self.settings)
                rows.append(parsed)
            except (ValidationError, ValueError) as exc:
                message = (
                    str(exc.errors()[0]["msg"])
                    if isinstance(exc, ValidationError)
                    else str(exc)
                )
                errors.append(
                    SectionAliasImportError(
                        sheet=sheet.title,
                        row_number=row_number,
                        message=message,
                    )
                )
        return rows, errors

    async def confirm_import(
        self,
        preview: SectionAliasImportPreview,
        *,
        mode: ImportMode = ImportMode.UPSERT,
    ) -> SectionAliasImportResult:
        _ensure_can_configure(self.user)
        if not preview.valid or preview.errors:
            raise business_error(
                "Import preview contains validation errors.",
                field="file",
            )
        profile = await self.profiles.get_by_id(
            preview.profile_id,
            for_update=True,
        )
        if profile is None:
            raise not_found("Section Alias Profile")
        if not profile.is_active:
            raise business_error(
                "Section alias profile must be active.",
                field="profileId",
            )
        result = SectionAliasImportResult(profile_id=profile.id)
        definitions_by_code = {
            item.canonical_code: item
            for item in await self.definitions.list_for_profile(profile.id)
        }
        seen_definition_codes: set[str] = set()
        for row in preview.definitions:
            if row.canonical_code in seen_definition_codes:
                result.skipped += 1
                continue
            seen_definition_codes.add(row.canonical_code)
            entity = definitions_by_code.get(row.canonical_code)
            values = row.model_dump(
                exclude={"row_number"},
                by_alias=False,
            )
            if entity is None:
                entity = SectionDefinition(
                    profile_id=profile.id,
                    **values,
                    created_by=self.user.id,
                    updated_by=self.user.id,
                )
                await self.definitions.add(entity)
                definitions_by_code[entity.canonical_code] = entity
                result.definitions_created += 1
            elif mode is ImportMode.CREATE_ONLY:
                result.skipped += 1
            else:
                for key, value in values.items():
                    setattr(entity, key, value)
                entity.updated_by = self.user.id
                result.definitions_updated += 1
        seen_alias_keys: set[tuple[object, object, str]] = set()
        for alias_row in preview.aliases:
            definition = definitions_by_code.get(alias_row.canonical_code)
            if definition is None:
                raise business_error(
                    f"Unknown canonical section {alias_row.canonical_code}.",
                    field="canonicalCode",
                )
            normalized = normalise_alias(
                alias_row.alias_text,
                alias_row.match_type,
            )
            alias_key = (definition.id, alias_row.language_code, normalized)
            if alias_key in seen_alias_keys:
                result.skipped += 1
                continue
            seen_alias_keys.add(alias_key)
            alias_entity = await self.aliases.get_duplicate(
                section_definition_id=definition.id,
                language_code=alias_row.language_code,
                normalised_alias=normalized,
            )
            values = alias_row.model_dump(
                exclude={"row_number", "canonical_code"},
                by_alias=False,
            )
            if alias_entity is None:
                alias_entity = SectionAlias(
                    section_definition_id=definition.id,
                    normalised_alias=normalized,
                    **values,
                    created_by=self.user.id,
                    updated_by=self.user.id,
                )
                await self.aliases.add(alias_entity)
                result.aliases_created += 1
            elif mode is ImportMode.CREATE_ONLY:
                result.skipped += 1
            else:
                for key, value in values.items():
                    setattr(alias_entity, key, value)
                alias_entity.normalised_alias = normalized
                alias_entity.updated_by = self.user.id
                result.aliases_updated += 1
        try:
            await self.session.flush()
            await self.audit_logs.create(
                user_id=self.user.id,
                action=AuditAction.IMPORT_SECTION_ALIASES,
                entity_type="section_alias_profile",
                entity_id=profile.id,
                description=(
                    f"Section aliases for {profile.code} were imported "
                    f"using {mode.value} mode."
                ),
                old_values=None,
                new_values={
                    **audit_dump(result),
                    "mode": mode.value,
                },
                ip_address=self.metadata.ip_address,
                user_agent=self.metadata.user_agent,
            )
            await self.session.commit()
        except IntegrityError as exc:
            await self.session.rollback()
            raise business_error(
                "Import conflicts with existing section alias data.",
                field="file",
            ) from exc
        return result
