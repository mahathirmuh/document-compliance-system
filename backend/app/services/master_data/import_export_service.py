"""Safe XLSX template, preview, confirm, and export workflows."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ValidationError
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.core.config import Settings
from app.core.exceptions import ApplicationError
from app.models.department import Department
from app.models.document_status import DocumentStatus
from app.models.document_type import DocumentType
from app.models.section import Section
from app.models.user import User
from app.models.validation_rule import ValidationRule
from app.repositories.audit_log import AuditLogRepository
from app.repositories.department_repository import DepartmentRepository
from app.repositories.document_status_repository import (
    DocumentStatusRepository,
)
from app.repositories.document_type_repository import DocumentTypeRepository
from app.repositories.section_repository import SectionRepository
from app.repositories.validation_rule_repository import ValidationRuleRepository
from app.schemas.common import ErrorDetail
from app.schemas.department import DepartmentCreate
from app.schemas.document_status import DocumentStatusCreate
from app.schemas.document_type import DocumentTypeCreate
from app.schemas.master_data import (
    ImportConfirmResponse,
    ImportEntityType,
    ImportMode,
    ImportPreviewResponse,
    ImportPreviewRow,
    ImportRowStatus,
)
from app.schemas.section import SectionCreate
from app.schemas.validation_rule import ValidationRuleCreate
from app.services.auth.auth_service import RequestMetadata
from app.utils.datetime import utc_now

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

TEMPLATE_HEADERS: dict[ImportEntityType, tuple[str, ...]] = {
    ImportEntityType.DEPARTMENTS: (
        "code",
        "name",
        "description",
        "is_active",
    ),
    ImportEntityType.SECTIONS: (
        "department_code",
        "code",
        "name",
        "description",
        "is_active",
    ),
    ImportEntityType.DOCUMENT_TYPES: (
        "code",
        "name",
        "category",
        "description",
        "requires_section",
        "is_active",
    ),
    ImportEntityType.DOCUMENT_STATUSES: (
        "code",
        "name",
        "description",
        "display_order",
        "is_initial",
        "is_final",
        "is_obsolete",
        "is_active",
    ),
    ImportEntityType.VALIDATION_RULES: (
        "code",
        "name",
        "document_type_code",
        "required_indonesian",
        "required_english",
        "required_chinese",
        "minimum_indonesian_coverage",
        "minimum_english_coverage",
        "minimum_chinese_coverage",
        "validate_language_order",
        "language_order",
        "validate_sections",
        "required_sections",
        "validate_tables",
        "minimum_compliance_score",
        "partial_compliance_score",
        "is_default",
        "is_active",
    ),
}


@dataclass(slots=True)
class ParsedImportRow:
    row_number: int
    raw: dict[str, Any]
    payload: BaseModel | None
    existing: Any | None
    key: str | tuple[str, str] | None
    status: ImportRowStatus
    errors: list[str]
    context: dict[str, Any]

    def preview_data(self) -> dict[str, Any]:
        if self.payload is None:
            return {_camelize(key): _json_safe(value) for key, value in self.raw.items()}
        data = self.payload.model_dump(mode="json", by_alias=True)
        if "department_code" in self.raw:
            data["departmentCode"] = self.raw["department_code"]
        if "document_type_code" in self.raw:
            data["documentTypeCode"] = self.raw["document_type_code"]
        return data


def _camelize(value: str) -> str:
    head, *tail = value.split("_")
    return head + "".join(part.capitalize() for part in tail)


def _json_safe(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _excel_safe(value: Any) -> Any:
    """Prevent spreadsheet formula execution for user-controlled text."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _boolean(value: Any, *, default: bool | None = None) -> bool:
    if value is None or value == "":
        if default is not None:
            return default
        raise ValueError("A boolean value is required.")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1"}:
        return True
    if normalized in {"false", "no", "n", "0"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}.")


def _integer(value: Any, *, default: int | None = None) -> int:
    if value is None or value == "":
        if default is not None:
            return default
        raise ValueError("An integer value is required.")
    if isinstance(value, bool):
        raise ValueError("Boolean is not a valid integer.")
    numeric = int(value)
    if isinstance(value, float) and value != numeric:
        raise ValueError("Integer value must not contain decimals.")
    return numeric


def _string(value: Any, *, optional: bool = False) -> str | None:
    if value is None:
        return None if optional else ""
    normalized = str(value).strip()
    return normalized or (None if optional else "")


def _list_value(value: Any, *, default: list[str]) -> list[str]:
    if value is None or value == "":
        return list(default)
    if isinstance(value, str):
        return [part.strip() for part in value.split(",") if part.strip()]
    if isinstance(value, (list, tuple)):
        return [str(part).strip() for part in value if str(part).strip()]
    raise ValueError("List value must be comma-separated text.")


def _validation_messages(exc: ValidationError) -> list[str]:
    messages: list[str] = []
    for error in exc.errors():
        field = ".".join(str(item) for item in error["loc"])
        messages.append(f"{field}: {error['msg']}")
    return messages


def _file_error(message: str, *, field: str = "file") -> ApplicationError:
    return ApplicationError(
        "Master data import failed.",
        status_code=400,
        errors=[ErrorDetail(field=field, message=message)],
    )


class MasterDataImportExportService:
    """Validate every uploaded workbook and own import/export transactions."""

    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        user: User,
        metadata: RequestMetadata,
    ) -> None:
        self.session = session
        self.settings = settings
        self.user = user
        self.metadata = metadata
        self.audit_logs = AuditLogRepository(session)
        self.departments = DepartmentRepository(session)
        self.sections = SectionRepository(session)
        self.document_types = DocumentTypeRepository(session)
        self.document_statuses = DocumentStatusRepository(session)
        self.validation_rules = ValidationRuleRepository(session)

    @staticmethod
    def template(entity_type: ImportEntityType) -> tuple[bytes, str]:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = entity_type.value.replace("-", " ").title()[:31]
        headers = TEMPLATE_HEADERS[entity_type]
        sheet.append(headers)
        MasterDataImportExportService._format_sheet(sheet, len(headers))
        output = BytesIO()
        workbook.save(output)
        return (
            output.getvalue(),
            f"{entity_type.value}_template.xlsx",
        )

    async def preview(
        self,
        entity_type: ImportEntityType,
        *,
        filename: str | None,
        content: bytes,
    ) -> ImportPreviewResponse:
        parsed = await self._analyze(
            entity_type,
            filename,
            content,
            for_update=False,
        )
        return self._preview_response(entity_type, parsed)

    async def confirm(
        self,
        entity_type: ImportEntityType,
        *,
        mode: ImportMode,
        filename: str | None,
        content: bytes,
    ) -> ImportConfirmResponse:
        # The file is deliberately parsed and validated again. No client-side
        # preview payload is trusted as an import source.
        rows = await self._analyze(
            entity_type,
            filename,
            content,
            for_update=True,
        )
        created = 0
        updated = 0
        skipped = 0
        failed = 0
        changed_codes: list[str] = []

        for row in rows:
            if row.status is ImportRowStatus.INVALID:
                failed += 1
                continue
            if row.status is ImportRowStatus.DUPLICATE and row.context.get(
                "inFileDuplicate"
            ):
                skipped += 1
                continue
            if row.existing is not None and mode is ImportMode.CREATE_ONLY:
                skipped += 1
                continue
            if row.payload is None:
                failed += 1
                continue
            if row.existing is None:
                self._create_entity(entity_type, row)
                created += 1
            else:
                await self._update_entity(entity_type, row)
                updated += 1
            code = getattr(row.payload, "code", None)
            if code is not None and len(changed_codes) < 100:
                changed_codes.append(str(code))

        result = ImportConfirmResponse(
            entity_type=entity_type,
            mode=mode,
            total_rows=len(rows),
            created=created,
            updated=updated,
            skipped=skipped,
            failed=failed,
        )
        try:
            await self.session.flush()
            await self.audit_logs.create(
                user_id=self.user.id,
                action=AuditAction.IMPORT_MASTER_DATA,
                entity_type="master_data_import",
                entity_id=None,
                description=(
                    f"Imported {entity_type.value} using mode {mode.value}."
                ),
                old_values=None,
                new_values={
                    **result.model_dump(mode="json", by_alias=True),
                    "changedCodes": changed_codes,
                },
                ip_address=self.metadata.ip_address,
                user_agent=self.metadata.user_agent,
            )
            await self.session.commit()
        except IntegrityError as exc:
            await self.session.rollback()
            raise ApplicationError(
                "Master data import could not be committed.",
                status_code=409,
                errors=[
                    ErrorDetail(
                        field=None,
                        message=(
                            "Imported rows conflict with data changed by "
                            "another request. Preview the workbook again."
                        ),
                    )
                ],
            ) from exc
        return result

    async def export(
        self,
        entity_type: ImportEntityType,
        *,
        search: str | None,
        is_active: bool | None,
        department_id: Any | None,
        document_type_id: Any | None,
        category: str | None,
    ) -> tuple[bytes, str]:
        max_rows = self.settings.master_data_export_max_rows
        items, total = await self._export_items(
            entity_type,
            max_rows=max_rows,
            search=search,
            is_active=is_active,
            department_id=department_id,
            document_type_id=document_type_id,
            category=category,
        )
        headers = TEMPLATE_HEADERS[entity_type]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = entity_type.value.replace("-", " ").title()[:31]
        sheet.append(headers)
        for item in items:
            values = self._export_row(entity_type, item)
            sheet.append(
                [_excel_safe(values.get(header)) for header in headers]
            )
        self._format_sheet(sheet, len(headers))
        metadata_sheet = workbook.create_sheet("_metadata")
        metadata_sheet.sheet_state = "hidden"
        metadata_sheet.append(("exported_at", utc_now().isoformat()))
        metadata_sheet.append(("entity_type", entity_type.value))
        metadata_sheet.append(("available_rows", total))
        metadata_sheet.append(("exported_rows", len(items)))
        output = BytesIO()
        workbook.save(output)

        await self.audit_logs.create(
            user_id=self.user.id,
            action=AuditAction.EXPORT_MASTER_DATA,
            entity_type="master_data_export",
            entity_id=None,
            description=f"Exported {entity_type.value} to XLSX.",
            old_values=None,
            new_values={
                "entityType": entity_type.value,
                "availableRows": total,
                "exportedRows": len(items),
                "truncated": total > len(items),
                "filters": {
                    "search": search,
                    "isActive": is_active,
                    "departmentId": (
                        str(department_id)
                        if department_id is not None
                        else None
                    ),
                    "documentTypeId": (
                        str(document_type_id)
                        if document_type_id is not None
                        else None
                    ),
                    "category": category,
                },
            },
            ip_address=self.metadata.ip_address,
            user_agent=self.metadata.user_agent,
        )
        await self.session.commit()
        filename = f"{entity_type.value}_{utc_now().date().isoformat()}.xlsx"
        return output.getvalue(), filename

    async def _analyze(
        self,
        entity_type: ImportEntityType,
        filename: str | None,
        content: bytes,
        *,
        for_update: bool,
    ) -> list[ParsedImportRow]:
        raw_rows = self._read_workbook(entity_type, filename, content)
        seen: set[str | tuple[str, str]] = set()
        parsed: list[ParsedImportRow] = []
        initial_claimed: str | None = None
        default_claimed: set[str] = set()

        for row_number, raw in raw_rows:
            row = await self._parse_row(
                entity_type,
                row_number,
                raw,
                for_update=for_update,
            )
            if row.status is not ImportRowStatus.INVALID and row.key is not None:
                if row.key in seen:
                    row.status = ImportRowStatus.DUPLICATE
                    row.errors.append("Duplicate row in uploaded workbook.")
                    row.context["inFileDuplicate"] = True
                else:
                    seen.add(row.key)

            if (
                entity_type is ImportEntityType.DOCUMENT_STATUSES
                and row.payload is not None
                and getattr(row.payload, "is_initial", False)
            ):
                code = str(getattr(row.payload, "code"))
                if initial_claimed is not None and initial_claimed != code:
                    row.status = ImportRowStatus.INVALID
                    row.errors.append(
                        "Only one imported document status may be initial."
                    )
                else:
                    initial_claimed = code

            if (
                entity_type is ImportEntityType.VALIDATION_RULES
                and row.payload is not None
                and getattr(row.payload, "is_default", False)
            ):
                scope = str(
                    getattr(row.payload, "document_type_id", None) or "global"
                )
                if scope in default_claimed:
                    row.status = ImportRowStatus.INVALID
                    row.errors.append(
                        "Only one imported default rule is allowed per scope."
                    )
                else:
                    default_claimed.add(scope)
            parsed.append(row)
        return parsed

    def _read_workbook(
        self,
        entity_type: ImportEntityType,
        filename: str | None,
        content: bytes,
    ) -> list[tuple[int, dict[str, Any]]]:
        if not filename or Path(filename).suffix.lower() != ".xlsx":
            raise _file_error("Only .xlsx files are accepted.")
        if (
            not content
            or not content.startswith(b"PK")
            or not is_zipfile(BytesIO(content))
        ):
            raise _file_error("The uploaded file is not a valid XLSX workbook.")
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise _file_error(
                "The uploaded XLSX workbook could not be read."
            ) from exc
        sheet = workbook.active
        expected = list(TEMPLATE_HEADERS[entity_type])
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        actual = [
            str(value).strip() if value is not None else ""
            for value in header_values
        ]
        if actual != expected:
            workbook.close()
            raise _file_error(
                "Invalid header. Expected exactly: " + ", ".join(expected)
            )
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            if len(rows) >= self.settings.master_data_import_max_rows:
                raise _file_error(
                    "Workbook exceeds the configured import row limit of "
                    f"{self.settings.master_data_import_max_rows}."
                )
            rows.append(
                (
                    row_number,
                    dict(zip(expected, values, strict=False)),
                )
            )
        workbook.close()
        return rows

    async def _parse_row(
        self,
        entity_type: ImportEntityType,
        row_number: int,
        raw: dict[str, Any],
        *,
        for_update: bool,
    ) -> ParsedImportRow:
        try:
            payload, context = await self._build_payload(
                entity_type,
                raw,
                for_update=for_update,
            )
        except (ValueError, ValidationError) as exc:
            errors = (
                _validation_messages(exc)
                if isinstance(exc, ValidationError)
                else [str(exc)]
            )
            return ParsedImportRow(
                row_number=row_number,
                raw=raw,
                payload=None,
                existing=None,
                key=None,
                status=ImportRowStatus.INVALID,
                errors=errors,
                context={},
            )

        key = self._key(entity_type, payload, context)
        existing = await self._existing(
            entity_type,
            payload,
            context,
            for_update=for_update,
        )
        errors: list[str] = []
        status = ImportRowStatus.VALID
        if existing is not None:
            if getattr(existing, "deleted_at", None) is not None:
                status = ImportRowStatus.INVALID
                errors.append(
                    "A soft-deleted record already uses this unique code."
                )
            else:
                status = ImportRowStatus.DUPLICATE
                errors.append("Record already exists in the database.")

        business_errors = await self._row_business_errors(
            entity_type,
            payload,
            existing,
            for_update=for_update,
        )
        if business_errors:
            status = ImportRowStatus.INVALID
            errors.extend(business_errors)
        return ParsedImportRow(
            row_number=row_number,
            raw=raw,
            payload=payload,
            existing=existing,
            key=key,
            status=status,
            errors=errors,
            context=context,
        )

    async def _build_payload(
        self,
        entity_type: ImportEntityType,
        raw: dict[str, Any],
        *,
        for_update: bool,
    ) -> tuple[BaseModel, dict[str, Any]]:
        if entity_type is ImportEntityType.DEPARTMENTS:
            return (
                DepartmentCreate(
                    code=_string(raw["code"]),
                    name=_string(raw["name"]),
                    description=_string(raw["description"], optional=True),
                    is_active=_boolean(raw["is_active"], default=True),
                ),
                {},
            )
        if entity_type is ImportEntityType.SECTIONS:
            department_code = str(_string(raw["department_code"])).upper()
            department = await self.departments.get_by_code(
                department_code,
                for_update=for_update,
            )
            if department is None:
                raise ValueError("department_code: Department was not found.")
            if not department.is_active:
                raise ValueError("department_code: Department must be active.")
            return (
                SectionCreate(
                    department_id=department.id,
                    code=_string(raw["code"]),
                    name=_string(raw["name"]),
                    description=_string(raw["description"], optional=True),
                    is_active=_boolean(raw["is_active"], default=True),
                ),
                {
                    "departmentCode": department_code,
                    "department": department,
                },
            )
        if entity_type is ImportEntityType.DOCUMENT_TYPES:
            return (
                DocumentTypeCreate(
                    code=_string(raw["code"]),
                    name=_string(raw["name"]),
                    category=_string(raw["category"], optional=True),
                    description=_string(raw["description"], optional=True),
                    requires_section=_boolean(
                        raw["requires_section"],
                        default=True,
                    ),
                    is_active=_boolean(raw["is_active"], default=True),
                ),
                {},
            )
        if entity_type is ImportEntityType.DOCUMENT_STATUSES:
            return (
                DocumentStatusCreate(
                    code=_string(raw["code"]),
                    name=_string(raw["name"]),
                    description=_string(raw["description"], optional=True),
                    display_order=_integer(raw["display_order"], default=0),
                    is_initial=_boolean(raw["is_initial"], default=False),
                    is_final=_boolean(raw["is_final"], default=False),
                    is_obsolete=_boolean(raw["is_obsolete"], default=False),
                    is_active=_boolean(raw["is_active"], default=True),
                ),
                {},
            )

        document_type_code = _string(
            raw["document_type_code"],
            optional=True,
        )
        document_type = None
        if document_type_code:
            document_type = await self.document_types.get_by_code(
                document_type_code,
                for_update=for_update,
            )
            if document_type is None:
                raise ValueError(
                    "document_type_code: Document type was not found."
                )
        return (
            ValidationRuleCreate(
                code=_string(raw["code"]),
                name=_string(raw["name"]),
                document_type_id=(
                    document_type.id if document_type is not None else None
                ),
                required_indonesian=_boolean(
                    raw["required_indonesian"],
                    default=True,
                ),
                required_english=_boolean(
                    raw["required_english"],
                    default=True,
                ),
                required_chinese=_boolean(
                    raw["required_chinese"],
                    default=True,
                ),
                minimum_indonesian_coverage=_integer(
                    raw["minimum_indonesian_coverage"],
                    default=95,
                ),
                minimum_english_coverage=_integer(
                    raw["minimum_english_coverage"],
                    default=95,
                ),
                minimum_chinese_coverage=_integer(
                    raw["minimum_chinese_coverage"],
                    default=95,
                ),
                validate_language_order=_boolean(
                    raw["validate_language_order"],
                    default=True,
                ),
                language_order=_list_value(
                    raw["language_order"],
                    default=["id", "en", "zh"],
                ),
                validate_sections=_boolean(
                    raw["validate_sections"],
                    default=False,
                ),
                required_sections=_list_value(
                    raw["required_sections"],
                    default=[
                        "TITLE",
                        "PURPOSE",
                        "SCOPE",
                        "RESPONSIBILITY",
                        "PROCEDURE",
                        "RECORDS",
                        "REFERENCE",
                    ],
                ),
                validate_tables=_boolean(
                    raw["validate_tables"],
                    default=False,
                ),
                minimum_compliance_score=_integer(
                    raw["minimum_compliance_score"],
                    default=95,
                ),
                partial_compliance_score=_integer(
                    raw["partial_compliance_score"],
                    default=70,
                ),
                is_default=_boolean(raw["is_default"], default=False),
                is_active=_boolean(raw["is_active"], default=True),
            ),
            {
                "documentTypeCode": (
                    str(document_type_code).upper()
                    if document_type_code
                    else None
                ),
                "documentType": document_type,
            },
        )

    @staticmethod
    def _key(
        entity_type: ImportEntityType,
        payload: BaseModel,
        context: dict[str, Any],
    ) -> str | tuple[str, str]:
        code = str(getattr(payload, "code"))
        if entity_type is ImportEntityType.SECTIONS:
            return str(context["departmentCode"]), code
        return code

    async def _existing(
        self,
        entity_type: ImportEntityType,
        payload: BaseModel,
        context: dict[str, Any],
        *,
        for_update: bool,
    ) -> Any | None:
        code = str(getattr(payload, "code"))
        if entity_type is ImportEntityType.SECTIONS:
            statement = select(Section).where(
                Section.department_id == getattr(payload, "department_id"),
                Section.code == code,
            )
            if for_update:
                statement = statement.with_for_update()
            return await self.session.scalar(statement)
        model = {
            ImportEntityType.DEPARTMENTS: Department,
            ImportEntityType.DOCUMENT_TYPES: DocumentType,
            ImportEntityType.DOCUMENT_STATUSES: DocumentStatus,
            ImportEntityType.VALIDATION_RULES: ValidationRule,
        }[entity_type]
        statement = select(model).where(model.code == code)
        if for_update:
            statement = statement.with_for_update()
        return await self.session.scalar(statement)

    async def _row_business_errors(
        self,
        entity_type: ImportEntityType,
        payload: BaseModel,
        existing: Any | None,
        *,
        for_update: bool,
    ) -> list[str]:
        errors: list[str] = []
        if (
            entity_type is ImportEntityType.DOCUMENT_STATUSES
            and getattr(payload, "is_initial")
        ):
            initial = await self.document_statuses.get_initial(
                exclude_id=(
                    existing.id if isinstance(existing, DocumentStatus) else None
                ),
                for_update=for_update,
            )
            if initial is not None:
                errors.append("Another initial document status already exists.")
        if (
            entity_type is ImportEntityType.VALIDATION_RULES
            and getattr(payload, "is_default")
        ):
            default = await self.validation_rules.get_default(
                getattr(payload, "document_type_id"),
                exclude_id=(
                    existing.id if isinstance(existing, ValidationRule) else None
                ),
                for_update=for_update,
            )
            if default is not None:
                errors.append(
                    "Another default validation rule already exists for this scope."
                )
        return errors

    @staticmethod
    def _preview_response(
        entity_type: ImportEntityType,
        rows: list[ParsedImportRow],
    ) -> ImportPreviewResponse:
        return ImportPreviewResponse(
            entity_type=entity_type,
            total_rows=len(rows),
            valid_rows=sum(
                row.status is ImportRowStatus.VALID for row in rows
            ),
            invalid_rows=sum(
                row.status is ImportRowStatus.INVALID for row in rows
            ),
            duplicate_rows=sum(
                row.status is ImportRowStatus.DUPLICATE for row in rows
            ),
            rows=[
                ImportPreviewRow(
                    row_number=row.row_number,
                    status=row.status,
                    data=row.preview_data(),
                    errors=row.errors,
                )
                for row in rows
            ],
            warnings=(
                [
                    "Duplicate rows are skipped in CREATE_ONLY mode and "
                    "updated in UPSERT mode when they already exist in the "
                    "database."
                ]
                if any(
                    row.status is ImportRowStatus.DUPLICATE for row in rows
                )
                else []
            ),
        )

    def _create_entity(
        self,
        entity_type: ImportEntityType,
        row: ParsedImportRow,
    ) -> None:
        assert row.payload is not None
        values = row.payload.model_dump(by_alias=False)
        if entity_type is ImportEntityType.VALIDATION_RULES:
            values["language_order_json"] = values.pop("language_order")
            values["required_sections_json"] = values.pop("required_sections")
        category = values.get("category")
        if category is not None:
            values["category"] = category.value
        model = {
            ImportEntityType.DEPARTMENTS: Department,
            ImportEntityType.SECTIONS: Section,
            ImportEntityType.DOCUMENT_TYPES: DocumentType,
            ImportEntityType.DOCUMENT_STATUSES: DocumentStatus,
            ImportEntityType.VALIDATION_RULES: ValidationRule,
        }[entity_type]
        entity = model(
            **values,
            created_by=self.user.id,
            updated_by=self.user.id,
        )
        self.session.add(entity)
        row.existing = entity
        if (
            entity_type is ImportEntityType.VALIDATION_RULES
            and entity.document_type_id is not None
            and entity.is_default
        ):
            document_type = row.context.get("documentType")
            if document_type is not None:
                document_type.default_validation_rule = entity
                document_type.updated_by = self.user.id

    async def _update_entity(
        self,
        entity_type: ImportEntityType,
        row: ParsedImportRow,
    ) -> None:
        assert row.payload is not None and row.existing is not None
        old_document_type_id = (
            row.existing.document_type_id
            if entity_type is ImportEntityType.VALIDATION_RULES
            else None
        )
        old_was_default = bool(
            entity_type is ImportEntityType.VALIDATION_RULES
            and row.existing.is_default
        )
        values = row.payload.model_dump(by_alias=False)
        if entity_type is ImportEntityType.VALIDATION_RULES:
            values["language_order_json"] = values.pop("language_order")
            values["required_sections_json"] = values.pop("required_sections")
        category = values.get("category")
        if category is not None:
            values["category"] = category.value
        for key, value in values.items():
            setattr(row.existing, key, value)
        row.existing.updated_by = self.user.id
        if (
            entity_type is ImportEntityType.VALIDATION_RULES
            and old_document_type_id is not None
            and old_was_default
            and (
                old_document_type_id != row.existing.document_type_id
                or not row.existing.is_default
            )
        ):
            old_type = await self.document_types.get_by_id(
                old_document_type_id,
                for_update=True,
            )
            if (
                old_type is not None
                and old_type.default_validation_rule_id == row.existing.id
            ):
                old_type.default_validation_rule_id = None
                old_type.updated_by = self.user.id
        if (
            entity_type is ImportEntityType.VALIDATION_RULES
            and row.existing.document_type_id is not None
            and row.existing.is_default
        ):
            document_type = row.context.get("documentType")
            if document_type is not None:
                document_type.default_validation_rule_id = row.existing.id
                document_type.updated_by = self.user.id

    async def _export_items(
        self,
        entity_type: ImportEntityType,
        *,
        max_rows: int,
        search: str | None,
        is_active: bool | None,
        department_id: Any | None,
        document_type_id: Any | None,
        category: str | None,
    ) -> tuple[list[Any], int]:
        common = {
            "search": search,
            "is_active": is_active,
            "page": 1,
            "page_size": max_rows,
            "sort_by": "code",
            "sort_order": "asc",
        }
        if entity_type is ImportEntityType.DEPARTMENTS:
            return await self.departments.list_page(**common)
        if entity_type is ImportEntityType.SECTIONS:
            return await self.sections.list_page(
                department_id=department_id,
                **common,
            )
        if entity_type is ImportEntityType.DOCUMENT_TYPES:
            return await self.document_types.list_page(
                category=category,
                **common,
            )
        if entity_type is ImportEntityType.DOCUMENT_STATUSES:
            common["sort_by"] = "displayOrder"
            return await self.document_statuses.list_page(**common)
        return await self.validation_rules.list_page(
            document_type_id=document_type_id,
            is_default=None,
            **common,
        )

    @staticmethod
    def _export_row(
        entity_type: ImportEntityType,
        entity: Any,
    ) -> dict[str, Any]:
        common = {
            "code": entity.code,
            "name": entity.name,
            "description": entity.description,
            "is_active": entity.is_active,
        }
        if entity_type is ImportEntityType.DEPARTMENTS:
            return common
        if entity_type is ImportEntityType.SECTIONS:
            return {
                "department_code": entity.department.code,
                **common,
            }
        if entity_type is ImportEntityType.DOCUMENT_TYPES:
            return {
                "code": entity.code,
                "name": entity.name,
                "category": entity.category,
                "description": entity.description,
                "requires_section": entity.requires_section,
                "is_active": entity.is_active,
            }
        if entity_type is ImportEntityType.DOCUMENT_STATUSES:
            return {
                **common,
                "display_order": entity.display_order,
                "is_initial": entity.is_initial,
                "is_final": entity.is_final,
                "is_obsolete": entity.is_obsolete,
            }
        return {
            "code": entity.code,
            "name": entity.name,
            "document_type_code": (
                entity.document_type.code
                if entity.document_type is not None
                else None
            ),
            "required_indonesian": entity.required_indonesian,
            "required_english": entity.required_english,
            "required_chinese": entity.required_chinese,
            "minimum_indonesian_coverage": (
                entity.minimum_indonesian_coverage
            ),
            "minimum_english_coverage": entity.minimum_english_coverage,
            "minimum_chinese_coverage": entity.minimum_chinese_coverage,
            "validate_language_order": entity.validate_language_order,
            "language_order": ",".join(entity.language_order_json),
            "validate_sections": entity.validate_sections,
            "required_sections": ",".join(entity.required_sections_json),
            "validate_tables": entity.validate_tables,
            "minimum_compliance_score": entity.minimum_compliance_score,
            "partial_compliance_score": entity.partial_compliance_score,
            "is_default": entity.is_default,
            "is_active": entity.is_active,
        }

    @staticmethod
    def _format_sheet(sheet: Any, column_count: int) -> None:
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(column_count)}{max(sheet.max_row, 1)}"
        )
        for column in range(1, column_count + 1):
            maximum = 12
            for cell in sheet.iter_cols(
                min_col=column,
                max_col=column,
                min_row=1,
                max_row=sheet.max_row,
            ):
                for item in cell:
                    maximum = max(maximum, len(str(item.value or "")) + 2)
            sheet.column_dimensions[get_column_letter(column)].width = min(
                maximum,
                60,
            )
