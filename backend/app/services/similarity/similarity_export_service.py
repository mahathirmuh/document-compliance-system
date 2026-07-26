"""Private JSON/XLSX export for scoped stored similarity results."""

from __future__ import annotations

import asyncio
import json
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook

from app.core.authorization import AuditAction, Permission, has_permission
from app.core.exceptions import AuthorizationError
from app.schemas.similarity import (
    SectionSimilaritySummaryResponse,
    TranslationSimilarityResultResponse,
)
from app.services.compliance.compliance_export_service import (
    spreadsheet_safe_value,
)
from app.services.documents.base import document_error
from app.services.similarity.similarity_query_service import (
    SimilarityQueryService,
    section_similarity_response,
    similarity_result_response,
    similarity_run_response,
)

_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]+")


@dataclass(frozen=True, slots=True)
class SimilarityExportArtifact:
    path: Path
    filename: str
    media_type: str


class SimilarityExportService(SimilarityQueryService):
    async def export(
        self,
        run_id: UUID,
        *,
        export_format: str,
    ) -> SimilarityExportArtifact:
        if not has_permission(
            self.user.role,
            Permission.SIMILARITY_EXPORT,
            is_superuser=self.user.is_superuser,
        ):
            raise AuthorizationError()
        normalized = export_format.strip().lower()
        if normalized not in {"json", "xlsx"}:
            raise document_error(
                "Export format must be either json or xlsx.",
                field="format",
                code="SIMILARITY_EXPORT_FORMAT_INVALID",
                title="Similarity export format is invalid.",
            )
        run = await self._run(run_id)
        result_total = await self.results.count_for_run(run.id)
        section_total = await self.sections.count_for_run(run.id)
        maximum = int(
            getattr(
                self.settings,
                "report_export_max_rows",
                getattr(
                    self.settings,
                    "compliance_export_max_rows",
                    500_000,
                ),
            )
        )
        if max(result_total, section_total) > maximum:
            raise document_error(
                "The similarity result exceeds the configured export limit.",
                code="SIMILARITY_EXPORT_LIMIT_EXCEEDED",
                status_code=413,
                title="Similarity result is too large to export.",
            )
        results = await self._all_results(run.id, result_total)
        sections = await self._all_sections(run.id, section_total)
        descriptor, raw_path = tempfile.mkstemp(
            prefix="similarity-result-", suffix=f".{normalized}"
        )
        os.close(descriptor)
        path = Path(raw_path)
        try:
            run_payload = similarity_run_response(run).model_dump(
                mode="json", by_alias=True
            )
            if normalized == "json":
                await asyncio.to_thread(
                    _write_json,
                    path,
                    {
                        "metadata": {
                            "reportType": "TRANSLATION_SIMILARITY",
                            "rawDocumentTextIncluded": False,
                            "disclaimer": (
                                "Similarity is a review signal, not legal or "
                                "linguistic proof."
                            ),
                        },
                        "run": run_payload,
                        "results": [
                            item.model_dump(mode="json", by_alias=True)
                            for item in results
                        ],
                        "sections": [
                            item.model_dump(mode="json", by_alias=True)
                            for item in sections
                        ],
                    },
                )
                media_type = "application/json"
            else:
                await asyncio.to_thread(
                    _write_workbook,
                    path,
                    run_payload,
                    results,
                    sections,
                )
                media_type = (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )
            await self.audits.create(
                user_id=self.user.id,
                action=AuditAction.EXPORT_TRANSLATION_SIMILARITY,
                entity_type="SimilarityRun",
                entity_id=run.id,
                description="Translation similarity result exported.",
                new_values={
                    "format": normalized,
                    "documentFileId": str(run.document_file_id),
                    "resultCount": len(results),
                    "sectionCount": len(sections),
                    "rawDocumentTextIncluded": False,
                },
                ip_address=self.metadata.ip_address,
                user_agent=self.metadata.user_agent,
            )
            await self.session.commit()
        except Exception:
            await self.session.rollback()
            remove_similarity_export_artifact(path)
            raise
        code = _SAFE_FILENAME_RE.sub(
            "-", run.revision.full_document_code
        ).strip("-")[:120]
        return SimilarityExportArtifact(
            path=path,
            filename=f"{code or 'document'}_similarity.{normalized}",
            media_type=media_type,
        )

    async def _all_results(
        self,
        run_id: UUID,
        total: int,
    ) -> list[TranslationSimilarityResultResponse]:
        output: list[TranslationSimilarityResultResponse] = []
        page = 1
        batch = max(
            1,
            min(
                500,
                int(
                    getattr(
                        self.settings,
                        "similarity_db_batch_size",
                        500,
                    )
                ),
            ),
        )
        while len(output) < total:
            items, _ = await self.results.list_for_run(
                run_id, page=page, page_size=batch
            )
            if not items:
                break
            output.extend(similarity_result_response(item) for item in items)
            page += 1
        return output

    async def _all_sections(
        self,
        run_id: UUID,
        total: int,
    ) -> list[SectionSimilaritySummaryResponse]:
        output: list[SectionSimilaritySummaryResponse] = []
        page = 1
        while len(output) < total:
            items, _ = await self.sections.list_for_run(
                run_id, page=page, page_size=500
            )
            if not items:
                break
            output.extend(section_similarity_response(item) for item in items)
            page += 1
        return output


def _write_json(path: Path, payload: dict[str, object]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as output:
        json.dump(
            payload,
            output,
            ensure_ascii=False,
            separators=(",", ":"),
        )


def _write_workbook(
    path: Path,
    run: dict[str, object],
    results: list[TranslationSimilarityResultResponse],
    sections: list[SectionSimilaritySummaryResponse],
) -> None:
    workbook = Workbook(write_only=True)
    overview = workbook.create_sheet("Overview")
    overview.append(["Field", "Value"])
    for key, value in run.items():
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        overview.append(
            [spreadsheet_safe_value(key), spreadsheet_safe_value(value)]
        )
    _append_schema_rows(
        workbook,
        "Pairwise Results",
        [
            item.model_dump(mode="json", by_alias=True)
            for item in results
        ],
    )
    _append_schema_rows(
        workbook,
        "Section Summary",
        [
            item.model_dump(mode="json", by_alias=True)
            for item in sections
        ],
    )
    workbook.save(path)


def _append_schema_rows(
    workbook: Workbook,
    name: str,
    rows: list[dict[str, object]],
) -> None:
    sheet = workbook.create_sheet(name)
    if not rows:
        sheet.append(["No data"])
        return
    headers = list(rows[0])
    sheet.append([spreadsheet_safe_value(value) for value in headers])
    for row in rows:
        values: list[object] = []
        for header in headers:
            value = row.get(header)
            if isinstance(value, (dict, list)):
                value = json.dumps(
                    value, ensure_ascii=False, separators=(",", ":")
                )
            values.append(spreadsheet_safe_value(value))
        sheet.append(values)


def remove_similarity_export_artifact(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
