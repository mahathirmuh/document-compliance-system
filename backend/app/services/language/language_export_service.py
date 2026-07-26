"""Bounded, private JSON/XLSX export of language detection results."""

from __future__ import annotations

import json
import os
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path
from uuid import UUID

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.authorization import AuditAction
from app.core.config import Settings
from app.models.language_detection_run import LanguageDetectionRun
from app.models.user import User
from app.repositories.language_block_result_repository import (
    LanguageBlockReadRow,
    LanguageBlockResultRepository,
)
from app.repositories.language_container_summary_repository import (
    LanguageContainerSummaryRepository,
)
from app.schemas.language_detection import LanguageDetectionRunResponse
from app.services.auth.auth_service import RequestMetadata
from app.services.documents.base import document_error
from app.services.documents.xlsx_safety import excel_safe
from app.services.language.language_detection_job_service import (
    LanguageResultService,
    language_block_response,
    language_container_response,
)
from app.services.language.language_runtime_config import (
    LanguageRuntimeConfig,
)


@dataclass(frozen=True, slots=True)
class LanguageExportArtifact:
    path: Path
    filename: str
    media_type: str


class LanguageExportService(LanguageResultService):
    """Export exact stored decisions, scores, and provenance."""

    def __init__(
        self,
        session: AsyncSession,
        settings: Settings,
        user: User,
        metadata: RequestMetadata,
    ) -> None:
        super().__init__(session, settings, user, metadata)
        self.config = LanguageRuntimeConfig.from_settings(settings)
        self.export_blocks = LanguageBlockResultRepository(session)
        self.export_containers = LanguageContainerSummaryRepository(session)

    async def export(
        self,
        run_id: UUID,
        *,
        export_format: str,
    ) -> LanguageExportArtifact:
        normalized_format = export_format.strip().lower()
        if normalized_format not in {"json", "xlsx"}:
            raise document_error(
                "Export format must be either json or xlsx.",
                field="format",
                title="Language export format is invalid.",
            )
        run = await self._scoped_run(run_id)
        total_blocks = await self.export_blocks.count(run.id)
        if total_blocks > self.config.export_maximum_blocks:
            raise document_error(
                "The language result exceeds the configured export limit.",
                status_code=413,
                title="Language result is too large to export.",
            )
        run_response = await self._run_response(run)
        descriptor, raw_path = tempfile.mkstemp(
            prefix="language-detection-",
            suffix=f".{normalized_format}",
        )
        os.close(descriptor)
        path = Path(raw_path)
        try:
            if normalized_format == "json":
                await self._write_json(path, run, run_response)
                media_type = "application/json"
            else:
                await self._write_xlsx(path, run, run_response)
                media_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            await self.audit(
                action=AuditAction.EXPORT_LANGUAGE_RESULT,
                entity_type="LanguageDetectionRun",
                entity_id=run.id,
                description="Language detection result exported.",
                new_values={
                    "format": normalized_format,
                    "documentFileId": str(run.document_file_id),
                    "preliminary": True,
                },
            )
            await self.session.commit()
        except Exception:
            await self.session.rollback()
            remove_language_export_artifact(path)
            raise
        return LanguageExportArtifact(
            path=path,
            filename=(
                f"{_safe_filename(run.revision.full_document_code)}"
                f"_language-detection.{normalized_format}"
            ),
            media_type=media_type,
        )

    async def _write_json(
        self,
        path: Path,
        run: LanguageDetectionRun,
        run_response: LanguageDetectionRunResponse,
    ) -> None:
        with path.open("w", encoding="utf-8", newline="\n") as output:
            output.write('{"run":')
            _write_json_value(
                output,
                run_response.model_dump(mode="json", by_alias=True),
            )
            output.write(',"containers":[')
            first = True
            page = 1
            while True:
                items, total = await self.export_containers.list(
                    run.id,
                    page=page,
                    page_size=500,
                )
                for item in items:
                    if not first:
                        output.write(",")
                    _write_json_value(
                        output,
                        language_container_response(item).model_dump(
                            mode="json",
                            by_alias=True,
                        ),
                    )
                    first = False
                if page * 500 >= total:
                    break
                page += 1
            output.write('],"blocks":[')
            first = True
            page = 1
            while True:
                rows, total = await self.export_blocks.list(
                    run.id,
                    page=page,
                    page_size=500,
                )
                for row in rows:
                    if not first:
                        output.write(",")
                    _write_json_value(
                        output,
                        language_block_response(row).model_dump(
                            mode="json",
                            by_alias=True,
                        ),
                    )
                    first = False
                if page * 500 >= total:
                    break
                page += 1
            output.write("]}")

    async def _write_xlsx(
        self,
        path: Path,
        run: LanguageDetectionRun,
        run_response: LanguageDetectionRunResponse,
    ) -> None:
        workbook = Workbook(write_only=True)
        summary_sheet = workbook.create_sheet("Summary")
        container_sheet = workbook.create_sheet("Containers")
        block_sheet = workbook.create_sheet("Blocks")
        container_sheet.append(
            [
                "Container",
                "Container Type",
                "Container Index",
                "Dominant Language",
                "Eligible Blocks",
                "Indonesian Blocks",
                "English Blocks",
                "Chinese Blocks",
                "Mixed Blocks",
                "Unknown Blocks",
                "Other Blocks",
                "Language Presence",
                "Preliminary Coverage",
            ]
        )
        page = 1
        while True:
            items, total = await self.export_containers.list(
                run.id,
                page=page,
                page_size=500,
            )
            for item in items:
                container_sheet.append(
                    [
                        excel_safe(item.container_name or ""),
                        item.container_type,
                        item.container_index,
                        item.dominant_language,
                        item.eligible_blocks,
                        item.indonesian_blocks,
                        item.english_blocks,
                        item.chinese_blocks,
                        item.mixed_blocks,
                        item.unknown_blocks,
                        item.other_blocks,
                        json.dumps(
                            item.language_presence_json,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        ),
                        json.dumps(
                            item.coverage_json,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        ),
                    ]
                )
            if page * 500 >= total:
                break
            page += 1

        block_sheet.append(
            [
                "Source Type",
                "Container",
                "Source Reference",
                "Text",
                "Language",
                "Primary Language",
                "Confidence",
                "Mixed",
                "Character Count",
                "Han Characters",
                "Latin Characters",
                "Eligibility",
                "Eligibility Reason",
                "Source Confidence",
                "Detected Languages",
                "Script Statistics",
            ]
        )
        page = 1
        while True:
            rows, total = await self.export_blocks.list(
                run.id,
                page=page,
                page_size=500,
            )
            for row in rows:
                self._append_block_row(block_sheet, row)
            if page * 500 >= total:
                break
            page += 1

        summary_sheet.append(
            [
                "Language",
                "Presence",
                "Block Count",
                "Character Count",
                "Block Coverage",
                "Character Coverage",
                "Average Confidence",
            ]
        )
        averages = await self.export_blocks.average_confidence_by_language(run.id)
        summary = run_response
        raw_other_characters = (run.metadata_json or {}).get(
            "otherCharacters",
            0,
        )
        other_characters = (
            raw_other_characters
            if isinstance(raw_other_characters, int)
            and not isinstance(raw_other_characters, bool)
            and raw_other_characters >= 0
            else 0
        )
        count_values = {
            "id": (summary.indonesian_blocks, summary.indonesian_characters),
            "en": (summary.english_blocks, summary.english_characters),
            "zh": (summary.chinese_blocks, summary.chinese_characters),
            "mixed": (summary.mixed_blocks, summary.mixed_characters),
            "unknown": (summary.unknown_blocks, summary.unknown_characters),
            "other": (summary.other_blocks, other_characters),
        }
        block_coverage = summary.coverage.block_coverage.model_dump()
        character_coverage = summary.coverage.character_coverage.model_dump()
        presence = summary.language_presence.model_dump(mode="json")
        for code in ("id", "en", "zh", "mixed", "unknown", "other"):
            block_count, character_count = count_values[code]
            summary_sheet.append(
                [
                    code,
                    presence.get(code, "NOT_APPLICABLE"),
                    block_count,
                    character_count,
                    block_coverage.get(code, 0.0),
                    character_coverage.get(code, 0.0),
                    averages.get(code),
                ]
            )
        summary_sheet.append([])
        summary_sheet.append(
            [
                "Disclaimer",
                (
                    "Preliminary coverage only; this export does not "
                    "represent translation equivalence or final compliance."
                ),
            ]
        )
        workbook.save(path)

    @staticmethod
    def _append_block_row(
        sheet: object,
        row: LanguageBlockReadRow,
    ) -> None:
        result = row.result
        metadata = result.metadata_json or {}
        container = metadata.get("containerIndex", "")
        sheet.append(  # type: ignore[attr-defined]
            [
                result.source_type.value,
                container,
                excel_safe(result.source_reference),
                excel_safe(row.text),
                result.language_code.value,
                result.primary_language_code.value,
                float(result.confidence),
                result.is_mixed,
                result.character_count,
                result.han_character_count,
                result.latin_character_count,
                result.eligibility_status.value,
                (
                    result.eligibility_reason.value
                    if result.eligibility_reason is not None
                    else ""
                ),
                row.source_confidence,
                json.dumps(
                    result.detected_languages_json,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                json.dumps(
                    result.script_statistics_json,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            ]
        )


def _write_json_value(output: object, value: object) -> None:
    json.dump(
        value,
        output,
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _safe_filename(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return normalized[:180] or "document"


def remove_language_export_artifact(path: Path) -> None:
    """Delete only the exact temporary artifact produced by this service."""
    try:
        if path.is_file() or path.is_symlink():
            path.unlink(missing_ok=True)
    except OSError:
        return
