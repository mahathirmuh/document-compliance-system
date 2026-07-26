"""Formula-safe multi-sheet XLSX advanced reports."""

from __future__ import annotations

import io
import json
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.advanced_reporting import AdvancedReportFilters
from app.services.reporting.report_dataset_service import ReportDataset


class ReportXlsxService:
    def __init__(self, *, maximum_rows: int) -> None:
        self.maximum_rows = maximum_rows

    def build(
        self,
        dataset: ReportDataset,
        *,
        report_name: str,
        filters: AdvancedReportFilters,
    ) -> bytes:
        workbook = Workbook()
        summary = workbook.active
        assert summary is not None
        summary.title = "Summary"
        summary.append(["Report Name", _safe_value(report_name)])
        summary.append(["Report Type", dataset.report_type.value])
        summary.append(["Generated At", _safe_value(datetime.now(UTC))])
        summary.append(["Disclaimer", (
            "Automated quality metrics are review signals and do not "
            "constitute legal or linguistic proof."
        )])
        summary.append([])
        summary.append(["Metric", "Value"])
        for cell in summary[6]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for key, value in dataset.summary.items():
            summary.append([key, _safe_value(value)])
        summary.freeze_panes = "A6"
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 70

        filter_sheet = workbook.create_sheet("Filters")
        filter_sheet.append(["Filter", "Value"])
        for cell in filter_sheet[1]:
            cell.font = Font(bold=True)
        for key, value in filters.model_dump(
            mode="json", by_alias=True
        ).items():
            filter_sheet.append([key, _safe_value(value)])
        filter_sheet.freeze_panes = "A2"
        filter_sheet.auto_filter.ref = filter_sheet.dimensions
        filter_sheet.column_dimensions["A"].width = 32
        filter_sheet.column_dimensions["B"].width = 70

        for name, rows in dataset.tables.items():
            self._add_table(workbook, name, rows)
        if dataset.data_series:
            self._add_table(
                workbook, "Chart Data", dataset.data_series
            )
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _add_table(
        self,
        workbook: Workbook,
        name: str,
        rows: list[dict[str, object]],
    ) -> None:
        if len(rows) > self.maximum_rows:
            raise ValueError("Report dataset exceeds the XLSX row limit.")
        safe_name = self._unique_name(workbook, name)
        sheet = workbook.create_sheet(safe_name)
        if not rows:
            sheet.append(["No data"])
            return
        headers = list(rows[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in rows:
            sheet.append(
                [_safe_value(row.get(header)) for header in headers]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, header in enumerate(headers, start=1):
            maximum = max(
                len(str(header)),
                *(
                    len(str(sheet.cell(row, column_index).value or ""))
                    for row in range(2, sheet.max_row + 1)
                ),
            )
            sheet.column_dimensions[
                get_column_letter(column_index)
            ].width = min(maximum + 2, 60)
            if any(
                token in str(header).lower()
                for token in ("score", "similarity", "percentage")
            ) and sheet.max_row > 2:
                letter = get_column_letter(column_index)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )

    @staticmethod
    def _unique_name(workbook: Workbook, name: str) -> str:
        base = "".join(
            "_" if character in r"[]:*?/\\" else character
            for character in name
        )[:31] or "Data"
        candidate = base
        suffix = 2
        while candidate in workbook.sheetnames:
            tail = f"_{suffix}"
            candidate = f"{base[: 31 - len(tail)]}{tail}"
            suffix += 1
        return candidate


def _safe_value(value: object) -> object:
    if isinstance(value, datetime) and value.tzinfo is not None:
        value = value.astimezone(UTC).replace(tzinfo=None)
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
