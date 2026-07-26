"""Focused persistence, workflow, scope, and export tests for Phase 8 findings."""

from __future__ import annotations

import json
from collections.abc import Callable
from datetime import date
from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from uuid import UUID, uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects import postgresql
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.authorization import AuditAction, UserRole
from app.core.config import get_settings
from app.core.exceptions import ApplicationError
from app.main import app
from app.models.audit_log import AuditLog
from app.models.department import Department
from app.models.document import Document
from app.models.document_file import DocumentFile, DocumentFileStatus
from app.models.document_revision import DocumentRevision
from app.models.validation_finding import ValidationFinding
from app.schemas.finding import (
    FindingCreateManualRequest,
    FindingFilter,
    FindingUpdateRequest,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.auth.token_service import TokenService
from app.services.compliance.findings.finding_export_service import (
    FindingExportService,
    remove_finding_export_artifact,
)
from app.services.compliance.findings.finding_management_service import (
    FindingManagementService,
)

TestSessionFactory = async_sessionmaker[AsyncSession]
UserFactory = Callable[..., Any]

_METADATA = RequestMetadata(
    ip_address="127.0.0.1",
    user_agent="phase8-finding-test",
)


def test_finding_filters_apply_exact_section_id_and_canonical_code() -> None:
    section_id = uuid4()
    filters = FindingFilter(
        detected_section_id=section_id,
        section_code=" purpose ",
    )

    predicates = FindingManagementService._filter_predicates(
        SimpleNamespace(),
        filters,
        department_ids=None,
    )
    statement = select(ValidationFinding).where(*predicates)
    sql = str(
        statement.compile(
            dialect=postgresql.dialect(),
            compile_kwargs={"literal_binds": True},
        )
    )

    assert str(section_id) in sql
    assert "validation_findings.detected_section_id" in sql
    assert "upper(detected_sections.canonical_code) = 'PURPOSE'" in sql


async def _department(
    session_factory: TestSessionFactory,
    *,
    code: str,
) -> Department:
    department = Department(
        id=uuid4(),
        code=code,
        name=f"{code} Department",
        is_active=True,
    )
    async with session_factory() as session:
        session.add(department)
        await session.commit()
    return department


async def _source(
    session_factory: TestSessionFactory,
    *,
    department_id: UUID,
    suffix: str,
) -> tuple[Document, DocumentRevision, DocumentFile]:
    document_id = uuid4()
    revision_id = uuid4()
    file_id = uuid4()
    document = Document(
        id=document_id,
        company_code="MTI",
        department_id=department_id,
        document_type_id=uuid4(),
        document_number=suffix,
        base_document_code=f"MTI-{suffix}-POL-001",
        title=f"Finding source {suffix}",
    )
    revision = DocumentRevision(
        id=revision_id,
        document_id=document_id,
        revision_code="Rev.000",
        revision_number=0,
        full_document_code=f"MTI-{suffix}-POL-001_Rev.000",
        document_status_id=uuid4(),
        is_current=True,
    )
    document_file = DocumentFile(
        id=file_id,
        document_id=document_id,
        document_revision_id=revision_id,
        original_filename=f"MTI-{suffix}-POL-001_Rev.000.pdf",
        sanitized_filename=f"MTI-{suffix}-POL-001_Rev.000.pdf",
        file_extension="pdf",
        mime_type="application/pdf",
        detected_mime_type="application/pdf",
        file_size=100,
        sha256_hash=sha256(suffix.encode()).hexdigest(),
        storage_key=f"documents/originals/findings/{suffix}.pdf",
        file_status=DocumentFileStatus.AVAILABLE,
        is_primary=True,
        is_current=True,
    )
    async with session_factory() as session:
        session.add_all([document, revision, document_file])
        await session.commit()
    return document, revision, document_file


def _manual_payload(
    source: tuple[Document, DocumentRevision, DocumentFile],
    *,
    severity: str = "MAJOR",
    title: str = "Manual multilingual review",
    source_reference: str = "PDF:page=1:block=2",
) -> FindingCreateManualRequest:
    document, revision, document_file = source
    return FindingCreateManualRequest(
        documentId=document.id,
        documentRevisionId=revision.id,
        documentFileId=document_file.id,
        severity=severity,
        title=title,
        description="The multilingual structure requires a human review.",
        recommendation="Confirm the approved structure.",
        sourceReference=source_reference,
        pageNumber=1,
    )


@pytest.mark.asyncio
async def test_manual_finding_is_retained_scoped_and_audited(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department_a = await _department(session_factory, code="FA")
    department_b = await _department(session_factory, code="FB")
    source = await _source(
        session_factory,
        department_id=department_a.id,
        suffix="FA",
    )
    controller = await create_user(
        email="finding-controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department_a.id,
    )
    outsider = await create_user(
        email="finding-outsider@example.com",
        role=UserRole.DEPARTMENT_USER,
        department_id=department_b.id,
    )

    async with session_factory() as session:
        created = await FindingManagementService(
            session,
            controller,
            _METADATA,
        ).create_manual(_manual_payload(source))
        assert created.finding_code.value == "MANUAL_FINDING"
        assert created.finding_type.value == "MANUAL"
        assert created.status.value == "OPEN"
        assert created.is_system_generated is False
        assert isinstance(created.history, list)
        assert created.history[0].action == "CREATE_MANUAL_FINDING"
        finding_id = created.id

    async with session_factory() as session:
        persisted = await session.get(ValidationFinding, finding_id)
        assert persisted is not None
        assert persisted.compliance_run_id is None
        assert persisted.created_by == controller.id
        actions = list(
            (
                await session.scalars(
                    select(AuditLog.action).where(AuditLog.entity_id == finding_id)
                )
            ).all()
        )
        assert actions == [AuditAction.CREATE_MANUAL_FINDING]

    async with session_factory() as session:
        service = FindingManagementService(session, outsider, _METADATA)
        result = await service.list(FindingFilter())
        assert result.total_items == 0
        with pytest.raises(ApplicationError) as outside_department:
            await service.list(FindingFilter(departmentId=department_a.id))
        assert outside_department.value.status_code == 403
        assert outside_department.value.errors is not None
        assert (
            outside_department.value.errors[0].code == "FINDING_DEPARTMENT_SCOPE_DENIED"
        )
        with pytest.raises(ApplicationError) as error:
            await service.get(finding_id)
        assert error.value.status_code == 404


@pytest.mark.asyncio
async def test_workflow_update_assignment_and_invalid_transition_are_atomic(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="WF")
    other_department = await _department(session_factory, code="WO")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="WF",
    )
    controller = await create_user(
        email="finding-workflow@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    assignee = await create_user(
        email="finding-assignee@example.com",
        role=UserRole.REVIEWER,
        department_id=department.id,
    )
    scoped_editor = await create_user(
        email="finding-scoped-editor@example.com",
        role=UserRole.DEPARTMENT_USER,
        department_id=department.id,
    )
    outside_assignee = await create_user(
        email="finding-outside-assignee@example.com",
        role=UserRole.REVIEWER,
        department_id=other_department.id,
    )

    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        finding = await service.create_manual(_manual_payload(source))
        finding = await service.update(
            finding.id,
            FindingUpdateRequest(
                severity="CRITICAL",
                title="Confirmed structural exception",
            ),
        )
        assert finding.severity.value == "CRITICAL"
        assert finding.title == "Confirmed structural exception"

        finding = await service.review(
            finding.id,
            comment="Confirmed by the document reviewer.",
        )
        assert finding.status.value == "IN_REVIEW"
        assert finding.model_dump(mode="json", by_alias=True)["reviewedBy"][
            "id"
        ] == str(controller.id)
        assert finding.reviewed_at is not None

        finding = await service.resolve(
            finding.id,
            comment="Corrected in the controlled source.",
        )
        assert finding.status.value == "RESOLVED"
        assert finding.model_dump(mode="json", by_alias=True)["resolvedBy"][
            "id"
        ] == str(controller.id)

        with pytest.raises(ApplicationError) as invalid:
            await service.resolve(
                finding.id,
                comment="A duplicate transition is not valid.",
            )
        assert invalid.value.status_code == 409
        assert invalid.value.errors is not None
        assert invalid.value.errors[0].code == "FINDING_INVALID_STATUS_TRANSITION"

        finding = await service.reopen(
            finding.id,
            reason="The issue reappeared after file replacement.",
        )
        finding = await service.mark_false_positive(
            finding.id,
            reason="The approved exception applies to this source.",
        )
        assert finding.status.value == "FALSE_POSITIVE"
        assert finding.model_dump(mode="json", by_alias=True)["falsePositiveBy"][
            "id"
        ] == str(controller.id)

        finding = await service.assign(
            finding.id,
            assigned_to=assignee.id,
        )
        assert finding.model_dump(mode="json", by_alias=True)["assignedTo"][
            "id"
        ] == str(assignee.id)
        assert isinstance(finding.history, list)
        assert len(finding.history) == 7

    async with session_factory() as session:
        with pytest.raises(ApplicationError) as outside_assignment:
            await FindingManagementService(
                session,
                scoped_editor,
                _METADATA,
            ).assign(
                finding.id,
                assigned_to=outside_assignee.id,
            )
        assert outside_assignment.value.status_code == 400
        assert outside_assignment.value.errors is not None
        assert outside_assignment.value.errors[0].code == "FINDING_ASSIGNMENT_INVALID"

    async with session_factory() as session:
        actions = list(
            (
                await session.scalars(
                    select(AuditLog.action)
                    .where(AuditLog.entity_id == finding.id)
                    .order_by(AuditLog.created_at, AuditLog.id)
                )
            ).all()
        )
        assert len(actions) == 7
        assert set(actions) == {
            AuditAction.CREATE_MANUAL_FINDING,
            AuditAction.UPDATE_FINDING,
            AuditAction.REVIEW_FINDING,
            AuditAction.RESOLVE_FINDING,
            AuditAction.REOPEN_FINDING,
            AuditAction.MARK_FINDING_FALSE_POSITIVE,
            AuditAction.ASSIGN_FINDING,
        }


@pytest.mark.asyncio
async def test_bulk_assign_and_review_deduplicate_and_audit_each_finding(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="BK")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="BK",
    )
    controller = await create_user(
        email="finding-bulk-controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    editor = await create_user(
        email="finding-bulk-editor@example.com",
        role=UserRole.DEPARTMENT_USER,
        department_id=department.id,
    )
    reviewer = await create_user(
        email="finding-bulk-reviewer@example.com",
        role=UserRole.REVIEWER,
        department_id=department.id,
    )

    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        first = await service.create_manual(
            _manual_payload(source, title="First bulk finding")
        )
        second = await service.create_manual(
            _manual_payload(source, title="Second bulk finding")
        )

    async with session_factory() as session:
        assigned = await FindingManagementService(
            session,
            editor,
            _METADATA,
        ).bulk_assign(
            [first.id, first.id, second.id],
            assigned_to=reviewer.id,
            maximum_items=2,
        )
        assert assigned.action == "ASSIGN"
        assert assigned.processed_count == 2
        assert assigned.finding_ids == [first.id, second.id]

    async with session_factory() as session:
        reviewed = await FindingManagementService(
            session,
            reviewer,
            _METADATA,
        ).bulk_review(
            [second.id, first.id],
            comment="  Review all selected findings.  ",
            maximum_items=2,
        )
        assert reviewed.action == "REVIEW"
        assert reviewed.processed_count == 2
        assert reviewed.finding_ids == [second.id, first.id]

    async with session_factory() as session:
        findings = list(
            (
                await session.scalars(
                    select(ValidationFinding).where(
                        ValidationFinding.id.in_([first.id, second.id])
                    )
                )
            ).all()
        )
        assert {finding.status.value for finding in findings} == {"IN_REVIEW"}
        assert {finding.assigned_to for finding in findings} == {reviewer.id}
        assert {finding.reviewed_by for finding in findings} == {reviewer.id}
        assert {finding.review_comment for finding in findings} == {
            "Review all selected findings."
        }

        bulk_audits = list(
            (
                await session.scalars(
                    select(AuditLog).where(
                        AuditLog.entity_id.in_([first.id, second.id]),
                        AuditLog.action.in_(
                            [
                                AuditAction.ASSIGN_FINDING,
                                AuditAction.REVIEW_FINDING,
                            ]
                        ),
                    )
                )
            ).all()
        )
        assert len(bulk_audits) == 4
        assert {
            (audit.entity_id, audit.action, audit.user_id) for audit in bulk_audits
        } == {
            (first.id, AuditAction.ASSIGN_FINDING, editor.id),
            (second.id, AuditAction.ASSIGN_FINDING, editor.id),
            (first.id, AuditAction.REVIEW_FINDING, reviewer.id),
            (second.id, AuditAction.REVIEW_FINDING, reviewer.id),
        }


@pytest.mark.asyncio
async def test_bulk_action_limit_is_enforced_from_endpoint_settings(
    api_client: AsyncClient,
    token_service: TokenService,
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="BL")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="BL",
    )
    controller = await create_user(
        email="finding-bulk-limit@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    assignee = await create_user(
        email="finding-bulk-limit-assignee@example.com",
        role=UserRole.REVIEWER,
        department_id=department.id,
    )
    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        first = await service.create_manual(
            _manual_payload(source, title="Limit finding one")
        )
        second = await service.create_manual(
            _manual_payload(source, title="Limit finding two")
        )

    settings = get_settings().model_copy(update={"finding_bulk_action_max_items": 1})
    app.dependency_overrides[get_settings] = lambda: settings
    response = await api_client.post(
        "/api/v1/findings/bulk-actions",
        headers={
            "Authorization": (f"Bearer {token_service.create_access_token(controller)}")
        },
        json={
            "action": "ASSIGN",
            "findingIds": [str(first.id), str(second.id)],
            "assignedTo": str(assignee.id),
        },
    )
    assert response.status_code == 422, response.text
    assert response.json()["errors"][0]["code"] == (
        "FINDING_BULK_ACTION_LIMIT_EXCEEDED"
    )

    async with session_factory() as session:
        findings = list(
            (
                await session.scalars(
                    select(ValidationFinding).where(
                        ValidationFinding.id.in_([first.id, second.id])
                    )
                )
            ).all()
        )
        assert all(finding.assigned_to is None for finding in findings)
        audits = list(
            (
                await session.scalars(
                    select(AuditLog).where(
                        AuditLog.entity_id.in_([first.id, second.id]),
                        AuditLog.action == AuditAction.ASSIGN_FINDING,
                    )
                )
            ).all()
        )
        assert audits == []


@pytest.mark.asyncio
async def test_bulk_review_rolls_back_when_one_transition_is_invalid(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="BI")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="BI",
    )
    controller = await create_user(
        email="finding-bulk-invalid@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        valid = await service.create_manual(
            _manual_payload(source, title="Valid bulk review")
        )
        invalid = await service.create_manual(
            _manual_payload(source, title="Invalid bulk review")
        )
        invalid = await service.resolve(
            invalid.id,
            comment="Already resolved before the bulk review.",
        )

    async with session_factory() as session:
        with pytest.raises(ApplicationError) as error:
            await FindingManagementService(
                session,
                controller,
                _METADATA,
            ).bulk_review(
                [valid.id, invalid.id],
                comment="This batch must be all or nothing.",
                maximum_items=10,
            )
        assert error.value.status_code == 409
        assert error.value.errors is not None
        assert error.value.errors[0].code == "FINDING_INVALID_STATUS_TRANSITION"

    async with session_factory() as session:
        valid_row = await session.get(ValidationFinding, valid.id)
        invalid_row = await session.get(ValidationFinding, invalid.id)
        assert valid_row is not None
        assert invalid_row is not None
        assert valid_row.status.value == "OPEN"
        assert valid_row.reviewed_by is None
        assert invalid_row.status.value == "RESOLVED"
        review_audits = list(
            (
                await session.scalars(
                    select(AuditLog).where(
                        AuditLog.entity_id.in_([valid.id, invalid.id]),
                        AuditLog.action == AuditAction.REVIEW_FINDING,
                    )
                )
            ).all()
        )
        assert review_audits == []


@pytest.mark.asyncio
async def test_bulk_review_scope_failure_rolls_back_visible_findings(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department_a = await _department(session_factory, code="BS")
    department_b = await _department(session_factory, code="BO")
    source_a = await _source(
        session_factory,
        department_id=department_a.id,
        suffix="BS",
    )
    source_b = await _source(
        session_factory,
        department_id=department_b.id,
        suffix="BO",
    )
    controller = await create_user(
        email="finding-bulk-scope-controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department_a.id,
    )
    reviewer = await create_user(
        email="finding-bulk-scope-reviewer@example.com",
        role=UserRole.REVIEWER,
        department_id=department_a.id,
    )
    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        visible = await service.create_manual(
            _manual_payload(source_a, title="Visible bulk finding")
        )
        outside = await service.create_manual(
            _manual_payload(source_b, title="Outside bulk finding")
        )

    async with session_factory() as session:
        with pytest.raises(ApplicationError) as error:
            await FindingManagementService(
                session,
                reviewer,
                _METADATA,
            ).bulk_review(
                [visible.id, outside.id],
                comment="Do not partially review this batch.",
                maximum_items=10,
            )
        assert error.value.status_code == 404
        assert error.value.errors is not None
        assert error.value.errors[0].code == "FINDING_NOT_FOUND"

    async with session_factory() as session:
        findings = list(
            (
                await session.scalars(
                    select(ValidationFinding).where(
                        ValidationFinding.id.in_([visible.id, outside.id])
                    )
                )
            ).all()
        )
        assert {finding.status.value for finding in findings} == {"OPEN"}
        audits = list(
            (
                await session.scalars(
                    select(AuditLog).where(
                        AuditLog.entity_id.in_([visible.id, outside.id]),
                        AuditLog.action == AuditAction.REVIEW_FINDING,
                    )
                )
            ).all()
        )
        assert audits == []


def test_bulk_action_openapi_contract_is_discriminated_and_safe() -> None:
    schema = app.openapi()
    operation = schema["paths"]["/api/v1/findings/bulk-actions"]["post"]
    request = operation["requestBody"]["content"]["application/json"]["schema"]
    assert request["discriminator"] == {
        "propertyName": "action",
        "mapping": {
            "ASSIGN": "#/components/schemas/FindingBulkAssignRequest",
            "REVIEW": "#/components/schemas/FindingBulkReviewRequest",
        },
    }
    assert set(operation["responses"]) >= {
        "200",
        "400",
        "403",
        "404",
        "409",
        "422",
    }
    assert "413" not in operation["responses"]

    schemas = schema["components"]["schemas"]
    assert (
        schemas["FindingBulkAssignRequest"]["properties"]["action"]["const"] == "ASSIGN"
    )
    assert (
        schemas["FindingBulkReviewRequest"]["properties"]["action"]["const"] == "REVIEW"
    )
    assert (
        schemas["FindingBulkAssignRequest"]["properties"]["findingIds"]["maxItems"]
        == 10_000
    )
    assert (
        schemas["FindingBulkReviewRequest"]["properties"]["findingIds"]["maxItems"]
        == 10_000
    )
    assert set(schemas["FindingBulkActionResponse"]["properties"]) == {
        "action",
        "processedCount",
        "findingIds",
    }


@pytest.mark.asyncio
async def test_in_review_finding_can_return_to_open(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="RO")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="RO",
    )
    controller = await create_user(
        email="finding-return-open@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        finding = await service.create_manual(_manual_payload(source))
        finding = await service.review(
            finding.id,
            comment="Review started.",
        )
        assert finding.status.value == "IN_REVIEW"

        finding = await service.return_to_open(
            finding.id,
            comment="More source evidence is required.",
        )

        assert finding.status.value == "OPEN"
        returned = next(
            event
            for event in finding.history
            if event.previous_status is not None
            and event.previous_status.value == "IN_REVIEW"
            and event.new_status.value == "OPEN"
        )
        assert returned.comment == ("More source evidence is required.")


@pytest.mark.asyncio
async def test_accept_risk_requires_resolve_permission(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department = await _department(session_factory, code="AR")
    source = await _source(
        session_factory,
        department_id=department.id,
        suffix="AR",
    )
    controller = await create_user(
        email="finding-risk-controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department.id,
    )
    reviewer = await create_user(
        email="finding-risk-reviewer@example.com",
        role=UserRole.REVIEWER,
        department_id=department.id,
    )
    department_user = await create_user(
        email="finding-risk-department@example.com",
        role=UserRole.DEPARTMENT_USER,
        department_id=department.id,
    )
    async with session_factory() as session:
        finding = await FindingManagementService(
            session,
            controller,
            _METADATA,
        ).create_manual(_manual_payload(source))

    async with session_factory() as session:
        with pytest.raises(ApplicationError) as forbidden:
            await FindingManagementService(
                session,
                department_user,
                _METADATA,
            ).accept_risk(
                finding.id,
                reason="Temporary approved exception.",
                expiry_date=date(2027, 1, 31),
            )
        assert forbidden.value.status_code == 403

    async with session_factory() as session:
        accepted = await FindingManagementService(
            session,
            reviewer,
            _METADATA,
        ).accept_risk(
            finding.id,
            reason="Temporary approved exception.",
            expiry_date=date(2027, 1, 31),
        )
        assert accepted.status.value == "ACCEPTED_RISK"
        assert accepted.model_dump(mode="json", by_alias=True)["acceptedRiskBy"][
            "id"
        ] == str(reviewer.id)
        assert accepted.accepted_risk_expiry_date == date(2027, 1, 31)


@pytest.mark.asyncio
async def test_filtered_exports_are_scoped_bounded_and_spreadsheet_safe(
    session_factory: TestSessionFactory,
    create_user: UserFactory,
) -> None:
    department_a = await _department(session_factory, code="EA")
    department_b = await _department(session_factory, code="EB")
    source_a = await _source(
        session_factory,
        department_id=department_a.id,
        suffix="EA",
    )
    source_b = await _source(
        session_factory,
        department_id=department_b.id,
        suffix="EB",
    )
    controller = await create_user(
        email="finding-export-controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=department_a.id,
    )
    async with session_factory() as session:
        service = FindingManagementService(session, controller, _METADATA)
        unsafe_payload = _manual_payload(
            source_a,
            severity="CRITICAL",
            title='=HYPERLINK("https://invalid.example")',
            source_reference=r"C:\private\controlled-source.pdf",
        ).model_copy(
            update={
                "location": {
                    "storage_path": r"C:\private\controlled-source.pdf",
                    "safe": "page-1",
                    "nested": {"local_path": "/private/source.pdf"},
                }
            }
        )
        first = await service.create_manual(unsafe_payload)
        assert first.source_reference == "[redacted source reference]"
        assert first.location == {"safe": "page-1", "nested": {}}
        await service.create_manual(
            _manual_payload(
                source_b,
                severity="MINOR",
                title="Other department",
            )
        )

    filters = FindingFilter(
        departmentId=department_a.id,
        createdBySystem=False,
        page=1,
        pageSize=20,
    )
    settings = get_settings().model_copy(update={"finding_export_max_rows": 10})
    async with session_factory() as session:
        artifact = await FindingExportService(
            session,
            settings,
            controller,
            _METADATA,
        ).export(filters, export_format="xlsx")
    try:
        workbook = load_workbook(artifact.path, read_only=True, data_only=False)
        sheet = workbook["Findings"]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        assert rows[0][0:7] == (
            "Document Code",
            "Revision",
            "Department",
            "Finding Code",
            "Finding Type",
            "Severity",
            "Status",
        )
        assert len(rows) == 2
        assert rows[1][7].startswith("'=")
        assert rows[1][15] == "[redacted source reference]"
    finally:
        remove_finding_export_artifact(artifact.path)
    assert not artifact.path.exists()

    async with session_factory() as session:
        artifact = await FindingExportService(
            session,
            settings,
            controller,
            _METADATA,
        ).export(filters, export_format="json")
    try:
        payload = json.loads(artifact.path.read_text(encoding="utf-8"))
        assert payload["totalItems"] == 1
        assert payload["items"][0]["id"] == str(first.id)
        assert payload["items"][0]["sourceReference"] == "[redacted source reference]"
    finally:
        remove_finding_export_artifact(Path(artifact.path))

    limited = settings.model_copy(update={"finding_export_max_rows": 1})
    unscoped = FindingFilter(createdBySystem=False)
    async with session_factory() as session:
        with pytest.raises(ApplicationError) as too_large:
            await FindingExportService(
                session,
                limited,
                controller,
                _METADATA,
            ).export(unscoped, export_format="json")
        assert too_large.value.status_code == 413
        assert too_large.value.errors is not None
        assert too_large.value.errors[0].code == "FINDING_EXPORT_LIMIT_EXCEEDED"

    async with session_factory() as session:
        export_actions = list(
            (
                await session.scalars(
                    select(AuditLog.action).where(
                        AuditLog.action == AuditAction.EXPORT_FINDINGS
                    )
                )
            ).all()
        )
        assert export_actions == [
            AuditAction.EXPORT_FINDINGS,
            AuditAction.EXPORT_FINDINGS,
        ]
