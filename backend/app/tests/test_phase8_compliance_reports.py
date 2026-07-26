"""Focused latest-run, scope, permission, and export report coverage."""

from __future__ import annotations

from collections.abc import AsyncIterator, Callable
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from uuid import UUID, uuid4

import pytest
import pytest_asyncio
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.api.dependencies.auth import get_token_service
from app.api.v1.endpoints.compliance_reports import router
from app.core.authorization import UserRole
from app.core.exception_handlers import register_exception_handlers
from app.database.session import get_db_session
from app.models.compliance_enums import (
    ComplianceRunStatus,
    ComplianceStatus,
    FindingCode,
    FindingSeverity,
    FindingStatus,
    FindingType,
)
from app.models.compliance_run import ComplianceRun
from app.models.department import Department
from app.models.document import Document
from app.models.document_file import DocumentFile, DocumentFileStatus
from app.models.document_revision import DocumentRevision
from app.models.document_type import DocumentType
from app.models.section import Section
from app.models.validation_finding import ValidationFinding
from app.models.validation_rule import ValidationRule
from app.services.auth.token_service import TokenService

TestSessionFactory = async_sessionmaker[AsyncSession]
UserFactory = Callable[..., Any]


@pytest_asyncio.fixture
async def report_api_client(
    session_factory: TestSessionFactory,
    token_service: TokenService,
) -> AsyncIterator[AsyncClient]:
    async def override_session() -> AsyncIterator[AsyncSession]:
        async with session_factory() as session:
            try:
                yield session
            except Exception:
                await session.rollback()
                raise

    application = FastAPI()
    register_exception_handlers(application)
    application.include_router(router, prefix="/api/v1")
    application.dependency_overrides[get_db_session] = override_session
    application.dependency_overrides[get_token_service] = (
        lambda: token_service
    )
    async with AsyncClient(
        transport=ASGITransport(app=application),
        base_url="http://test",
    ) as client:
        yield client


async def _headers(
    create_user: UserFactory,
    token_service: TokenService,
    *,
    email: str,
    role: UserRole,
    department_id: UUID | None = None,
    is_superuser: bool = False,
) -> dict[str, str]:
    user = await create_user(
        email=email,
        role=role,
        department_id=department_id,
        is_superuser=is_superuser,
    )
    return {
        "Authorization": f"Bearer {token_service.create_access_token(user)}"
    }


async def _seed_report_data(
    session_factory: TestSessionFactory,
) -> tuple[UUID, UUID]:
    """Create one file with an obsolete and a latest official run."""

    async with session_factory() as session:
        department = Department(
            code="QMS",
            name="Quality Management",
        )
        other_department = Department(
            code="ICT",
            name="Technology",
        )
        document_type = DocumentType(
            code="SOP",
            name="Standard Operating Procedure",
        )
        session.add_all(
            [department, other_department, document_type]
        )
        await session.flush()
        section = Section(
            department_id=department.id,
            code="QA",
            name="Quality Assurance",
        )
        rule = ValidationRule(
            code="TRI",
            name="Trilingual Structure",
        )
        session.add_all([section, rule])
        await session.flush()
        document = Document(
            company_code="MTI",
            department_id=department.id,
            section_id=section.id,
            document_type_id=document_type.id,
            document_number="001",
            base_document_code="MTI-QMS-QA-SOP-001",
            title="=Unsafe report title",
        )
        session.add(document)
        await session.flush()
        revision = DocumentRevision(
            document_id=document.id,
            revision_code="Rev.001",
            revision_number=1,
            full_document_code="MTI-QMS-QA-SOP-001-Rev.001",
            document_status_id=uuid4(),
            validation_rule_id=rule.id,
            is_current=True,
        )
        session.add(revision)
        await session.flush()
        document.current_revision_id = revision.id
        document_file = DocumentFile(
            document_id=document.id,
            document_revision_id=revision.id,
            original_filename="procedure.docx",
            sanitized_filename="procedure.docx",
            file_extension="docx",
            mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "wordprocessingml.document"
            ),
            detected_mime_type=(
                "application/vnd.openxmlformats-officedocument."
                "wordprocessingml.document"
            ),
            file_size=1024,
            sha256_hash="a" * 64,
            storage_key="documents/originals/procedure.docx",
            file_status=DocumentFileStatus.AVAILABLE,
        )
        session.add(document_file)
        await session.flush()

        old_run = _run(
            document=document,
            revision=revision,
            document_file=document_file,
            rule=rule,
            status=ComplianceStatus.NON_COMPLIANT,
            score=42,
            source_hash="b" * 64,
            missing_languages=["en", "zh"],
            missing_sections=["PURPOSE", "SCOPE"],
            critical=1,
            major=1,
            completed_at=datetime(
                2026,
                6,
                1,
                1,
                tzinfo=timezone.utc,
            ),
        )
        latest_run = _run(
            document=document,
            revision=revision,
            document_file=document_file,
            rule=rule,
            status=ComplianceStatus.PARTIALLY_COMPLIANT,
            score=88,
            source_hash="c" * 64,
            missing_languages=["zh"],
            missing_sections=["PURPOSE"],
            critical=0,
            major=1,
            completed_at=datetime(
                2026,
                7,
                1,
                1,
                tzinfo=timezone.utc,
            ),
        )
        session.add_all([old_run, latest_run])
        await session.flush()
        document_file.latest_compliance_run_id = latest_run.id
        session.add_all(
            [
                _finding(
                    old_run,
                    severity=FindingSeverity.CRITICAL,
                    code=FindingCode.MISSING_CHINESE,
                    finding_type=FindingType.LANGUAGE_PRESENCE,
                    title="Obsolete missing Chinese",
                ),
                _finding(
                    latest_run,
                    severity=FindingSeverity.MAJOR,
                    code=FindingCode.MISSING_CHINESE,
                    finding_type=FindingType.LANGUAGE_PRESENCE,
                    title="Current missing Chinese",
                    status=FindingStatus.IN_REVIEW,
                ),
                ValidationFinding(
                    compliance_run_id=None,
                    document_id=document.id,
                    document_revision_id=revision.id,
                    document_file_id=document_file.id,
                    validation_rule_id=rule.id,
                    finding_code=FindingCode.MANUAL_FINDING,
                    finding_type=FindingType.MANUAL,
                    severity=FindingSeverity.CRITICAL,
                    status=FindingStatus.OPEN,
                    title="Manual control gap",
                    description="A reviewer recorded a manual gap.",
                    is_system_generated=False,
                ),
            ]
        )
        rule.code = "TRI-RENAMED"
        rule.name = "Renamed after validation"
        await session.commit()
        return department.id, other_department.id


def _run(
    *,
    document: Document,
    revision: DocumentRevision,
    document_file: DocumentFile,
    rule: ValidationRule,
    status: ComplianceStatus,
    score: int,
    source_hash: str,
    missing_languages: list[str],
    missing_sections: list[str],
    critical: int,
    major: int,
    completed_at: datetime,
) -> ComplianceRun:
    detected_languages = [
        language
        for language in ("id", "en", "zh")
        if language not in missing_languages
    ]
    return ComplianceRun(
        compliance_job_id=uuid4(),
        document_id=document.id,
        document_revision_id=revision.id,
        document_file_id=document_file.id,
        extraction_run_id=uuid4(),
        language_detection_run_id=uuid4(),
        validation_rule_id=rule.id,
        rule_snapshot_json={
            "ruleCode": rule.code,
            "ruleName": rule.name,
            "structuralValidationOnly": True,
        },
        source_content_hash=source_hash,
        status=ComplianceRunStatus.COMPLETED,
        compliance_status=status,
        compliance_score=score,
        maximum_score=100,
        document_code_score=10,
        language_presence_score=20,
        language_coverage_score=12,
        section_completeness_score=15,
        language_order_score=10,
        translation_group_score=15,
        table_completeness_score=5,
        total_findings=critical + major,
        critical_findings=critical,
        major_findings=major,
        minor_findings=0,
        information_findings=0,
        open_findings=critical + major,
        required_languages_json=["id", "en", "zh"],
        detected_languages_json=detected_languages,
        missing_languages_json=missing_languages,
        required_sections_json=["TITLE", "PURPOSE"],
        detected_sections_json=["TITLE"],
        missing_sections_json=missing_sections,
        warnings_json=[],
        metrics_json={
            "validators": {
                "LANGUAGE_PRESENCE": {
                    "metrics": {
                        "presence": {
                            language: (
                                "NOT_PRESENT"
                                if language in missing_languages
                                else "PRESENT"
                            )
                            for language in ("id", "en", "zh")
                        }
                    }
                },
                "REQUIRED_SECTIONS": {
                    "metrics": {
                        "completeSections": 1,
                        "totalRequiredSections": 2,
                    }
                },
                "LANGUAGE_ORDER": {
                    "metrics": {
                        "evaluatedGroups": 2,
                        "invalidGroups": 0,
                    }
                },
            },
            "structuralValidationOnly": True,
            "semanticSimilarityEvaluated": False,
        },
        completed_at=completed_at,
    )


def _finding(
    run: ComplianceRun,
    *,
    severity: FindingSeverity,
    code: FindingCode,
    finding_type: FindingType,
    title: str,
    status: FindingStatus = FindingStatus.OPEN,
) -> ValidationFinding:
    return ValidationFinding(
        compliance_run_id=run.id,
        document_id=run.document_id,
        document_revision_id=run.document_revision_id,
        document_file_id=run.document_file_id,
        validation_rule_id=run.validation_rule_id,
        finding_code=code,
        finding_type=finding_type,
        severity=severity,
        status=status,
        title=title,
        description=f"{title}.",
        is_system_generated=True,
    )


@pytest.mark.asyncio
async def test_reports_use_latest_official_run_and_keep_manual_findings(
    report_api_client: AsyncClient,
    session_factory: TestSessionFactory,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    department_id, _ = await _seed_report_data(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        email="auditor-report@example.com",
        role=UserRole.AUDITOR,
    )

    overview = await report_api_client.get(
        "/api/v1/compliance/overview",
        headers=headers,
    )
    assert overview.status_code == 200, overview.text
    overview_data = overview.json()["data"]
    assert overview_data["totalValidatedDocuments"] == 1
    assert overview_data["partiallyCompliant"] == 1
    assert overview_data["nonCompliant"] == 0
    assert overview_data["openCriticalFindings"] == 0
    assert overview_data["openMajorFindings"] == 1
    assert overview_data["missingLanguages"] == [
        {"languageCode": "zh", "count": 1}
    ]

    compliance = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=headers,
        params={"departmentId": str(department_id)},
    )
    assert compliance.status_code == 200, compliance.text
    compliance_data = compliance.json()["data"]
    assert compliance_data["totalItems"] == 1
    item = compliance_data["items"][0]
    assert item["score"] == 88.0
    assert item["languagePresence"]["zh"] == "NOT_PRESENT"
    assert item["sectionCompleteness"] == 50.0
    assert item["languageOrderValid"] is True
    assert item["validationRule"] == "Trilingual Structure"

    findings = await report_api_client.get(
        "/api/v1/reports/findings",
        headers=headers,
        params={"page": 1, "pageSize": 20},
    )
    assert findings.status_code == 200, findings.text
    findings_data = findings.json()["data"]
    assert findings_data["summary"]["totalFindings"] == 2
    assert findings_data["summary"]["critical"] == 1
    assert findings_data["summary"]["major"] == 1
    assert findings_data["findings"]["totalItems"] == 2
    assert {
        row["title"] for row in findings_data["findings"]["items"]
    } == {"Current missing Chinese", "Manual control gap"}
    assert {
        row["document"]["baseDocumentCode"]
        for row in findings_data["findings"]["items"]
    } == {"MTI-QMS-QA-SOP-001"}


@pytest.mark.asyncio
async def test_report_permissions_scope_and_formula_safe_export(
    report_api_client: AsyncClient,
    session_factory: TestSessionFactory,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    department_id, other_department_id = await _seed_report_data(
        session_factory
    )
    reviewer_headers = await _headers(
        create_user,
        token_service,
        email="reviewer-report@example.com",
        role=UserRole.REVIEWER,
        department_id=department_id,
    )
    forbidden_scope = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=reviewer_headers,
        params={"departmentId": str(other_department_id)},
    )
    assert forbidden_scope.status_code == 403

    forbidden_export = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=reviewer_headers,
        params={"format": "xlsx"},
    )
    assert forbidden_export.status_code == 403

    viewer_headers = await _headers(
        create_user,
        token_service,
        email="viewer-report@example.com",
        role=UserRole.VIEWER,
        department_id=department_id,
    )
    allowed_overview = await report_api_client.get(
        "/api/v1/compliance/overview",
        headers=viewer_headers,
    )
    assert allowed_overview.status_code == 200
    forbidden_report = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=viewer_headers,
    )
    assert forbidden_report.status_code == 403

    auditor_headers = await _headers(
        create_user,
        token_service,
        email="exporter-report@example.com",
        role=UserRole.AUDITOR,
    )
    exported = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=auditor_headers,
        params={"format": "xlsx"},
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    workbook = load_workbook(
        BytesIO(exported.content),
        read_only=True,
        data_only=False,
    )
    try:
        rows = list(workbook["Compliance"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[1][1] == "'=Unsafe report title"
    assert rows[1][-1] is False

    exported_json = await report_api_client.get(
        "/api/v1/reports/compliance",
        headers=auditor_headers,
        params={"format": "json"},
    )
    assert exported_json.status_code == 200, exported_json.text
    payload = exported_json.json()
    assert payload["latestOfficialRunsOnly"] is True
    assert payload["structuralValidationOnly"] is True
    assert payload["semanticSimilarityEvaluated"] is False
    assert payload["totalItems"] == 1
    assert payload["items"][0]["score"] == 88.0
