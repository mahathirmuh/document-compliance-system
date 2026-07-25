"""Tests for explicit, local-only Phase 7 model setup helpers."""

from __future__ import annotations

import hashlib
from pathlib import Path

import pymupdf
import pytest
from docx import Document
from openpyxl import load_workbook

from scripts.download_language_model import (
    download_model,
    validate_existing,
)
from scripts.download_ocr_models import (
    model_directory_ready,
    normalized_checksum,
)
from scripts.generate_phase7_sample_documents import generate


def test_existing_language_model_is_reused_without_network(
    tmp_path: Path,
) -> None:
    destination = tmp_path / "lid.176.bin"
    content = b"generated-test-model"
    destination.write_bytes(content)
    checksum = hashlib.sha256(content).hexdigest()

    result = download_model(
        destination=destination,
        source_url="https://example.invalid/lid.176.bin",
        expected_sha256=checksum,
    )

    assert result == destination.resolve()
    assert validate_existing(destination, checksum)


@pytest.mark.parametrize(
    ("source_url", "checksum"),
    [
        ("http://example.com/lid.176.bin", None),
        ("not-a-url", None),
        ("https://example.com/lid.176.bin", "invalid"),
    ],
)
def test_language_model_setup_rejects_unsafe_configuration(
    tmp_path: Path,
    source_url: str,
    checksum: str | None,
) -> None:
    with pytest.raises(ValueError):
        download_model(
            destination=tmp_path / "lid.176.bin",
            source_url=source_url,
            expected_sha256=checksum,
        )


def test_ocr_model_directory_requires_a_real_model_file(
    tmp_path: Path,
) -> None:
    profile = tmp_path / "latin" / "recognition"
    profile.mkdir(parents=True)
    (profile / ".gitkeep").touch()
    assert not model_directory_ready(profile)

    (profile / "inference.json").write_text("{}", encoding="utf-8")
    assert model_directory_ready(profile)


def test_ocr_model_checksum_validation() -> None:
    checksum = "a" * 64
    assert normalized_checksum(checksum.upper()) == checksum
    assert normalized_checksum("") is None
    with pytest.raises(ValueError):
        normalized_checksum("not-a-checksum")


def test_generated_phase7_sample_document_inventory(
    tmp_path: Path,
) -> None:
    outputs = generate(tmp_path / "sample-documents")
    assert len(outputs) == 14
    assert all(path.is_file() and path.stat().st_size > 0 for path in outputs)

    ocr_directory = tmp_path / "sample-documents" / "ocr"
    expected_ocr_names = {
        "scanned-indonesian.pdf",
        "scanned-english.pdf",
        "scanned-chinese.pdf",
        "scanned-mixed.pdf",
        "rotated.pdf",
        "low-resolution.pdf",
        "partial-scan.pdf",
        "blank-scan.pdf",
    }
    assert {path.name for path in ocr_directory.iterdir()} == expected_ocr_names
    for name in expected_ocr_names - {"partial-scan.pdf"}:
        with pymupdf.open(ocr_directory / name) as document:
            assert all(not page.get_text().strip() for page in document)
    with pymupdf.open(ocr_directory / "partial-scan.pdf") as partial:
        assert len(partial) == 2
        assert len(partial[0].get_text().strip()) >= 50
        assert not partial[1].get_text().strip()

    language_directory = tmp_path / "sample-documents" / "language"
    expected_language_names = {
        "indonesian.docx",
        "english.docx",
        "chinese.docx",
        "three-language.docx",
        "mixed-language.xlsx",
        "short-text.xlsx",
    }
    assert {
        path.name for path in language_directory.iterdir()
    } == expected_language_names
    chinese_document = Document(language_directory / "chinese.docx")
    assert any("中文" in paragraph.text for paragraph in chinese_document.paragraphs)
    mixed_workbook = load_workbook(
        language_directory / "mixed-language.xlsx",
        read_only=True,
    )
    try:
        values = [
            str(cell)
            for row in mixed_workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
        assert any("质量" in value for value in values)
    finally:
        mixed_workbook.close()
