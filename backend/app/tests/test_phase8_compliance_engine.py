"""Focused acceptance tests for the pure Phase 8 compliance engine."""

from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace
from uuid import UUID, uuid4

import pytest
from celery.exceptions import SoftTimeLimitExceeded
from openpyxl import load_workbook

from app.schemas.compliance import (
    ComplianceScoreBreakdownResponse,
    ComplianceScoreComponent,
    ComplianceScorePenalties,
)
from app.schemas.compliance_internal import (
    ComplianceBlockData,
    ComplianceTableCellData,
    ComplianceTableData,
    SectionAliasData,
    TranslationGroupData,
    TranslationGroupMemberData,
    ValidationRuleSnapshot,
    ValidatorResult,
)
from app.services.compliance.compliance_comparison_service import (
    ComplianceComparisonService,
)
from app.services.compliance.compliance_context_service import (
    ComplianceContextBuildError,
    ComplianceContextService,
)
from app.services.compliance.compliance_export_service import (
    ComplianceExportService,
    safe_source_reference,
    spreadsheet_safe_value,
)
from app.services.compliance.compliance_pipeline import (
    COMPLIANCE_GROUPING_FAILED,
    COMPLIANCE_SECTION_DETECTION_FAILED,
    COMPLIANCE_VALIDATION_FAILED,
    CompliancePipeline,
    CompliancePipelineCancelled,
    CompliancePipelineStageError,
)
from app.services.compliance.compliance_result_export_service import (
    _write_workbook,
)
from app.services.compliance.compliance_score_service import (
    ComplianceScoreService,
    ComplianceWeightError,
)
from app.services.compliance.compliance_status_service import (
    ComplianceStatusService,
)
from app.services.compliance.findings.finding_deduplication_service import (
    FindingDeduplicationService,
)
from app.services.compliance.findings.finding_factory import FindingFactory
from app.services.compliance.findings.finding_resolution_service import (
    FindingResolutionService,
    FindingTransitionError,
)
from app.services.compliance.grouping._group_builder import member_from_block
from app.services.compliance.grouping.paragraph_grouping_service import (
    ParagraphGroupingService,
)
from app.services.compliance.grouping.positional_grouping_service import (
    PositionalGroupingService,
)
from app.services.compliance.grouping.table_grouping_service import (
    TableGroupingService,
)
from app.services.compliance.sections.heading_candidate_service import (
    HeadingCandidateService,
)
from app.services.compliance.sections.section_alias_service import (
    DuplicateSectionAliasError,
    SectionAliasService,
    UnsafeSectionAliasRegexError,
    normalise_heading,
)
from app.services.compliance.sections.section_boundary_service import (
    SectionBoundaryService,
)
from app.services.compliance.sections.section_matcher import SectionMatcher
from app.services.compliance.validators.container_completeness_validator import (
    ContainerCompletenessValidator,
)
from app.services.compliance.validators.language_coverage_validator import (
    LanguageCoverageValidator,
)
from app.services.compliance.validators.language_order_validator import (
    LanguageOrderValidator,
)
from app.services.compliance.validators.language_presence_validator import (
    LanguagePresenceValidator,
)
from app.services.compliance.validators.table_multilingual_validator import (
    TableMultilingualValidator,
)
from app.services.compliance.validators.translation_group_validator import (
    TranslationGroupValidator,
)


def _rule(**updates: object) -> ValidationRuleSnapshot:
    base = ValidationRuleSnapshot(
        rule_code="TEST-3LANG",
        required_languages=["id", "en", "zh"],
        required_sections=[],
        language_order=["id", "en", "zh"],
        validation_options={
            "presenceMinimumBlocks": 1,
            "presenceMinimumCharacters": 10,
        },
    )
    return base.model_copy(update=updates, deep=True)


def _block(
    text: str,
    language: str,
    order: int,
    *,
    container_id: UUID | None = None,
    container_type: str = "DOCX_BODY",
    block_type: str = "PARAGRAPH",
    confidence: float = 0.95,
    location: dict[str, object] | None = None,
    metadata: dict[str, object] | None = None,
) -> ComplianceBlockData:
    return ComplianceBlockData(
        id=uuid4(),
        container_id=container_id or _block.default_container,
        container_type=container_type,
        container_name="Container 1",
        container_index=1,
        block_order=order,
        block_type=block_type,
        source_reference=f"{container_type}:block={order}",
        text=text,
        normalised_text=text,
        heading_level=1 if block_type == "HEADING" else None,
        page_number=1 if container_type == "PDF_PAGE" else None,
        language_code=language,
        language_confidence=confidence,
        character_count=len(text),
        location=location or {},
        metadata=metadata or {},
    )


_block.default_container = uuid4()  # type: ignore[attr-defined]


def _context(
    blocks: list[ComplianceBlockData],
    *,
    rule: ValidationRuleSnapshot | None = None,
    source_format: str = "DOCX",
    tables: list[ComplianceTableData] | None = None,
    aliases: list[SectionAliasData] | None = None,
):
    return ComplianceContextService().build(
        rule=rule or _rule(),
        source_format=source_format,
        blocks=blocks,
        tables=tables or [],
        section_aliases=aliases or [],
        document_code="DOC-001",
        expected_document_code="DOC-001",
    )


@pytest.mark.parametrize(
    ("heading", "expected"),
    [
        ("1. TUJUAN:", "tujuan"),
        ("2.3 Purpose：", "purpose"),
        ("III. 目的：", "目的"),
    ],
)
def test_heading_normalisation_preserves_language_text(
    heading: str,
    expected: str,
) -> None:
    assert normalise_heading(heading) == expected


@pytest.mark.parametrize(
    ("heading", "language", "alias"),
    [
        ("1. TUJUAN:", "id", "Tujuan"),
        ("PURPOSE", "en", "Purpose"),
        ("目的：", "zh", "目的"),
    ],
)
def test_section_matcher_exact_id_en_zh(
    heading: str,
    language: str,
    alias: str,
) -> None:
    match = SectionMatcher().match(
        heading,
        [
            SectionAliasData(
                canonical_code="PURPOSE",
                language_code=language,
                alias_text=alias,
                match_type="EXACT",
            ),
        ],
    )

    assert match is not None
    assert match.canonical_code == "PURPOSE"
    assert match.language_code == language
    assert match.confidence == 1


def test_section_match_strategy_prefix_regex_contains_and_fuzzy() -> None:
    aliases = [
        SectionAliasData(
            canonical_code="PURPOSE",
            language_code="en",
            alias_text="Purpose",
            match_type="PREFIX",
        ),
        SectionAliasData(
            canonical_code="RESPONSIBILITY",
            language_code="en",
            alias_text=r"^Responsibilities$",
            match_type="REGEX",
            is_regex=True,
        ),
        SectionAliasData(
            canonical_code="PURPOSE",
            language_code="zh",
            alias_text="目的",
            match_type="CONTAINS",
        ),
        SectionAliasData(
            canonical_code="SCOPE",
            language_code="en",
            alias_text="Scope",
            match_type="FUZZY",
        ),
    ]

    assert SectionMatcher().match("Purpose - operations", aliases).match_type == (
        "PREFIX"
    )
    assert SectionMatcher().match("Responsibilities", aliases).match_type == ("REGEX")
    assert SectionMatcher().match("2. Responsibilities:", aliases).match_type == (
        "REGEX"
    )
    assert SectionMatcher().match("文件目的", aliases).match_type == "CONTAINS"
    assert SectionMatcher().match("Scop", aliases).match_type == "FUZZY"
    assert SectionMatcher().match("Unrelated", aliases) is None


def test_alias_duplicate_inactive_profile_and_regex_safety() -> None:
    duplicate = SectionAliasData(
        canonical_code="PURPOSE",
        language_code="en",
        alias_text="Purpose",
        match_type="EXACT",
    )
    with pytest.raises(DuplicateSectionAliasError):
        SectionAliasService().validate_unique([duplicate, duplicate])
    inactive = duplicate.model_copy(update={"is_active": False})
    assert SectionMatcher().match("Purpose", [inactive]) is None
    with pytest.raises(UnsafeSectionAliasRegexError):
        SectionAliasService().validate_regex(r"(a+)+$")


def test_heading_candidates_and_multilingual_section_boundary() -> None:
    container_id = uuid4()
    blocks = [
        _block(
            "1. TUJUAN",
            "id",
            1,
            container_id=container_id,
            block_type="HEADING",
        ),
        _block(
            "1. PURPOSE",
            "en",
            2,
            container_id=container_id,
            block_type="HEADING",
        ),
        _block("Isi tujuan yang cukup.", "id", 3, container_id=container_id),
        _block(
            "Purpose content is here.",
            "en",
            4,
            container_id=container_id,
        ),
        _block(
            "2. SCOPE",
            "en",
            5,
            container_id=container_id,
            block_type="HEADING",
        ),
        _block("Scope content.", "en", 6, container_id=container_id),
    ]
    aliases = [
        SectionAliasData(
            canonical_code="PURPOSE",
            language_code="id",
            alias_text="Tujuan",
            match_type="EXACT",
        ),
        SectionAliasData(
            canonical_code="PURPOSE",
            language_code="en",
            alias_text="Purpose",
            match_type="EXACT",
        ),
        SectionAliasData(
            canonical_code="SCOPE",
            language_code="en",
            alias_text="Scope",
            match_type="EXACT",
        ),
    ]
    candidates = HeadingCandidateService().detect(
        blocks,
        alias_texts=[alias.alias_text for alias in aliases],
    )
    matches = SectionMatcher().match_all(candidates, aliases)
    sections = SectionBoundaryService().build(matches, blocks)

    assert [item.canonical_code for item in sections] == ["PURPOSE", "SCOPE"]
    assert sections[0].heading_text == "1. TUJUAN / 1. PURPOSE"
    assert sections[0].start_block_order == 3
    assert sections[0].end_block_order == 4


def test_docx_and_pdf_structural_grouping_order() -> None:
    docx = [
        _block("Teks bahasa Indonesia panjang.", "id", 1),
        _block("A sufficiently long English text.", "en", 2),
        _block("足够长的中文文本内容。", "zh", 3),
    ]
    docx_groups = ParagraphGroupingService().group(
        docx,
        ["id", "en", "zh"],
    )
    assert len(docx_groups) == 1
    assert docx_groups[0].is_complete
    assert docx_groups[0].is_order_valid

    pdf = [
        _block(
            "English PDF paragraph.",
            "en",
            1,
            container_type="PDF_PAGE",
            location={"page": 1, "bbox": [10, 10, 200, 30]},
        ),
        _block(
            "Paragraf PDF Indonesia.",
            "id",
            2,
            container_type="PDF_PAGE",
            location={"page": 1, "bbox": [10, 35, 200, 55]},
        ),
        _block(
            "PDF 中文段落。",
            "zh",
            3,
            container_type="PDF_PAGE",
            location={"page": 1, "bbox": [10, 60, 200, 80]},
        ),
    ]
    pdf_group = PositionalGroupingService().group(
        pdf,
        ["id", "en", "zh"],
    )[0]
    assert pdf_group.is_complete
    assert not pdf_group.is_order_valid
    assert pdf_group.metrics["semanticSimilarityEvaluated"] is False


def _table(
    *,
    rows_layout: bool = False,
    missing_chinese: bool = False,
) -> ComplianceTableData:
    if rows_layout:
        values = [
            (0, 0, "Indonesia", "unknown"),
            (0, 1, "Teks Indonesia panjang", "id"),
            (1, 0, "English", "unknown"),
            (1, 1, "Long English text", "en"),
            (2, 0, "Chinese", "unknown"),
            (2, 1, "" if missing_chinese else "中文内容", "zh"),
        ]
    else:
        values = [
            (0, 0, "Indonesia", "unknown"),
            (0, 1, "English", "unknown"),
            (0, 2, "Chinese", "unknown"),
            (1, 0, "Teks Indonesia panjang", "id"),
            (1, 1, "Long English text", "en"),
            (1, 2, "" if missing_chinese else "中文内容", "zh"),
        ]
    cells = [
        ComplianceTableCellData(
            id=uuid4(),
            row_index=row,
            column_index=column,
            coordinate=f"{chr(65 + column)}{row + 1}",
            text=text,
            normalised_text=text,
            language_code=language,
            language_confidence=0.95 if language != "unknown" else 0,
        )
        for row, column, text, language in values
    ]
    return ComplianceTableData(
        id=uuid4(),
        container_id=_block.default_container,  # type: ignore[attr-defined]
        source_reference="XLSX:sheet=Register:table=1",
        table_index=1,
        row_count=3 if rows_layout else 2,
        column_count=2 if rows_layout else 3,
        cells=cells,
    )


@pytest.mark.parametrize(
    ("rows_layout", "expected_layout"),
    [
        (False, "LANGUAGES_AS_COLUMNS"),
        (True, "LANGUAGES_AS_ROWS"),
    ],
)
def test_xlsx_columns_rows_and_incomplete_cells(
    rows_layout: bool,
    expected_layout: str,
) -> None:
    service = TableGroupingService()
    complete = _table(rows_layout=rows_layout)
    layout = service.detect_layout(complete)
    groups = service.group(complete, ["id", "en", "zh"])
    incomplete = service.group(
        _table(rows_layout=rows_layout, missing_chinese=True),
        ["id", "en", "zh"],
    )

    assert layout.layout == expected_layout
    assert groups and groups[0].is_complete
    assert incomplete and not incomplete[0].is_complete
    assert incomplete[0].metrics["missingCells"]["zh"]


@pytest.mark.asyncio
async def test_language_presence_coverage_and_low_confidence() -> None:
    blocks = [
        _block("Teks Indonesia yang cukup.", "id", 1),
        _block("English text with evidence.", "en", 2),
        _block("中文语言内容具有充分证据。", "zh", 3),
        _block("??? unknown linguistic block", "unknown", 4),
        _block("Mixed 中文 and text", "mixed", 5),
    ]
    rule = _rule(
        minimum_language_block_coverage={"id": 10, "en": 10, "zh": 10},
        minimum_language_character_coverage={"id": 10, "en": 10, "zh": 10},
        maximum_unknown_block_percentage=10,
        maximum_mixed_block_percentage=10,
    )
    context = _context(blocks, rule=rule)
    presence = await LanguagePresenceValidator().validate(context)
    coverage = await LanguageCoverageValidator().validate(context)

    assert presence.score == presence.maximum_score
    assert coverage.metrics["blockCoverage"]["unknown"] == 20
    assert {finding.finding_code for finding in coverage.findings} >= {
        "UNKNOWN_TEXT_EXCEEDS_THRESHOLD",
        "MIXED_TEXT_EXCEEDS_THRESHOLD",
    }

    low_confidence = _context(
        [
            _block(
                "Only weak Chinese evidence here 中文",
                "zh",
                1,
                confidence=0.2,
            ),
            _block("Strong Indonesian evidence.", "id", 2),
            _block("Strong English evidence.", "en", 3),
        ],
        rule=_rule(),
    )
    low_result = await LanguagePresenceValidator().validate(low_confidence)
    assert "MISSING_CHINESE" in {
        finding.finding_code for finding in low_result.findings
    }


@pytest.mark.asyncio
async def test_container_section_and_order_validators() -> None:
    container_id = uuid4()
    blocks = [
        _block(
            "PURPOSE",
            "en",
            1,
            container_id=container_id,
            block_type="HEADING",
        ),
        _block(
            "Tujuan dengan isi cukup.",
            "id",
            2,
            container_id=container_id,
        ),
        _block(
            "Purpose with enough content.",
            "en",
            3,
            container_id=container_id,
        ),
    ]
    alias = SectionAliasData(
        canonical_code="PURPOSE",
        language_code="en",
        alias_text="Purpose",
        match_type="EXACT",
    )
    rule = _rule(
        validate_sections=True,
        validate_container_completeness=True,
        required_sections=["PURPOSE", "SCOPE", "RESPONSIBILITIES"],
        validation_options={
            "presenceMinimumBlocks": 1,
            "presenceMinimumCharacters": 5,
            "requireAllLanguagesPerContainer": True,
            "ignoreContainersBelowCharacterCount": 1,
        },
    )
    context = _context(blocks, rule=rule, aliases=[alias])
    analysed = await CompliancePipeline().run(context)
    required = next(
        result
        for result in analysed.validator_results
        if result.validator_code == "REQUIRED_SECTIONS"
    )
    container = await ContainerCompletenessValidator().validate(
        analysed.context,
    )

    assert "MISSING_REQUIRED_SECTION" in {
        finding.finding_code for finding in required.findings
    }
    missing_required_sections = {
        finding.detected_section_code
        for finding in analysed.findings
        if finding.finding_code == "MISSING_REQUIRED_SECTION"
    }
    assert missing_required_sections == {"SCOPE", "RESPONSIBILITIES"}
    assert "MISSING_SECTION_CHINESE" in {
        finding.finding_code for finding in required.findings
    }
    assert container.metrics["evaluatedContainers"] == 1
    assert any(finding.language_code == "zh" for finding in container.findings)


def _group(
    languages: list[str],
    *,
    confidence: float = 0.95,
    group_type: str = "PARAGRAPH_GROUP",
) -> TranslationGroupData:
    members = [
        TranslationGroupMemberData(
            block_id=uuid4(),
            extracted_block_id=uuid4(),
            language_code=language,
            block_order=index,
            text_snapshot=f"{language} content",
            confidence=confidence,
            source_reference=f"DOCX:p={index}",
            source_type="NATIVE_EXTRACTION",
        )
        for index, language in enumerate(languages, 1)
    ]
    return TranslationGroupData(
        container_id=_block.default_container,  # type: ignore[attr-defined]
        group_index=0,
        group_type=group_type,
        source_reference="DOCX:group=0",
        expected_languages=["id", "en", "zh"],
        detected_languages=list(dict.fromkeys(languages)),
        language_order=languages,
        members=members,
        is_complete=set(languages) >= {"id", "en", "zh"},
        is_order_valid=languages == ["id", "en", "zh"],
        confidence=confidence,
    )


@pytest.mark.asyncio
async def test_language_order_group_completeness_and_low_confidence() -> None:
    base = _context([], rule=_rule(validate_translation_groups=True))
    context = base.model_copy(
        update={
            "translation_groups": [
                _group(["en", "id", "zh"]),
                _group(["id", "en"], confidence=0.5),
            ],
        },
    )
    order = await LanguageOrderValidator().validate(context)
    completeness = await TranslationGroupValidator().validate(context)

    assert order.metrics["invalidGroups"] == 1
    assert order.metrics["lowConfidenceGroups"] == 1
    assert any(
        finding.severity == "INFORMATION"
        for finding in completeness.findings
        if finding.metrics.get("lowConfidence")
    )
    assert not any(
        finding.severity in {"CRITICAL", "MAJOR"}
        for finding in completeness.findings
        if finding.metrics.get("lowConfidence")
    )


@pytest.mark.asyncio
async def test_table_validator_reports_missing_xlsx_language() -> None:
    table = _table(missing_chinese=True)
    context = _context(
        [],
        rule=_rule(validate_tables=True),
        source_format="XLSX",
        tables=[table],
    )
    result = await TableMultilingualValidator().validate(context)

    assert result.status == "FAILED"
    assert result.findings[0].finding_code == ("XLSX_ROW_TRANSLATION_INCOMPLETE")


@pytest.mark.asyncio
async def test_table_validator_skips_non_applicable_text_only_document() -> None:
    context = _context(
        [
            _block(
                "This controlled procedure applies to every department.",
                "en",
                1,
            )
        ],
        rule=_rule(validate_tables=True),
        source_format="DOCX",
        tables=[],
    )

    table_result = await TableMultilingualValidator().validate(context)
    decision = ComplianceStatusService().determine(
        100,
        [],
        context=context,
        rule=context.rule,
        validator_results=[table_result],
    )

    assert table_result.status == "SKIPPED"
    assert table_result.score == table_result.maximum_score == 5
    assert table_result.metrics == {
        "enabled": True,
        "applicable": False,
        "evaluatedTableGroups": 0,
        "completeTableGroups": 0,
        "incompleteTableGroups": 0,
        "invalidHeaderOrders": 0,
        "layouts": [],
    }
    assert decision.status == "COMPLIANT"


def test_scoring_penalties_cap_and_status_precedence() -> None:
    results = [
        ValidatorResult(
            validator_code="ALL",
            status="PASSED",
            score=100,
            maximum_score=100,
        ),
    ]
    factory = FindingFactory()
    findings = [
        factory.create("MISSING_CHINESE", severity="CRITICAL"),
        factory.create("LOW_ENGLISH_COVERAGE", severity="MAJOR"),
        factory.create("LANGUAGE_ORDER_INVALID", severity="MINOR"),
    ]
    score = ComplianceScoreService().calculate(results, findings, _rule())

    assert score.weighted_score == 100
    assert score.score_before_cap == 94
    assert score.score_cap == 69
    assert score.final_score == 69
    decision = ComplianceStatusService().determine(
        score,
        findings,
        rule=_rule(),
    )
    assert decision.status == "NON_COMPLIANT"

    review = ComplianceStatusService().determine(
        88,
        [],
        rule=_rule(needs_review_score=50),
        metrics={"ocrConfidenceTooLow": True},
    )
    assert review.status == "NEEDS_REVIEW"
    not_evaluated = ComplianceStatusService().determine(
        100,
        [],
        context=SimpleNamespace(
            prerequisites={"extractionAvailable": False},
            rule=_rule(),
        ),
    )
    assert not_evaluated.status == "NOT_EVALUATED"


def test_weight_total_rejected_and_perfect_pipeline_maximum_is_100() -> None:
    invalid = _rule(document_code_weight=11)
    with pytest.raises(ComplianceWeightError):
        ComplianceScoreService().validate_weights(invalid)
    with pytest.raises(ComplianceWeightError):
        ComplianceScoreService().calculate(
            [
                ValidatorResult(
                    validator_code="OVER",
                    status="PASSED",
                    score=101,
                    maximum_score=101,
                ),
            ],
            rule=_rule(),
        )


def test_finding_dedup_repeat_manual_and_workflow() -> None:
    factory = FindingFactory()
    duplicate = factory.create(
        "MISSING_CHINESE",
        source_reference="DOCX:p=1",
        language_code="zh",
    )
    dedup = FindingDeduplicationService()
    retained = dedup.deduplicate([duplicate, duplicate])
    assert len(retained) == 1
    assert retained[0].metrics["occurrenceCount"] == 2

    previous_id = uuid4()
    previous = duplicate.model_copy(update={"id": previous_id})
    repeated = dedup.link_repeated([duplicate], [previous])[0]
    assert repeated.is_repeat
    assert repeated.previous_finding_id == previous_id

    manual = factory.manual(
        severity="MINOR",
        title="<b>Manual</b>",
        description="Review this.",
        recommendation="Check.",
    )
    merged = dedup.merge_revalidation([duplicate], [manual])
    assert any(not item.is_system_generated for item in merged)
    assert manual.title == "Manual"

    workflow = FindingResolutionService()
    reviewed = workflow.review(
        duplicate,
        actor_id=uuid4(),
        comment="Confirmed.",
    )
    assert reviewed.status == "IN_REVIEW"
    resolved = workflow.resolve(
        reviewed,
        actor_id=uuid4(),
        comment="Corrected.",
    )
    reopened = workflow.reopen(
        resolved,
        actor_id=uuid4(),
        reason="Reappeared.",
    )
    assert reopened.status == "REOPENED"
    with pytest.raises(FindingTransitionError):
        workflow.resolve(duplicate, actor_id=uuid4(), comment="")


def test_comparison_and_export_helpers_are_non_mutating_and_safe() -> None:
    factory = FindingFactory()
    previous_finding = factory.create(
        "MISSING_CHINESE",
        source_reference="DOCX:p=1",
    )
    current_finding = factory.create(
        "LANGUAGE_ORDER_INVALID",
        source_reference="DOCX:p=2",
    )
    comparison = ComplianceComparisonService().compare(
        {
            "complianceScore": 70,
            "complianceStatus": "NON_COMPLIANT",
            "detectedLanguages": ["id", "en"],
            "detectedSections": ["PURPOSE"],
            "findings": [previous_finding],
            "translationGroups": {"total": 3, "complete": 1},
        },
        {
            "complianceScore": 85,
            "complianceStatus": "PARTIALLY_COMPLIANT",
            "detectedLanguages": ["id", "en", "zh"],
            "detectedSections": ["PURPOSE", "SCOPE"],
            "findings": [current_finding],
            "translationGroups": {"total": 3, "complete": 2},
        },
    )
    assert comparison.score_change == 15
    assert comparison.languages_added == ("zh",)
    assert len(comparison.new_findings) == 1
    assert len(comparison.not_reproduced_findings) == 1
    assert previous_finding.status == "OPEN"

    assert spreadsheet_safe_value("=1+1") == "'=1+1"
    assert safe_source_reference(r"C:\secret\file.pdf") == (
        "[redacted source reference]"
    )
    assert safe_source_reference("documents/originals/private.pdf") == (
        "[redacted source reference]"
    )
    assert (
        safe_source_reference("PDF:documents/originals/private.pdf")
        == "[redacted source reference]"
    )
    export = ComplianceExportService().json_payload(
        {"status": "COMPLETED", "storagePath": r"C:\secret"},
        findings=[current_finding],
    )
    assert "storagePath" not in export["summary"]


def test_compliance_export_writes_required_xlsx_and_redacts_all_sources(
    tmp_path: Path,
) -> None:
    windows_group_path = r"C:\private\translation-group.pdf"
    unix_member_path = "/srv/compliance/private/member.txt"
    prefixed_windows_finding_path = r"PDF:C:\private\finding.pdf"
    relative_storage_path = "documents/originals/private.pdf"
    safe_reference = "PDF:page=2:block=4"
    run = {
        "document": {
            "baseDocumentCode": "MTI-HRM-POL-001",
        },
        "revision": {
            "revisionCode": "Rev.003",
            "fullDocumentCode": "MTI-HRM-POL-001_Rev.003",
        },
        "validationRule": {
            "code": "RENAMED-LIVE-RULE",
            "name": "Renamed Live Rule",
        },
        "ruleSnapshot": {
            "ruleCode": "TRILINGUAL-STRICT",
            "ruleName": "Trilingual Strict",
        },
        "complianceStatus": "PARTIALLY_COMPLIANT",
        "complianceScore": 82.5,
        "completedAt": "2026-07-26T10:15:00+08:00",
        "requiredLanguages": ["id", "en", "zh"],
        "missingLanguages": ["zh"],
        "requiredSections": ["PURPOSE", "SCOPE"],
        "missingSections": ["SCOPE"],
        "totalFindings": 3,
        "storagePath": r"C:\private\run.json",
    }
    score = ComplianceScoreBreakdownResponse(
        documentCode=ComplianceScoreComponent(earned=10, maximum=10),
        languagePresence=ComplianceScoreComponent(earned=20, maximum=25),
        languageCoverage=ComplianceScoreComponent(earned=12, maximum=15),
        sectionCompleteness=ComplianceScoreComponent(earned=16, maximum=20),
        languageOrder=ComplianceScoreComponent(earned=8, maximum=10),
        translationGroups=ComplianceScoreComponent(earned=12, maximum=15),
        tableCompleteness=ComplianceScoreComponent(earned=4, maximum=5),
        penalties=ComplianceScorePenalties(major=-1, minor=-0.5),
        weightedScore=82,
        scoreCap=None,
        finalScore=80.5,
    )
    translation_groups = [
        {
            "id": str(uuid4()),
            "sourceReference": windows_group_path,
            "members": [
                {
                    "languageCode": "id",
                    "sourceReference": unix_member_path,
                    "position": {
                        "source_reference": "file:///srv/private/member.txt",
                        "page": 1,
                    },
                    "storagePath": "/srv/private/source.bin",
                }
            ],
        }
    ]
    findings = [
        {
            "id": str(uuid4()),
            "findingCode": "MISSING_CHINESE",
            "sourceReference": prefixed_windows_finding_path,
        },
        {
            "id": str(uuid4()),
            "findingCode": "LANGUAGE_ORDER_INVALID",
            "sourceReference": safe_reference,
        },
        {
            "id": str(uuid4()),
            "findingCode": "MISSING_REQUIRED_SECTION",
            "sourceReference": relative_storage_path,
        },
    ]
    exporter = ComplianceExportService()

    payload = exporter.json_payload(
        run,
        score_breakdown=score,
        translation_groups=translation_groups,
        findings=findings,
    )
    groups_payload = payload["translationGroups"]
    findings_payload = payload["findings"]
    assert isinstance(groups_payload, list)
    assert isinstance(findings_payload, list)
    assert groups_payload[0]["sourceReference"] == ("[redacted source reference]")
    assert groups_payload[0]["members"][0]["sourceReference"] == (
        "[redacted source reference]"
    )
    assert (
        groups_payload[0]["members"][0]["position"]["source_reference"]
        == "[redacted source reference]"
    )
    assert findings_payload[0]["sourceReference"] == ("[redacted source reference]")
    assert findings_payload[1]["sourceReference"] == safe_reference
    assert findings_payload[2]["sourceReference"] == ("[redacted source reference]")
    serialized_payload = json.dumps(payload, ensure_ascii=False)
    for secret in (
        windows_group_path,
        unix_member_path,
        prefixed_windows_finding_path,
        relative_storage_path,
        "/srv/private/source.bin",
    ):
        assert secret not in serialized_payload
    assert "storagePath" not in serialized_payload

    sheets = exporter.workbook_data(
        run,
        score_breakdown=score,
        translation_groups=translation_groups,
        findings=findings,
    )
    workbook_path = tmp_path / "compliance-export.xlsx"
    _write_workbook(workbook_path, sheets)
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
    )
    try:
        assert workbook.sheetnames == [
            "Summary",
            "Score Breakdown",
            "Languages",
            "Sections",
            "Translation Groups",
            "Findings",
        ]
        summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
        assert summary_rows[0] == (
            "Document Code",
            "Revision",
            "Validation Rule",
            "Compliance Status",
            "Compliance Score",
            "Validated At",
            "Required Languages",
            "Missing Languages",
            "Required Sections",
            "Missing Sections",
            "Total Findings",
        )
        assert summary_rows[1][:6] == (
            "MTI-HRM-POL-001",
            "Rev.003",
            "TRILINGUAL-STRICT",
            "PARTIALLY_COMPLIANT",
            82.5,
            "2026-07-26T10:15:00+08:00",
        )

        score_rows = list(workbook["Score Breakdown"].iter_rows(values_only=True))
        validator_index = score_rows[0].index("validator")
        assert [row[validator_index] for row in score_rows[1:8]] == [
            "documentCode",
            "languagePresence",
            "languageCoverage",
            "sectionCompleteness",
            "languageOrder",
            "translationGroups",
            "tableCompleteness",
        ]
        assert score_rows[8][validator_index] == "FINAL"

        workbook_values = "\n".join(
            str(value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
    finally:
        workbook.close()

    for secret in (
        windows_group_path,
        unix_member_path,
        prefixed_windows_finding_path,
        relative_storage_path,
        "/srv/private/source.bin",
    ):
        assert secret not in workbook_values
    assert workbook_values.count("[redacted source reference]") >= 3


def test_context_accepts_phase7_read_rows_and_persists_group_provenance() -> None:
    source_id = uuid4()
    result_id = uuid4()
    container_id = uuid4()
    persisted = SimpleNamespace(
        id=result_id,
        extracted_block_id=source_id,
        ocr_block_id=None,
        container_id=container_id,
        source_type="NATIVE_EXTRACTION",
        source_reference="XLSX:sheet=Register:cell=A2",
        language_code="id",
        confidence=0.96,
        character_count=25,
        eligibility_status="ELIGIBLE",
        metadata_json={
            "containerIndex": 1,
            "blockOrder": 2,
            "source": {"coordinate": "A2"},
        },
    )
    row = SimpleNamespace(
        result=persisted,
        text="Teks Indonesia yang cukup",
        source_confidence=None,
    )
    context = ComplianceContextService().build(
        rule=_rule(),
        source_format="XLSX",
        blocks=[row],
    )
    member = member_from_block(context.blocks[0])

    assert member.extracted_block_id == source_id
    assert member.language_block_result_id == result_id
    assert member.source_type == "NATIVE_EXTRACTION"

    table_context = ComplianceContextService().build(
        rule=_rule(),
        source_format="XLSX",
        tables=[
            ComplianceTableData(
                id=uuid4(),
                container_id=container_id,
                source_reference="XLSX:sheet=Register:table=1",
                table_index=1,
                row_count=2,
                column_count=1,
                cells=[
                    ComplianceTableCellData(
                        id=uuid4(),
                        row_index=1,
                        column_index=0,
                        coordinate="A2",
                        text="Teks Indonesia yang cukup",
                        normalised_text="Teks Indonesia yang cukup",
                    ),
                ],
            ),
        ],
        language_results=[row],
    )
    table_member = member_from_block(table_context.tables[0].cells[0])
    assert table_member.extracted_block_id == source_id
    assert table_member.language_block_result_id == result_id
    assert table_member.position["coordinate"] == "A2"


@pytest.mark.asyncio
async def test_perfect_pipeline_weighted_score_and_maximum_are_100() -> None:
    container_id = uuid4()
    blocks = [
        _block(
            "PURPOSE",
            "en",
            1,
            container_id=container_id,
            block_type="HEADING",
        ),
    ]
    order = 2
    for _ in range(2):
        for text, language in (
            ("Teks Indonesia yang cukup panjang.", "id"),
            ("English text with sufficient evidence.", "en"),
            ("中文文本具有足够的语言证据。", "zh"),
        ):
            blocks.append(
                _block(
                    text,
                    language,
                    order,
                    container_id=container_id,
                ),
            )
            order += 1
    alias = SectionAliasData(
        canonical_code="PURPOSE",
        language_code="en",
        alias_text="Purpose",
        match_type="EXACT",
    )
    rule = _rule(
        validate_sections=True,
        validate_translation_groups=True,
        required_sections=["PURPOSE"],
    )
    context = _context(blocks, rule=rule, aliases=[alias])
    pipeline = await CompliancePipeline().run(context)

    assert pipeline.score.weighted_score == 100
    assert pipeline.score.maximum_score == 100
    assert pipeline.score.final_score == 100
    assert pipeline.status.status == "COMPLIANT"
    assert not pipeline.findings


@pytest.mark.asyncio
async def test_pipeline_reports_bounded_operational_progress() -> None:
    stages: list[tuple[str, int]] = []

    async def progress(stage: str, percentage: int) -> None:
        stages.append((stage, percentage))

    await CompliancePipeline().run(
        _context(
            [
                _block("Teks Indonesia yang cukup.", "id", 1),
                _block("Sufficient English content.", "en", 2),
                _block("足够的中文内容。", "zh", 3),
            ]
        ),
        progress_callback=progress,
    )

    assert stages == [
        ("DETECTING_SECTIONS", 15),
        ("GROUPING_CONTENT", 30),
        ("VALIDATING_LANGUAGES", 45),
        ("VALIDATING_SECTIONS", 60),
        ("VALIDATING_ORDER", 70),
        ("VALIDATING_TABLES", 80),
        ("GENERATING_FINDINGS", 88),
        ("CALCULATING_SCORE", 93),
    ]


def _pipeline_stage_failure(
    stage: str,
    error: Exception,
) -> tuple[CompliancePipeline, object]:
    def fail_sync(*_: object, **__: object):
        raise error

    async def fail_async(*_: object, **__: object):
        raise error

    if stage == COMPLIANCE_SECTION_DETECTION_FAILED:
        return (
            CompliancePipeline(
                section_detector=SimpleNamespace(detect=fail_sync),
            ),
            _context(
                [],
                rule=_rule(
                    validate_sections=True,
                    validate_language_order=False,
                ),
            ),
        )
    if stage == COMPLIANCE_GROUPING_FAILED:
        return (
            CompliancePipeline(
                grouping_service=SimpleNamespace(group=fail_sync),
            ),
            _context(
                [],
                rule=_rule(
                    validate_language_order=True,
                    validate_translation_groups=True,
                ),
            ),
        )
    return (
        CompliancePipeline(
            validators=[
                SimpleNamespace(
                    code="LANGUAGE_PRESENCE",
                    validate=fail_async,
                ),
            ],
        ),
        _context(
            [],
            rule=_rule(
                validate_language_order=False,
                validate_translation_groups=False,
            ),
        ),
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("stage", "public_message"),
    [
        (
            COMPLIANCE_SECTION_DETECTION_FAILED,
            "Document section detection could not be completed.",
        ),
        (
            COMPLIANCE_GROUPING_FAILED,
            "Multilingual content grouping could not be completed.",
        ),
        (
            COMPLIANCE_VALIDATION_FAILED,
            "Compliance validation could not be completed.",
        ),
    ],
)
async def test_pipeline_exposes_stable_stage_error_contract(
    stage: str,
    public_message: str,
) -> None:
    private_error = RuntimeError(f"private implementation failure: {stage}")
    pipeline, context = _pipeline_stage_failure(stage, private_error)

    with pytest.raises(CompliancePipelineStageError) as exc_info:
        await pipeline.run(context)  # type: ignore[arg-type]

    assert exc_info.value.code == stage
    assert exc_info.value.public_message == public_message
    assert str(exc_info.value) == public_message
    assert "private implementation failure" not in str(exc_info.value)
    assert exc_info.value.__cause__ is private_error


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "stage",
    [
        COMPLIANCE_SECTION_DETECTION_FAILED,
        COMPLIANCE_GROUPING_FAILED,
        COMPLIANCE_VALIDATION_FAILED,
    ],
)
@pytest.mark.parametrize(
    "passthrough_error_type",
    [
        CompliancePipelineCancelled,
        SoftTimeLimitExceeded,
        ComplianceContextBuildError,
    ],
)
async def test_pipeline_does_not_wrap_passthrough_errors(
    stage: str,
    passthrough_error_type: type[Exception],
) -> None:
    passthrough_error = passthrough_error_type("synthetic passthrough")
    pipeline, context = _pipeline_stage_failure(stage, passthrough_error)

    with pytest.raises(passthrough_error_type) as exc_info:
        await pipeline.run(context)  # type: ignore[arg-type]

    assert exc_info.value is passthrough_error
