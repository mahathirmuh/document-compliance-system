"""Focused Phase 9 glossary model, matching, validation, and import tests."""

from __future__ import annotations

import json
from datetime import UTC, datetime, timedelta
from io import BytesIO
from uuid import UUID, uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.models.glossary_enums import (
    GlossaryExceptionScopeType,
    GlossaryExceptionType,
    GlossaryLanguageCode,
    GlossaryScopeType,
    GlossaryTermSeverity,
    GlossaryTermType,
    GlossaryVariantType,
)
from app.models.glossary_exception import GlossaryException
from app.models.glossary_profile import GlossaryProfile
from app.models.glossary_term import GlossaryTerm
from app.models.glossary_term_variant import GlossaryTermVariant
from app.models.glossary_translation import GlossaryTranslation
from app.repositories.glossary_profile_repository import (
    GlossaryProfileRepository,
)
from app.schemas.glossary import (
    GlossaryImportMode,
    GlossaryProfileCreate,
    GlossaryTermCreate,
    GlossaryTranslationCreate,
    GlossaryVariantCreate,
)
from app.services.auth.auth_service import RequestMetadata
from app.services.glossary.contracts import (
    GlossaryTextBlock,
    GlossaryValidationScope,
)
from app.services.glossary.glossary_export_service import (
    GlossaryExportService,
)
from app.services.glossary.glossary_import_service import (
    GlossaryImportService,
)
from app.services.glossary.glossary_matching_service import (
    GlossaryMatchingService,
)
from app.services.glossary.glossary_profile_service import (
    GlossaryProfileService,
)
from app.services.glossary.glossary_service import GlossaryService
from app.services.glossary.glossary_validation_service import (
    GlossaryValidationService,
)
from app.services.glossary.matching.regex_term_matcher import (
    RegexTermMatcher,
    UnsafeGlossaryRegexError,
)


def _term(
    *,
    term_code: str = "DOC_CONTROL",
    term_type: GlossaryTermType = GlossaryTermType.PREFERRED,
    case_sensitive: bool = False,
    whole_word: bool = True,
    allow_inflection: bool = False,
    is_regex: bool = False,
) -> GlossaryTerm:
    return GlossaryTerm(
        id=uuid4(),
        glossary_profile_id=uuid4(),
        term_code=term_code,
        concept_name=term_code.replace("_", " ").title(),
        term_type=term_type,
        severity=(
            GlossaryTermSeverity.MAJOR
            if term_type is GlossaryTermType.FORBIDDEN
            else GlossaryTermSeverity.MINOR
        ),
        is_case_sensitive=case_sensitive,
        match_whole_word=whole_word,
        allow_inflection=allow_inflection,
        is_regex=is_regex,
        is_active=True,
        created_by=None,
        updated_by=None,
    )


def _translation(
    term: GlossaryTerm,
    language: GlossaryLanguageCode,
    text: str,
    *,
    preferred: bool = True,
    forbidden: bool = False,
    required: bool = False,
) -> GlossaryTranslation:
    item = GlossaryTranslation(
        id=uuid4(),
        glossary_term_id=term.id,
        language_code=language,
        term_text=text,
        normalised_term=text.casefold(),
        is_preferred=preferred,
        is_forbidden=forbidden,
        is_required=required,
        priority=0,
        is_active=True,
    )
    item.variants = []
    return item


def _block(
    text: str,
    language: str,
    *,
    group_id: UUID | None = None,
    confidence: float = 1.0,
) -> GlossaryTextBlock:
    return GlossaryTextBlock(
        text=text,
        language_code=language,
        source_type="NATIVE_EXTRACTION",
        source_reference=f"TEST:{uuid4()}",
        extracted_block_id=uuid4(),
        translation_group_id=group_id,
        confidence=confidence,
    )


def test_preferred_whole_word_and_case_sensitive_matching() -> None:
    term = _term()
    term.translations = [
        _translation(term, GlossaryLanguageCode.ENGLISH, "Document Control")
    ]
    service = GlossaryMatchingService()
    matches, warnings = service.match(
        [_block("Document Controller; document control.", "en")],
        [term],
    )
    assert warnings == []
    assert [item.matched_text for item in matches] == ["document control"]

    strict = _term(term_code="STRICT", case_sensitive=True)
    strict.translations = [
        _translation(strict, GlossaryLanguageCode.ENGLISH, "Policy")
    ]
    matches, _ = service.match(
        [_block("policy Policy", "en")],
        [strict],
    )
    assert [item.matched_text for item in matches] == ["Policy"]


def test_mandarin_substring_and_allowed_variant() -> None:
    term = _term()
    translation = _translation(
        term,
        GlossaryLanguageCode.CHINESE,
        "文件控制",
    )
    variant = GlossaryTermVariant(
        id=uuid4(),
        glossary_translation_id=translation.id,
        variant_text="文档控制",
        normalised_variant="文档控制",
        variant_type=GlossaryVariantType.SYNONYM,
        is_allowed=True,
        is_active=True,
    )
    translation.variants = [variant]
    term.translations = [translation]
    matches, _ = GlossaryMatchingService().match(
        [_block("本程序适用于文件控制，也称文档控制。", "zh")],
        [term],
    )
    assert len(matches) == 2
    assert {item.matched_text for item in matches} == {
        "文件控制",
        "文档控制",
    }
    assert next(
        item for item in matches if item.matched_text == "文档控制"
    ).is_allowed_variant


def test_language_mismatch_is_a_review_signal() -> None:
    term = _term()
    term.translations = [
        _translation(
            term,
            GlossaryLanguageCode.ENGLISH,
            "Document Control",
        ),
        _translation(
            term,
            GlossaryLanguageCode.CHINESE,
            "\u6587\u4ef6\u63a7\u5236",
        ),
    ]
    result = GlossaryValidationService().validate(
        blocks=[
            _block(
                "English context containing \u6587\u4ef6\u63a7\u5236.",
                "en",
            )
        ],
        terms=[term],
    )
    mismatch = next(
        item
        for item in result.findings
        if item.finding_code == "GLOSSARY_TERM_LANGUAGE_MISMATCH"
    )
    assert mismatch.language_code == "en"
    assert mismatch.metrics["configuredLanguage"] == "zh"


def test_regex_safety_and_bounded_regex_match() -> None:
    matcher = RegexTermMatcher(maximum_length=100, timeout_ms=50)
    spans = matcher.find(
        "Reference ABC-123 applies.",
        r"ABC-\d{3}",
        case_sensitive=True,
    )
    assert [(item.start, item.end) for item in spans] == [(10, 17)]
    with pytest.raises(UnsafeGlossaryRegexError):
        matcher.validate(r"(a+)+$")
    with pytest.raises(UnsafeGlossaryRegexError):
        matcher.validate(r"(foo)\1")


def test_forbidden_required_translation_consistency_and_confidence() -> None:
    group_id = uuid4()
    preferred = _term()
    en = _translation(
        preferred,
        GlossaryLanguageCode.ENGLISH,
        "Document Control",
        required=True,
    )
    legacy = _translation(
        preferred,
        GlossaryLanguageCode.ENGLISH,
        "Document Management Control",
        preferred=False,
    )
    zh = _translation(
        preferred,
        GlossaryLanguageCode.CHINESE,
        "文件控制",
        required=True,
    )
    preferred.translations = [en, legacy, zh]

    forbidden = _term(
        term_code="OLD_COMPANY",
        term_type=GlossaryTermType.FORBIDDEN,
    )
    forbidden.translations = [
        _translation(
            forbidden,
            GlossaryLanguageCode.ENGLISH,
            "Old Company",
            forbidden=True,
        )
    ]
    result = GlossaryValidationService().validate(
        blocks=[
            _block(
                "Document Control and Old Company.",
                "en",
                group_id=group_id,
                confidence=0.5,
            ),
            _block(
                "Document Management Control applies elsewhere.",
                "en",
            ),
        ],
        terms=[preferred, forbidden],
    )
    codes = {item.finding_code for item in result.findings}
    assert "FORBIDDEN_GLOSSARY_TERM" in codes
    assert "MISSING_GLOSSARY_TRANSLATION" in codes
    assert "NON_PREFERRED_GLOSSARY_TERM" in codes
    assert "INCONSISTENT_GLOSSARY_TRANSLATION" in codes
    assert "GLOSSARY_MATCH_LOW_CONFIDENCE" in codes


def test_effective_and_expired_forbidden_exceptions() -> None:
    term = _term(
        term_code="FORBIDDEN",
        term_type=GlossaryTermType.FORBIDDEN,
    )
    term.translations = [
        _translation(
            term,
            GlossaryLanguageCode.ENGLISH,
            "Prohibited",
            forbidden=True,
        )
    ]
    now = datetime.now(UTC)
    active = GlossaryException(
        id=uuid4(),
        glossary_term_id=term.id,
        scope_type=GlossaryExceptionScopeType.GLOBAL,
        exception_type=GlossaryExceptionType.ALLOW_FORBIDDEN_TERM,
        reason="Approved temporary wording.",
        effective_from=now.date() - timedelta(days=1),
        effective_to=now.date() + timedelta(days=1),
        is_active=True,
        created_at=now,
        created_by=None,
    )
    result = GlossaryValidationService().validate(
        blocks=[_block("Prohibited wording.", "en")],
        terms=[term],
        exceptions=[active],
        scope=GlossaryValidationScope(),
    )
    assert not any(
        item.finding_code == "FORBIDDEN_GLOSSARY_TERM"
        for item in result.findings
    )
    assert result.exception_applied_count == 1

    active.effective_to = now.date() - timedelta(days=1)
    result = GlossaryValidationService().validate(
        blocks=[_block("Prohibited wording.", "en")],
        terms=[term],
        exceptions=[active],
    )
    codes = {item.finding_code for item in result.findings}
    assert "FORBIDDEN_GLOSSARY_TERM" in codes
    assert "GLOSSARY_EXCEPTION_EXPIRED" in codes


@pytest.mark.asyncio
async def test_profile_resolution_orders_specific_before_global(
    session_factory,
) -> None:
    department_id = uuid4()
    document_type_id = uuid4()
    async with session_factory() as session:
        global_profile = GlossaryProfile(
            code="GLOBAL",
            name="Global",
            scope_type=GlossaryScopeType.GLOBAL,
            is_default=True,
            is_active=True,
        )
        specific = GlossaryProfile(
            code="SPECIFIC",
            name="Specific",
            scope_type=GlossaryScopeType.DEPARTMENT_DOCUMENT_TYPE,
            department_id=department_id,
            document_type_id=document_type_id,
            is_default=True,
            is_active=True,
        )
        session.add_all([global_profile, specific])
        await session.commit()
        resolved = await GlossaryProfileRepository(
            session
        ).resolve_for_scope(
            department_id=department_id,
            document_type_id=document_type_id,
        )
        assert [item.code for item in resolved] == ["SPECIFIC", "GLOBAL"]


@pytest.mark.asyncio
async def test_glossary_import_preview_confirm_and_template(
    session_factory,
    create_user,
) -> None:
    user = await create_user(is_superuser=True)
    content, filename = GlossaryImportService.template()
    workbook = load_workbook(BytesIO(content))
    workbook["Profiles"].append(
        ["GLOBAL", "Global", None, "GLOBAL", None, None, True, True]
    )
    workbook["Terms"].append(
        [
            "GLOBAL",
            "PPE",
            "Personal Protective Equipment",
            None,
            "REQUIRED",
            "MINOR",
            False,
            True,
            False,
            False,
            True,
            None,
        ]
    )
    for language, text in (
        ("id", "Alat Pelindung Diri"),
        ("en", "Personal Protective Equipment"),
        ("zh", "个人防护装备"),
    ):
        workbook["Translations"].append(
            [
                "PPE",
                language,
                text,
                True,
                False,
                True,
                0,
                None,
                None,
                True,
            ]
        )
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    async with session_factory() as session:
        service = GlossaryImportService(
            session,
            user,
            RequestMetadata(ip_address="127.0.0.1", user_agent="pytest"),
        )
        preview = await service.preview(
            filename=filename,
            content=output.getvalue(),
        )
        assert preview.valid
        result = await service.confirm(
            mode=GlossaryImportMode.CREATE_ONLY,
            filename=filename,
            content=output.getvalue(),
        )
        assert result.created["profiles"] == 1
        assert result.created["terms"] == 1
        assert result.created["translations"] == 3
        assert await session.scalar(
            select(func.count(GlossaryTerm.id))
        ) == 1
        exported, json_name, media_type = await GlossaryExportService(
            session,
            user,
            RequestMetadata(ip_address="127.0.0.1", user_agent="pytest"),
        ).export(export_format="json")
        assert json_name == "glossary_export.json"
        assert media_type == "application/json"
        payload = json.loads(exported)
        assert payload["terms"][0]["termCode"] == "PPE"
        assert {
            item["languageCode"]
            for item in payload["terms"][0]["translations"]
        } == {"id", "en", "zh"}

        xlsx, xlsx_name, _ = await GlossaryExportService(
            session,
            user,
            RequestMetadata(ip_address="127.0.0.1", user_agent="pytest"),
        ).export(export_format="xlsx")
        assert xlsx_name == "glossary_export.xlsx"
        exported_workbook = load_workbook(BytesIO(xlsx), read_only=True)
        assert exported_workbook.sheetnames == [
            "Profiles",
            "Terms",
            "Translations",
            "Variants",
            "Exceptions",
        ]
        exported_workbook.close()


@pytest.mark.asyncio
async def test_import_preview_rejects_unsafe_regex(
    session_factory,
    create_user,
) -> None:
    user = await create_user(is_superuser=True)
    content, filename = GlossaryImportService.template()
    workbook = load_workbook(BytesIO(content))
    workbook["Profiles"].append(
        ["GLOBAL", "Global", None, "GLOBAL", None, None, True, True]
    )
    workbook["Terms"].append(
        [
            "GLOBAL",
            "UNSAFE",
            "Unsafe",
            None,
            "PREFERRED",
            "MINOR",
            False,
            False,
            False,
            True,
            True,
            None,
        ]
    )
    workbook["Translations"].append(
        [
            "UNSAFE",
            "en",
            "(a+)+$",
            True,
            False,
            False,
            0,
            None,
            None,
            True,
        ]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    async with session_factory() as session:
        preview = await GlossaryImportService(
            session,
            user,
            RequestMetadata(ip_address=None, user_agent="pytest"),
        ).preview(
            filename=filename,
            content=output.getvalue(),
        )
    assert not preview.valid
    assert preview.invalid_rows == 1
    assert any("quantifier" in item.message for item in preview.issues)


@pytest.mark.asyncio
async def test_glossary_crud_translations_variant_and_archive(
    session_factory,
    create_user,
) -> None:
    user = await create_user(is_superuser=True)
    metadata = RequestMetadata(ip_address="127.0.0.1", user_agent="pytest")
    async with session_factory() as session:
        profile = await GlossaryProfileService(
            session,
            user,
            metadata,
        ).create(
            GlossaryProfileCreate(
                code="OPERATIONS",
                name="Operations glossary",
                scope_type=GlossaryScopeType.GLOBAL,
                is_default=True,
            )
        )
        service = GlossaryService(session, user, metadata)
        term = await service.create(
            GlossaryTermCreate(
                glossary_profile_id=profile.id,
                term_code="LOCKOUT",
                concept_name="Lockout Tagout",
                term_type=GlossaryTermType.PREFERRED,
                severity=GlossaryTermSeverity.MAJOR,
            )
        )
        translations = {}
        for language, text in (
            (GlossaryLanguageCode.INDONESIAN, "Penguncian dan Penandaan"),
            (GlossaryLanguageCode.ENGLISH, "Lockout Tagout"),
            (GlossaryLanguageCode.CHINESE, "\u4e0a\u9501\u6302\u724c"),
        ):
            translations[language] = await service.add_translation(
                term.id,
                GlossaryTranslationCreate(
                    language_code=language,
                    term_text=text,
                    is_preferred=True,
                    is_required=True,
                ),
            )
        variant = await service.add_variant(
            translations[GlossaryLanguageCode.ENGLISH].id,
            GlossaryVariantCreate(
                variant_text="LOTO",
                variant_type=GlossaryVariantType.FORBIDDEN_VARIANT,
                is_allowed=True,
            ),
        )
        assert not variant.is_allowed
        assert len((await service.get(term.id)).translations) == 3
        assert not (await service.archive(term.id)).is_active
        assert (await service.restore(term.id)).is_active
