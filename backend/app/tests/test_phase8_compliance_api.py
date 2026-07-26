"""Phase 8 compliance API, worker, history, export, and cancellation flow."""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace
from uuid import UUID, uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.authorization import AuditAction, UserRole
from app.core.config import get_settings
from app.models.audit_log import AuditLog
from app.models.document_file import DocumentFile
from app.models.validation_finding import ValidationFinding
from app.models.validation_rule import ValidationRule
from app.services.auth.token_service import TokenService
from app.services.compliance.compliance_worker_service import (
    ComplianceWorkerService,
)
from app.services.language.language_persistence_service import (
    LanguagePersistenceService,
)
from app.services.language.language_runtime_config import (
    LanguageRuntimeConfig,
)
from app.tests.test_phase7_language_persistence import (
    _language_job,
    _pipeline,
    _source_graph,
)
from app.workers.compliance_tasks import process_compliance_job

TestSessionFactory = async_sessionmaker[AsyncSession]


def _headers(user: object, tokens: TokenService) -> dict[str, str]:
    return {
        "Authorization": (
            f"Bearer {tokens.create_access_token(user)}"  # type: ignore[arg-type]
        )
    }


@pytest.mark.asyncio
async def test_compliance_queue_worker_results_export_and_cancel(
    api_client: AsyncClient,
    create_user,
    session_factory: TestSessionFactory,
    token_service: TokenService,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (
        document,
        revision,
        document_file,
        extraction_job,
        extraction_run,
        container,
        block,
    ) = _source_graph()
    user = await create_user(
        name="Compliance Controller",
        email="compliance.controller@example.com",
        role=UserRole.DOCUMENT_CONTROLLER,
        department_id=document.department_id,
    )
    config = LanguageRuntimeConfig(database_batch_size=1)
    async with session_factory() as session:
        rule = ValidationRule(
            code="PHASE8-API",
            name="Phase 8 API Rule",
            document_type_id=document.document_type_id,
            is_default=True,
            is_active=True,
        )
        session.add_all(
            [
                document,
                revision,
                document_file,
                extraction_job,
                extraction_run,
                container,
                block,
                rule,
            ]
        )
        await session.flush()
        revision.validation_rule_id = rule.id
        validation_rule_id = rule.id
        document_file.latest_extraction_run_id = extraction_run.id
        language_job = _language_job(
            document,
            revision,
            document_file,
            extraction_run,
        )
        session.add(language_job)
        await session.flush()
        language_run = await LanguagePersistenceService(
            session,
            config,
        ).persist_result(
            job=language_job,
            result=_pipeline(
                extraction_run,
                container,
                block,
                config,
            ),
            started_at=datetime.now(UTC),
            completed_at=datetime.now(UTC),
        )
        await session.commit()

    monkeypatch.setattr(
        process_compliance_job,
        "apply_async",
        lambda **_: SimpleNamespace(id="phase8-test-task"),
    )
    headers = _headers(user, token_service)
    queued = await api_client.post(
        "/api/v1/compliance/jobs",
        headers=headers,
        json={"documentFileId": str(document_file.id)},
    )
    assert queued.status_code == 202
    queued_data = queued.json()["data"]
    assert queued_data["status"] == "QUEUED"
    assert queued_data["reusedExistingResult"] is False
    job_id = queued_data["jobId"]

    status = await ComplianceWorkerService(
        get_settings().model_copy(update={"compliance_db_batch_size": 1}),
        session_factory=session_factory,
    ).process_job(
        UUID(job_id),
        worker_reference="phase8-worker-test",
        attempt_number=1,
    )
    assert status.value in {"COMPLETED", "PARTIALLY_COMPLETED"}

    async with session_factory() as session:
        stored_rule = await session.get(ValidationRule, validation_rule_id)
        assert stored_rule is not None
        stored_rule.code = "PHASE8-API-RENAMED"
        stored_rule.name = "Renamed after compliance validation"
        await session.commit()

    job_response = await api_client.get(
        f"/api/v1/compliance/jobs/{job_id}",
        headers=headers,
    )
    assert job_response.status_code == 200
    job_data = job_response.json()["data"]
    assert job_data["progress"] == 100
    assert job_data["document"]["baseDocumentCode"] == (document.base_document_code)
    run_id = job_data["resultSummary"]["runId"]

    run_response = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}",
        headers=headers,
    )
    summary_response = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/summary",
        headers=headers,
    )
    score_response = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/score-breakdown",
        headers=headers,
    )
    findings_response = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/findings",
        headers=headers,
    )
    history_response = await api_client.get(
        (f"/api/v1/document-files/{document_file.id}/compliance-history"),
        headers=headers,
    )
    assert run_response.status_code == 200
    assert summary_response.status_code == 200
    assert score_response.status_code == 200
    assert findings_response.status_code == 200
    assert history_response.status_code == 200
    run_data = run_response.json()["data"]
    summary_data = summary_response.json()["data"]
    finding_page = findings_response.json()["data"]
    finding_items = finding_page["items"]
    assert finding_page["totalItems"] == len(finding_items)
    assert finding_page["page"] == 1
    assert finding_page["pageSize"] == 100
    assert finding_items
    finding_detail = await api_client.get(
        f"/api/v1/findings/{finding_items[0]['id']}",
        headers=headers,
    )
    assert finding_detail.status_code == 200
    assert finding_detail.json()["data"]["history"][0]["action"] == (
        AuditAction.CREATE_FINDING.value
    )
    async with session_factory() as session:
        creation_audits = list(
            await session.scalars(
                select(AuditLog).where(
                    AuditLog.action == AuditAction.CREATE_FINDING,
                    AuditLog.entity_type == "ValidationFinding",
                )
            )
        )
    assert {str(row.entity_id) for row in creation_audits} == {
        item["id"] for item in finding_items
    }
    assert all(row.user_id == user.id for row in creation_audits)
    assert all(
        row.new_values_json
        and row.new_values_json["complianceRunId"] == run_id
        and row.new_values_json["isSystemGenerated"] is True
        and "sourceReference" not in row.new_values_json
        and "description" not in row.new_values_json
        for row in creation_audits
    )
    assert run_data["sourceContentHash"] == (language_run.source_content_hash)
    assert run_data["validationRule"] == {
        "id": str(validation_rule_id),
        "code": "PHASE8-API",
        "name": "Phase 8 API Rule",
        "version": 1,
    }
    assert run_data["ruleSnapshot"]["ruleCode"] == "PHASE8-API"
    assert run_data["ruleSnapshot"]["ruleName"] == "Phase 8 API Rule"
    assert summary_data["requiredLanguages"] == [
        "id",
        "en",
        "zh",
    ]
    assert [
        metric["languageCode"] for metric in summary_data["languageMetrics"]
    ] == summary_data["requiredLanguages"]
    finding_counts_by_language = {
        language: sum(item["languageCode"] == language for item in finding_items)
        for language in summary_data["requiredLanguages"]
    }
    assert {
        metric["languageCode"]: metric["findingCount"]
        for metric in summary_data["languageMetrics"]
    } == finding_counts_by_language
    assert all(
        metric["blockCoverage"] >= 0 and metric["characterCoverage"] >= 0
        for metric in summary_data["languageMetrics"]
    )
    expected_detected_languages = [
        language
        for language in summary_data["requiredLanguages"]
        if summary_data["languagePresence"][language] == "PRESENT"
    ]
    assert run_data["detectedLanguages"] == expected_detected_languages
    assert run_data["missingLanguages"] == [
        language
        for language in summary_data["requiredLanguages"]
        if language not in expected_detected_languages
    ]
    assert job_data["resultSummary"]["totalFindings"] == (run_data["totalFindings"])
    assert (
        job_data["resultSummary"]["criticalFindings"] == (run_data["criticalFindings"])
    )
    assert job_data["resultSummary"]["majorFindings"] == (run_data["majorFindings"])
    assert job_data["resultSummary"]["minorFindings"] == (run_data["minorFindings"])
    assert score_response.json()["data"]["finalScore"] >= 0
    assert history_response.json()["data"]["totalItems"] == 1

    exported = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/export",
        headers=headers,
        params={"format": "json"},
    )
    assert exported.status_code == 200
    assert exported.json()["summary"]["validationRule"] == {
        "id": str(validation_rule_id),
        "code": "PHASE8-API",
        "name": "Phase 8 API Rule",
        "version": 1,
    }
    assert exported.json()["limitations"] == {
        "semanticSimilarityEvaluated": False,
        "translationMeaningValidated": False,
    }
    exported_xlsx = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/export",
        headers=headers,
        params={"format": "xlsx"},
    )
    assert exported_xlsx.status_code == 200
    assert exported_xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    workbook = load_workbook(
        BytesIO(exported_xlsx.content),
        read_only=True,
        data_only=True,
    )
    try:
        summary_rows = list(
            workbook["Summary"].iter_rows(values_only=True)
        )
        validation_rule_column = summary_rows[0].index(
            "Validation Rule"
        )
        assert (
            summary_rows[1][validation_rule_column]
            == "PHASE8-API"
        )
    finally:
        workbook.close()
    async with session_factory() as session:
        export_audits = list(
            await session.scalars(
                select(AuditLog)
                .where(
                    AuditLog.action == AuditAction.EXPORT_COMPLIANCE_RESULT,
                    AuditLog.entity_id == UUID(run_id),
                )
                .order_by(AuditLog.created_at, AuditLog.id)
            )
        )
    assert len(export_audits) == 2
    assert sorted(
        audit.new_values_json["format"]
        for audit in export_audits
        if audit.new_values_json is not None
    ) == ["json", "xlsx"]
    assert all(
        audit.new_values_json is not None
        and audit.new_values_json["structuralValidationOnly"] is True
        and audit.new_values_json["documentFileId"] == str(document_file.id)
        for audit in export_audits
    )

    other_viewer = await create_user(
        name="Other Department Viewer",
        email="other.compliance.viewer@example.com",
        role=UserRole.VIEWER,
        department_id=uuid4(),
    )
    other_headers = _headers(other_viewer, token_service)
    outside_scope = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}",
        headers=other_headers,
    )
    forbidden_export = await api_client.get(
        f"/api/v1/compliance/runs/{run_id}/export",
        headers=other_headers,
        params={"format": "json"},
    )
    forbidden_start = await api_client.post(
        "/api/v1/compliance/jobs",
        headers=other_headers,
        json={"documentFileId": str(document_file.id)},
    )
    assert outside_scope.status_code == 404
    assert forbidden_export.status_code == 403
    assert forbidden_start.status_code == 403

    successful_revalidation = await api_client.post(
        f"/api/v1/compliance/runs/{run_id}/revalidate",
        headers=headers,
        json={"reason": "Verify repeat linkage against persisted findings."},
    )
    assert successful_revalidation.status_code == 202
    successful_job_id = successful_revalidation.json()["data"]["jobId"]
    successful_status = await ComplianceWorkerService(
        get_settings(),
        session_factory=session_factory,
    ).process_job(
        UUID(successful_job_id),
        worker_reference="phase8-repeat-test",
        attempt_number=1,
    )
    assert successful_status.value in {"COMPLETED", "PARTIALLY_COMPLETED"}
    successful_job = await api_client.get(
        f"/api/v1/compliance/jobs/{successful_job_id}",
        headers=headers,
    )
    repeated_run_id = successful_job.json()["data"]["resultSummary"]["runId"]
    repeated_findings = await api_client.get(
        f"/api/v1/compliance/runs/{repeated_run_id}/findings",
        headers=headers,
    )
    assert repeated_findings.status_code == 200
    repeated_finding_items = repeated_findings.json()["data"]["items"]
    assert repeated_finding_items
    assert all(item["isRepeat"] for item in repeated_finding_items)
    async with session_factory() as session:
        initial_ids = set(
            await session.scalars(
                select(ValidationFinding.id).where(
                    ValidationFinding.compliance_run_id == UUID(run_id)
                )
            )
        )
        persisted_repeats = list(
            await session.scalars(
                select(ValidationFinding).where(
                    ValidationFinding.compliance_run_id == UUID(repeated_run_id)
                )
            )
        )
    assert persisted_repeats
    assert all(
        finding.previous_finding_id in initial_ids for finding in persisted_repeats
    )

    repeated_summary_before = await api_client.get(
        f"/api/v1/compliance/runs/{repeated_run_id}/summary",
        headers=headers,
    )
    initial_open_count = repeated_summary_before.json()["data"]["findings"]["open"]
    workflow_finding_id = repeated_finding_items[0]["id"]
    reviewed = await api_client.post(
        f"/api/v1/findings/{workflow_finding_id}/review",
        headers=headers,
        json={"comment": "Confirmed during the retained-run review."},
    )
    resolved = await api_client.post(
        f"/api/v1/findings/{workflow_finding_id}/resolve",
        headers=headers,
        json={"comment": "Corrected in the controlled source."},
    )
    assert reviewed.status_code == 200
    assert resolved.status_code == 200
    repeated_summary_resolved = await api_client.get(
        f"/api/v1/compliance/runs/{repeated_run_id}/summary",
        headers=headers,
    )
    assert (
        repeated_summary_resolved.json()["data"]["findings"]["open"]
        == initial_open_count - 1
    )
    reopened = await api_client.post(
        f"/api/v1/findings/{workflow_finding_id}/reopen",
        headers=headers,
        json={"reason": "The structural issue reappeared."},
    )
    assert reopened.status_code == 200
    repeated_summary_reopened = await api_client.get(
        f"/api/v1/compliance/runs/{repeated_run_id}/summary",
        headers=headers,
    )
    assert (
        repeated_summary_reopened.json()["data"]["findings"]["open"]
        == initial_open_count
    )

    revalidated = await api_client.post(
        f"/api/v1/compliance/runs/{repeated_run_id}/revalidate",
        headers=headers,
        json={"reason": "Verify cancellation between validation stages."},
    )
    assert revalidated.status_code == 202
    revalidation_job_id = revalidated.json()["data"]["jobId"]
    cancelling_worker = ComplianceWorkerService(
        get_settings(),
        session_factory=session_factory,
    )
    original_set_progress = cancelling_worker._set_progress
    late_cancellations: list[str] = []

    async def cancel_after_persisting_progress(*args, **kwargs):
        await original_set_progress(*args, **kwargs)
        if kwargs["status"].value == "PERSISTING":
            cancelled = await api_client.post(
                f"/api/v1/compliance/jobs/{revalidation_job_id}/cancel",
                headers=headers,
            )
            assert cancelled.status_code == 200
            assert cancelled.json()["data"]["status"] == "CANCEL_REQUESTED"
            late_cancellations.append(cancelled.json()["data"]["status"])

    monkeypatch.setattr(
        cancelling_worker,
        "_set_progress",
        cancel_after_persisting_progress,
    )
    cancelled_status = await cancelling_worker.process_job(
        UUID(revalidation_job_id),
        worker_reference="phase8-cancel-test",
        attempt_number=1,
    )
    assert cancelled_status.value == "CANCELLED"
    assert late_cancellations == ["CANCEL_REQUESTED"]
    unchanged_history = await api_client.get(
        (f"/api/v1/document-files/{document_file.id}/compliance-history"),
        headers=headers,
    )
    assert unchanged_history.json()["data"]["totalItems"] == 2

    stale_revalidation = await api_client.post(
        f"/api/v1/compliance/runs/{repeated_run_id}/revalidate",
        headers=headers,
        json={"reason": "Verify a late source change blocks persistence."},
    )
    assert stale_revalidation.status_code == 202
    stale_job_id = stale_revalidation.json()["data"]["jobId"]
    stale_worker = ComplianceWorkerService(
        get_settings(),
        session_factory=session_factory,
    )
    original_stale_progress = stale_worker._set_progress
    late_source_changes: list[bool] = []

    async def replace_source_after_persisting_progress(*args, **kwargs):
        await original_stale_progress(*args, **kwargs)
        if kwargs["status"].value == "PERSISTING":
            async with session_factory() as session:
                await session.execute(
                    update(DocumentFile)
                    .where(DocumentFile.id == document_file.id)
                    .values(is_current=False)
                )
                await session.commit()
            late_source_changes.append(True)

    monkeypatch.setattr(
        stale_worker,
        "_set_progress",
        replace_source_after_persisting_progress,
    )
    try:
        stale_status = await stale_worker.process_job(
            UUID(stale_job_id),
            worker_reference="phase8-stale-source-test",
            attempt_number=1,
        )
    finally:
        async with session_factory() as session:
            await session.execute(
                update(DocumentFile)
                .where(DocumentFile.id == document_file.id)
                .values(is_current=True)
            )
            await session.commit()
    assert stale_status.value == "FAILED"
    assert late_source_changes == [True]
    stale_job = await api_client.get(
        f"/api/v1/compliance/jobs/{stale_job_id}",
        headers=headers,
    )
    assert stale_job.status_code == 200
    assert stale_job.json()["data"]["errorCode"] == "COMPLIANCE_SOURCE_NOT_AVAILABLE"

    failed_revalidation = await api_client.post(
        f"/api/v1/compliance/runs/{repeated_run_id}/revalidate",
        headers=headers,
        json={"reason": "Verify durable handling of unexpected worker errors."},
    )
    assert failed_revalidation.status_code == 202
    failed_job_id = failed_revalidation.json()["data"]["jobId"]
    lease_worker = ComplianceWorkerService(
        get_settings(),
        session_factory=session_factory,
    )
    async with (
        lease_worker._execution_lease(UUID(failed_job_id)) as execution_lease,
        lease_worker._execution_lease(UUID(failed_job_id)) as duplicate_execution_lease,
    ):
        assert execution_lease is True
        assert duplicate_execution_lease is False
    async with lease_worker._execution_lease(
        UUID(failed_job_id)
    ) as recovered_execution_lease:
        assert recovered_execution_lease is True
    async with session_factory() as session:
        _, first_lease = await lease_worker._start_job(
            session,
            UUID(failed_job_id),
            worker_reference="phase8-failure-test",
            attempt_number=1,
        )
    async with session_factory() as session:
        _, duplicate_lease = await lease_worker._start_job(
            session,
            UUID(failed_job_id),
            worker_reference="phase8-failure-test",
            attempt_number=1,
        )
    async with session_factory() as session:
        _, foreign_lease = await lease_worker._start_job(
            session,
            UUID(failed_job_id),
            worker_reference="phase8-foreign-worker",
            attempt_number=2,
        )
    assert first_lease is True
    assert duplicate_lease is False
    assert foreign_lease is False

    async def unexpected_pipeline_failure(*_args, **_kwargs):
        raise RuntimeError("synthetic internal failure")

    failing_worker = ComplianceWorkerService(
        get_settings(),
        session_factory=session_factory,
    )
    monkeypatch.setattr(
        failing_worker.pipeline,
        "run",
        unexpected_pipeline_failure,
    )
    failed_status = await failing_worker.process_job(
        UUID(failed_job_id),
        worker_reference="phase8-failure-test",
        attempt_number=1,
    )
    assert failed_status.value == "FAILED"

    failed_job = await api_client.get(
        f"/api/v1/compliance/jobs/{failed_job_id}",
        headers=headers,
    )
    assert failed_job.status_code == 200
    assert failed_job.json()["data"]["status"] == "FAILED"
    assert failed_job.json()["data"]["errorCode"] == "COMPLIANCE_VALIDATION_FAILED"
    assert "synthetic internal failure" not in failed_job.text

    sorted_jobs = await api_client.get(
        "/api/v1/compliance/jobs",
        headers=headers,
        params={
            "documentFileId": str(document_file.id),
            "sortBy": "progress",
            "sortOrder": "asc",
            "pageSize": 100,
        },
    )
    assert sorted_jobs.status_code == 200
    progress_values = [item["progress"] for item in sorted_jobs.json()["data"]["items"]]
    assert progress_values == sorted(progress_values)
