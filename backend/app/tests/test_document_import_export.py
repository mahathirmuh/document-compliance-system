"""Phase 4 Document Register XLSX import/export integration tests."""

from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
from zoneinfo import ZoneInfo

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker
from starlette.datastructures import UploadFile

from app.api.v1.endpoints.document_import import _read_bounded_upload
from app.core.authorization import AuditAction, UserRole
from app.core.config import get_settings
from app.core.exceptions import ApplicationError
from app.models.audit_log import AuditLog
from app.models.department import Department
from app.models.document import Document
from app.models.document_revision import DocumentRevision
from app.models.document_type import DocumentType
from app.services.auth.token_service import TokenService
from app.services.documents.date_filter import created_at_utc_bounds
from app.services.documents.document_export_service import (
    DOCUMENT_EXPORT_HEADERS,
    _excel_datetime,
)
from app.services.documents.document_import_service import (
    DOCUMENT_IMPORT_HEADERS,
    XLSX_CONTENT_TYPE,
    chunk_import_audit_changes,
)
from app.services.documents.xlsx_safety import excel_safe
from app.tests.test_documents_api import _headers, _seed_master

TestSessionFactory = async_sessionmaker[AsyncSession]


def test_excel_datetime_converts_aware_utc_to_local_naive() -> None:
    value = datetime(2026, 7, 25, 1, 30, tzinfo=UTC)

    converted = _excel_datetime(
        value,
        ZoneInfo("Asia/Makassar"),
    )

    expected = datetime(
        2026,
        7,
        25,
        9,
        30,
        tzinfo=UTC,
    ).replace(tzinfo=None)
    assert converted == expected
    assert converted.tzinfo is None


def test_created_at_bounds_use_application_timezone() -> None:
    start, end = created_at_utc_bounds(
        date(2026, 7, 25),
        date(2026, 7, 25),
        "Asia/Makassar",
    )

    assert start == datetime(2026, 7, 24, 16, tzinfo=UTC)
    assert end == datetime(2026, 7, 25, 16, tzinfo=UTC)


def test_excel_safe_removes_xml_controls_and_preserves_whitespace() -> None:
    assert excel_safe(
        "\x01=Dirty\x0bValue\x1f\t\n\r"
    ) == "'=DirtyValue\t\n\r"


def test_import_audit_chunks_preserve_more_than_200_changes() -> None:
    changes = [
        {
            "documentId": str(index),
            "baseDocumentCode": f"DOC-{index:03d}",
        }
        for index in range(205)
    ]

    chunks = chunk_import_audit_changes(changes)

    assert [len(chunk) for chunk in chunks] == [100, 100, 5]
    assert [change for chunk in chunks for change in chunk] == changes
    with pytest.raises(ValueError, match="at least 1"):
        chunk_import_audit_changes(changes, chunk_size=0)


def _row(
    *,
    number: str = "010",
    revision: str = "0",
    title: str = "Imported Procedure",
    department: str = "HRM",
    section: str | None = "IER",
    document_type: str = "SOP",
) -> list[Any]:
    return [
        "MTI",
        department,
        section,
        document_type,
        number,
        title,
        "Imported description",
        revision,
        "",
        "",
        "2026-07-25",
        "2026-08-01",
        "2027-01-01",
        "2027-08-01",
        "HRM",
        "Document Owner",
        "https://example.sharepoint.com/document",
        "LEGACY-001",
        "Imported row",
    ]


def _workbook(
    rows: list[list[Any]],
    *,
    headers: tuple[str, ...] = DOCUMENT_IMPORT_HEADERS,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Document Register"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _files(
    content: bytes,
    filename: str = "document_register.xlsx",
) -> dict[str, Any]:
    return {"file": (filename, content, XLSX_CONTENT_TYPE)}


@pytest.mark.asyncio
async def test_template_reference_format_and_import_permission(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    async with session_factory() as session:
        department = await session.scalar(
            select(Department).where(Department.code == "HRM")
        )
        assert department is not None
        department.name = "=Human\x0b Resource"
        await session.commit()
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-template",
    )
    response = await api_client.get(
        "/api/v1/documents/import/template",
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Document Register", "Reference"]
    sheet = workbook["Document Register"]
    assert tuple(cell.value for cell in sheet[1]) == DOCUMENT_IMPORT_HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    assert sheet["A1"].comment is not None
    reference = workbook["Reference"]
    assert reference.max_row > 1
    department_row = next(
        row
        for row in reference.iter_rows(values_only=True)
        if row[0] == "Department" and row[2] == "HRM"
    )
    assert department_row[3] == "'=Human Resource"

    viewer = await _headers(
        create_user,
        token_service,
        role=UserRole.VIEWER,
        suffix="import-viewer",
    )
    denied = await api_client.get(
        "/api/v1/documents/import/template",
        headers=viewer,
    )
    assert denied.status_code == 403


@pytest.mark.asyncio
async def test_preview_valid_duplicate_invalid_header_formula_and_file(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-preview",
    )
    content = _workbook([_row(), _row()])
    preview = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(content),
    )
    assert preview.status_code == 200, preview.text
    data = preview.json()["data"]
    assert data["totalRows"] == 2
    assert data["validCreateRows"] == 1
    assert data["duplicateRows"] == 1
    assert data["rows"][0]["baseDocumentCode"] == "MTI-HRM-IER-SOP-010"

    bad_header = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(_workbook([_row()], headers=("wrong",))),
    )
    assert bad_header.status_code == 400
    assert "Invalid header" in bad_header.text

    formula_row = _row(number="011")
    formula_row[5] = "=HYPERLINK(\"https://evil.example\")"
    formula = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(_workbook([formula_row])),
    )
    assert formula.status_code == 200
    assert formula.json()["data"]["invalidRows"] == 1
    assert "formulas are not accepted" in formula.text

    long_owner_row = _row(number="012")
    long_owner_row[15] = "O" * 151
    long_owner = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(_workbook([long_owner_row])),
    )
    assert long_owner.status_code == 200
    assert long_owner.json()["data"]["invalidRows"] == 1
    assert "at most 150" in long_owner.text

    non_xlsx = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files={"file": ("register.csv", b"a,b", "text/csv")},
    )
    assert non_xlsx.status_code == 400

    archive = BytesIO()
    with ZipFile(archive, "w", ZIP_DEFLATED) as workbook_zip:
        workbook_zip.writestr(
            "xl/sharedStrings.xml",
            b"A" * (2 * 1024 * 1024),
        )
    zip_bomb = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(archive.getvalue()),
    )
    assert zip_bomb.status_code == 400
    assert "unsafe compression ratio" in zip_bomb.text

    limited_settings = get_settings().model_copy(
        update={"document_import_max_file_size_mb": 1}
    )
    oversized = UploadFile(
        file=BytesIO(b"X" * (1024 * 1024 + 1)),
        filename="oversized.xlsx",
    )
    with pytest.raises(ApplicationError, match="import failed"):
        await _read_bounded_upload(oversized, limited_settings)


@pytest.mark.asyncio
async def test_confirm_create_add_revision_upsert_and_revalidation(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-confirm",
    )
    first = _workbook([_row()])
    confirmed = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(first),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert confirmed.status_code == 200, confirmed.text
    result = confirmed.json()["data"]
    assert result == {
        "mode": "CREATE_AND_ADD_REVISION",
        "totalRows": 1,
        "documentsCreated": 1,
        "revisionsAdded": 1,
        "metadataUpdated": 0,
        "duplicatesSkipped": 0,
        "invalidSkipped": 0,
        "failed": 0,
    }

    duplicate = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(first),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert duplicate.status_code == 200
    assert duplicate.json()["data"]["duplicatesSkipped"] == 1

    upsert_actor = await create_user(
        email="documents-import-upsert@example.com",
        role=UserRole.SUPER_ADMIN,
        is_superuser=True,
    )
    upsert_headers = {
        "Authorization": (
            "Bearer "
            f"{token_service.create_access_token(upsert_actor)}"
        )
    }
    same_revision_upsert = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=upsert_headers,
        files=_files(
            _workbook([_row(title="Upserted with existing revision")])
        ),
        data={"mode": "UPSERT_METADATA"},
    )
    assert same_revision_upsert.status_code == 200
    assert same_revision_upsert.json()["data"]["revisionsAdded"] == 0
    assert same_revision_upsert.json()["data"]["metadataUpdated"] == 1
    assert same_revision_upsert.json()["data"]["duplicatesSkipped"] == 1
    async with session_factory() as session:
        imported_document = await session.scalar(
            select(Document).where(
                Document.base_document_code == "MTI-HRM-IER-SOP-010"
            )
        )
        assert imported_document is not None
        assert imported_document.updated_by == upsert_actor.id
        import_audits = list(
            (
                await session.scalars(
                    select(AuditLog).where(
                        AuditLog.action
                        == AuditAction.IMPORT_DOCUMENT_REGISTER
                    )
                )
            ).all()
        )
        upsert_summary = next(
            audit
            for audit in import_audits
            if (
                audit.entity_type == "document_register_import"
                and audit.new_values_json is not None
                and audit.new_values_json.get("metadataUpdated") == 1
            )
        )
        assert upsert_summary.new_values_json is not None
        assert upsert_summary.new_values_json["traceChangeCount"] == 1
        trace_id = upsert_summary.new_values_json["traceId"]
        trace_chunks = [
            audit
            for audit in import_audits
            if (
                audit.entity_type == "document_register_import_chunk"
                and audit.new_values_json is not None
                and audit.new_values_json.get("traceId") == trace_id
            )
        ]
        assert len(trace_chunks) == 1
        assert trace_chunks[0].new_values_json is not None
        changes = trace_chunks[0].new_values_json["changes"]
        assert len(changes) == 1
        change = changes[0]
        assert change["documentId"] == str(imported_document.id)
        assert change["baseDocumentCode"] == "MTI-HRM-IER-SOP-010"
        assert change["revisionId"]
        assert change["operation"] == "UPSERT_METADATA"
        assert change["oldValues"]["title"] == "Imported Procedure"
        assert (
            change["newValues"]["title"]
            == "Upserted with existing revision"
        )

    second = _workbook([_row(revision="1", title="Ignored title")])
    added = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(second),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert added.status_code == 200, added.text
    assert added.json()["data"]["revisionsAdded"] == 1
    assert added.json()["data"]["metadataUpdated"] == 0

    third = _workbook([_row(revision="2", title="Upserted title")])
    upserted = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(third),
        data={"mode": "UPSERT_METADATA"},
    )
    assert upserted.status_code == 200, upserted.text
    assert upserted.json()["data"]["revisionsAdded"] == 1
    assert upserted.json()["data"]["metadataUpdated"] == 1

    listed = await api_client.get(
        "/api/v1/documents",
        headers=headers,
        params={"search": "MTI-HRM-IER-SOP-010"},
    )
    assert listed.status_code == 200, listed.text
    item = listed.json()["data"]["items"][0]
    assert item["title"] == "Upserted title"
    assert item["currentRevision"]["revisionCode"] == "Rev.002"
    detail = await api_client.get(
        f"/api/v1/documents/{item['id']}",
        headers=headers,
    )
    assert len(detail.json()["data"]["revisions"]) == 3


@pytest.mark.asyncio
async def test_export_headers_filters_audit_and_permission(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="export",
    )
    exact_code = "MTI-HRM-IER-SOP-020"
    external_reference_match = _row(
        number="022",
        title="External reference false positive",
    )
    external_reference_match[17] = exact_code
    content = _workbook(
        [
            _row(number="020", title="+Formula-like title"),
            _row(
                number="021",
                title=f"Title mentions {exact_code}",
            ),
            external_reference_match,
        ]
    )
    confirm = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(content),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert confirm.status_code == 200, confirm.text
    async with session_factory() as session:
        target = await session.scalar(
            select(Document).where(
                Document.base_document_code == exact_code
            )
        )
        assert target is not None
        target.title = "+Formula\x01-like title"
        target_revision = await session.scalar(
            select(DocumentRevision).where(
                DocumentRevision.document_id == target.id,
                DocumentRevision.revision_code == "Rev.000",
            )
        )
        assert target_revision is not None
        target_revision.remarks = "Dirty\x0bRemarks\x1f"
        await session.commit()
    broad_export = await api_client.get(
        "/api/v1/documents/export",
        headers=headers,
        params={"search": exact_code},
    )
    assert broad_export.status_code == 200, broad_export.text
    broad_workbook = load_workbook(
        BytesIO(broad_export.content),
        data_only=False,
    )
    assert broad_workbook["Document Register"].max_row == 4

    exact_list = await api_client.get(
        "/api/v1/documents",
        headers=headers,
        params={"baseDocumentCode": exact_code},
    )
    assert exact_list.status_code == 200, exact_list.text
    assert exact_list.json()["data"]["totalItems"] == 1
    assert exact_list.json()["data"]["items"][0]["baseDocumentCode"] == (
        exact_code
    )

    exported = await api_client.get(
        "/api/v1/documents/export",
        headers=headers,
        params={"baseDocumentCode": exact_code},
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    assert "document_register_" in exported.headers["content-disposition"]
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    sheet = workbook["Document Register"]
    assert tuple(cell.value for cell in sheet[1]) == DOCUMENT_EXPORT_HEADERS
    assert sheet.max_row == 2
    assert sheet.cell(2, 10).value == "'+Formula-like title"
    assert sheet.cell(2, 23).value == "DirtyRemarks"
    assert sheet.cell(2, 21).hyperlink is not None
    assert workbook["_metadata"].sheet_state == "hidden"
    generated_at = workbook["_metadata"]["B1"].value
    assert isinstance(generated_at, str)
    assert generated_at.endswith("+08:00")

    async with session_factory() as session:
        actions = set((await session.scalars(select(AuditLog.action))).all())
    assert AuditAction.IMPORT_DOCUMENT_REGISTER in actions
    assert AuditAction.EXPORT_DOCUMENT_REGISTER in actions

    viewer = await _headers(
        create_user,
        token_service,
        role=UserRole.VIEWER,
        suffix="export-viewer",
    )
    denied = await api_client.get(
        "/api/v1/documents/export",
        headers=viewer,
    )
    assert denied.status_code == 403


@pytest.mark.asyncio
async def test_import_supports_hyphenated_document_type_code(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    async with session_factory() as session:
        session.add(
            DocumentType(
                code="WORK-INSTRUCTION",
                name="Work Instruction",
                requires_section=True,
            )
        )
        await session.commit()
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-hyphen-type",
    )
    row = _row(
        number="WI-001",
        document_type="WORK-INSTRUCTION",
    )
    preview = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(_workbook([row])),
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["data"]["rows"][0]["baseDocumentCode"] == (
        "MTI-HRM-IER-WORK-INSTRUCTION-WI-001"
    )
    confirmed = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(_workbook([row])),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert confirmed.status_code == 200, confirmed.text
    assert confirmed.json()["data"]["documentsCreated"] == 1


@pytest.mark.asyncio
async def test_import_cannot_modify_archived_document(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-archived",
    )
    original = _workbook([_row(number="030", title="Archived original")])
    created = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(original),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert created.status_code == 200, created.text
    listed = await api_client.get(
        "/api/v1/documents",
        headers=headers,
        params={"search": "MTI-HRM-IER-SOP-030"},
    )
    document_id = listed.json()["data"]["items"][0]["id"]
    archived = await api_client.post(
        f"/api/v1/documents/{document_id}/archive",
        headers=headers,
        json={"reason": "Retention hold"},
    )
    assert archived.status_code == 200, archived.text

    added_revision = _workbook(
        [_row(number="030", revision="1", title="Forbidden revision")]
    )
    preview = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(added_revision),
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["data"]["invalidRows"] == 1
    assert "archived documents are read-only" in preview.text
    confirmed_revision = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(added_revision),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert confirmed_revision.status_code == 200
    assert confirmed_revision.json()["data"]["invalidSkipped"] == 1
    assert confirmed_revision.json()["data"]["revisionsAdded"] == 0

    upsert = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(
            _workbook(
                [_row(number="030", title="Forbidden metadata update")]
            )
        ),
        data={"mode": "UPSERT_METADATA"},
    )
    assert upsert.status_code == 200
    assert upsert.json()["data"]["invalidSkipped"] == 1
    assert upsert.json()["data"]["metadataUpdated"] == 0

    detail = await api_client.get(
        f"/api/v1/documents/{document_id}",
        headers=headers,
    )
    assert detail.status_code == 200, detail.text
    assert detail.json()["data"]["title"] == "Archived original"
    assert len(detail.json()["data"]["revisions"]) == 1


@pytest.mark.asyncio
async def test_created_date_filters_use_local_day_boundaries(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="created-date-boundary",
    )
    numbers = ("050", "051", "052", "053")
    confirmed = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(
            _workbook(
                [
                    _row(number=number, title=f"Boundary {number}")
                    for number in numbers
                ]
            )
        ),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert confirmed.status_code == 200, confirmed.text

    timestamps = {
        "050": datetime(2026, 7, 24, 15, 59, 59, tzinfo=UTC),
        "051": datetime(2026, 7, 24, 16, 0, 0, tzinfo=UTC),
        "052": datetime(2026, 7, 25, 15, 59, 59, tzinfo=UTC),
        "053": datetime(2026, 7, 25, 16, 0, 0, tzinfo=UTC),
    }
    async with session_factory() as session:
        documents = list(
            (
                await session.scalars(
                    select(Document).where(
                        Document.document_number.in_(numbers)
                    )
                )
            ).all()
        )
        for document in documents:
            document.created_at = timestamps[document.document_number]
        await session.commit()

    params = {
        "createdFrom": "2026-07-25",
        "createdTo": "2026-07-25",
        "sortBy": "baseDocumentCode",
        "sortOrder": "asc",
    }
    listed = await api_client.get(
        "/api/v1/documents",
        headers=headers,
        params=params,
    )
    assert listed.status_code == 200, listed.text
    assert {
        item["documentNumber"]
        for item in listed.json()["data"]["items"]
    } == {"051", "052"}

    exported = await api_client.get(
        "/api/v1/documents/export",
        headers=headers,
        params=params,
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content))
    sheet = workbook["Document Register"]
    assert {
        sheet.cell(row=row, column=8).value
        for row in range(2, sheet.max_row + 1)
    } == {"051", "052"}


@pytest.mark.asyncio
async def test_import_owner_department_historical_active_rules(
    api_client: AsyncClient,
    create_user: Any,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    master = await _seed_master(session_factory)
    headers = await _headers(
        create_user,
        token_service,
        suffix="import-owner-active",
    )
    async with session_factory() as session:
        other_department = await session.get(
            Department,
            master["other_department"].id,
        )
        assert other_department is not None
        other_department.is_active = False
        await session.commit()

    inactive_owner_create = _row(
        number="060",
        title="Inactive new owner",
    )
    inactive_owner_create[14] = "ICT"
    rejected_create = await api_client.post(
        "/api/v1/documents/import/preview",
        headers=headers,
        files=_files(_workbook([inactive_owner_create])),
    )
    assert rejected_create.status_code == 200
    assert rejected_create.json()["data"]["invalidRows"] == 1
    assert "owner department is inactive" in rejected_create.text
    rejected_create_confirm = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(_workbook([inactive_owner_create])),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert rejected_create_confirm.status_code == 200
    assert rejected_create_confirm.json()["data"]["invalidSkipped"] == 1
    assert rejected_create_confirm.json()["data"]["documentsCreated"] == 0

    original = _workbook(
        [_row(number="061", title="Historical owner original")]
    )
    created = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(original),
        data={"mode": "CREATE_AND_ADD_REVISION"},
    )
    assert created.status_code == 200, created.text
    async with session_factory() as session:
        owner_department = await session.get(
            Department,
            master["department"].id,
        )
        assert owner_department is not None
        owner_department.is_active = False
        await session.commit()

    unchanged_owner = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(
            _workbook(
                [
                    _row(
                        number="061",
                        title="Historical owner retained",
                    )
                ]
            )
        ),
        data={"mode": "UPSERT_METADATA"},
    )
    assert unchanged_owner.status_code == 200, unchanged_owner.text
    assert unchanged_owner.json()["data"]["metadataUpdated"] == 1

    changed_owner = _row(
        number="061",
        title="Forbidden inactive owner",
    )
    changed_owner[14] = "ICT"
    rejected_change = await api_client.post(
        "/api/v1/documents/import/confirm",
        headers=headers,
        files=_files(_workbook([changed_owner])),
        data={"mode": "UPSERT_METADATA"},
    )
    assert rejected_change.status_code == 200
    assert rejected_change.json()["data"]["invalidSkipped"] == 1
    assert rejected_change.json()["data"]["metadataUpdated"] == 0
