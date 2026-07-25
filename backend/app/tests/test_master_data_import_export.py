"""XLSX template, preview, confirm, export, and permission tests."""

from collections.abc import Callable
from io import BytesIO
from typing import Any
from uuid import UUID

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.authorization import AuditAction, UserRole
from app.models.audit_log import AuditLog
from app.models.document_type import DocumentType
from app.services.auth.token_service import TokenService
from app.services.master_data.import_export_service import (
    TEMPLATE_HEADERS,
    XLSX_CONTENT_TYPE,
)
from app.schemas.master_data import ImportEntityType

TestSessionFactory = async_sessionmaker[AsyncSession]
UserFactory = Callable[..., Any]


async def _admin_headers(
    create_user: UserFactory,
    token_service: TokenService,
    *,
    email: str = "xlsx-admin@example.com",
) -> dict[str, str]:
    user = await create_user(
        email=email,
        role=UserRole.SUPER_ADMIN,
        is_superuser=True,
    )
    return {"Authorization": f"Bearer {token_service.create_access_token(user)}"}


def _workbook(
    entity_type: ImportEntityType,
    rows: list[list[Any]],
    *,
    headers: tuple[str, ...] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or TEMPLATE_HEADERS[entity_type])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _files(content: bytes, filename: str = "import.xlsx") -> dict[str, Any]:
    return {"file": (filename, content, XLSX_CONTENT_TYPE)}


@pytest.mark.asyncio
async def test_template_has_exact_headers_and_formatting(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    headers = await _admin_headers(create_user, token_service)
    response = await api_client.get(
        "/api/v1/master-data/import/template/departments",
        headers=headers,
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[1]) == TEMPLATE_HEADERS[
        ImportEntityType.DEPARTMENTS
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


@pytest.mark.asyncio
async def test_preview_valid_duplicate_invalid_header_and_non_xlsx(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    headers = await _admin_headers(create_user, token_service)
    content = _workbook(
        ImportEntityType.DEPARTMENTS,
        [
            ["hrm", "Human Resource", None, True],
            ["HRM", "Duplicate", None, True],
            ["BAD CODE", "Invalid", None, True],
        ],
    )
    preview = await api_client.post(
        "/api/v1/master-data/import/preview",
        headers=headers,
        data={"entityType": "departments"},
        files=_files(content),
    )
    assert preview.status_code == 200, preview.text
    data = preview.json()["data"]
    assert data["entityType"] == "departments"
    assert data["totalRows"] == 3
    assert data["validRows"] == 1
    assert data["duplicateRows"] == 1
    assert data["invalidRows"] == 1
    assert data["rows"][0]["data"]["code"] == "HRM"

    invalid_header = _workbook(
        ImportEntityType.DEPARTMENTS,
        [["OPS", "Operations"]],
        headers=("code", "name"),
    )
    bad = await api_client.post(
        "/api/v1/master-data/import/preview",
        headers=headers,
        data={"entityType": "departments"},
        files=_files(invalid_header),
    )
    assert bad.status_code == 400
    assert "Invalid header" in bad.json()["errors"][0]["message"]

    non_xlsx = await api_client.post(
        "/api/v1/master-data/import/preview",
        headers=headers,
        data={"entityType": "departments"},
        files=_files(b"not a workbook", "departments.csv"),
    )
    assert non_xlsx.status_code == 400


@pytest.mark.asyncio
async def test_confirm_create_only_and_upsert_exact_contract(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    headers = await _admin_headers(create_user, token_service)
    content = _workbook(
        ImportEntityType.DEPARTMENTS,
        [["OPS", "Operations", "Initial", True]],
    )
    created = await api_client.post(
        "/api/v1/master-data/import/confirm",
        headers=headers,
        data={"entityType": "departments", "mode": "CREATE_ONLY"},
        files=_files(content),
    )
    assert created.status_code == 200, created.text
    assert created.json()["data"] == {
        "entityType": "departments",
        "mode": "CREATE_ONLY",
        "totalRows": 1,
        "created": 1,
        "updated": 0,
        "skipped": 0,
        "failed": 0,
    }

    create_only_again = await api_client.post(
        "/api/v1/master-data/import/confirm",
        headers=headers,
        data={"entityType": "departments", "mode": "CREATE_ONLY"},
        files=_files(content),
    )
    assert create_only_again.status_code == 200
    assert create_only_again.json()["data"]["skipped"] == 1

    updated_content = _workbook(
        ImportEntityType.DEPARTMENTS,
        [["OPS", "Operational Excellence", "Updated", False]],
    )
    updated = await api_client.post(
        "/api/v1/master-data/import/confirm",
        headers=headers,
        data={"entityType": "departments", "mode": "UPSERT"},
        files=_files(updated_content),
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["data"]["updated"] == 1
    listed = await api_client.get(
        "/api/v1/master-data/departments",
        headers=headers,
        params={"search": "Operational Excellence"},
    )
    assert listed.json()["data"]["items"][0]["isActive"] is False


@pytest.mark.asyncio
async def test_upsert_requires_update_permission_dynamically(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await create_user(
        email="future-create-only@example.com",
        role=UserRole.SUPER_ADMIN,
        is_superuser=False,
    )
    headers = {
        "Authorization": f"Bearer {token_service.create_access_token(user)}"
    }
    monkeypatch.setattr(
        "app.api.v1.endpoints.master_data_transfer.has_permission",
        lambda *_args, **_kwargs: False,
    )
    response = await api_client.post(
        "/api/v1/master-data/import/confirm",
        headers=headers,
        data={"entityType": "departments", "mode": "UPSERT"},
        files=_files(
            _workbook(
                ImportEntityType.DEPARTMENTS,
                [["OPS", "Operations", None, True]],
            )
        ),
    )
    assert response.status_code == 403
    assert "master_data:update" in response.json()["errors"][0]["message"]


@pytest.mark.asyncio
async def test_validation_rule_upsert_clears_old_document_type_default_fk(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    headers = await _admin_headers(create_user, token_service)
    document_type = await api_client.post(
        "/api/v1/master-data/document-types",
        headers=headers,
        json={
            "code": "SOP",
            "name": "Standard Operating Procedure",
            "category": "PROCEDURE",
        },
    )
    document_type_id = document_type.json()["data"]["id"]
    rule = await api_client.post(
        "/api/v1/master-data/validation-rules",
        headers=headers,
        json={
            "code": "SOP-DEFAULT",
            "name": "SOP Default",
            "documentTypeId": document_type_id,
            "isDefault": True,
        },
    )
    assert rule.status_code == 201, rule.text

    content = _workbook(
        ImportEntityType.VALIDATION_RULES,
        [
            [
                "SOP-DEFAULT",
                "Now Global Non-default",
                None,
                True,
                True,
                True,
                95,
                95,
                95,
                True,
                "id,en,zh",
                False,
                "TITLE,PURPOSE",
                False,
                95,
                70,
                False,
                True,
            ]
        ],
    )
    response = await api_client.post(
        "/api/v1/master-data/import/confirm",
        headers=headers,
        data={"entityType": "validation-rules", "mode": "UPSERT"},
        files=_files(content),
    )
    assert response.status_code == 200, response.text
    assert response.json()["data"]["updated"] == 1
    async with session_factory() as session:
        stored = await session.get(DocumentType, UUID(document_type_id))
        assert stored is not None
        assert stored.default_validation_rule_id is None


@pytest.mark.asyncio
async def test_export_applies_filter_and_records_audit(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
    session_factory: TestSessionFactory,
) -> None:
    headers = await _admin_headers(create_user, token_service)
    for code, active in (("ACTIVE", True), ("INACTIVE", False)):
        response = await api_client.post(
            "/api/v1/master-data/departments",
            headers=headers,
            json={"code": code, "name": code.title(), "isActive": active},
        )
        assert response.status_code == 201, response.text

    exported = await api_client.get(
        "/api/v1/master-data/export/departments",
        headers=headers,
        params={"isActive": True},
    )
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[1]) == TEMPLATE_HEADERS[
        ImportEntityType.DEPARTMENTS
    ]
    assert sheet.max_row == 2
    assert sheet["A2"].value == "ACTIVE"
    assert "_metadata" in workbook.sheetnames

    async with session_factory() as session:
        actions = list(
            (
                await session.scalars(
                    select(AuditLog.action).where(
                        AuditLog.action == AuditAction.EXPORT_MASTER_DATA
                    )
                )
            ).all()
        )
    assert actions == [AuditAction.EXPORT_MASTER_DATA]


@pytest.mark.asyncio
async def test_export_permission_is_enforced(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    viewer = await create_user(role=UserRole.VIEWER)
    headers = {
        "Authorization": f"Bearer {token_service.create_access_token(viewer)}"
    }
    response = await api_client.get(
        "/api/v1/master-data/export/departments",
        headers=headers,
    )
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_confirm_and_export_support_every_master_data_entity(
    api_client: AsyncClient,
    create_user: UserFactory,
    token_service: TokenService,
) -> None:
    headers = await _admin_headers(
        create_user,
        token_service,
        email="all-entities@example.com",
    )
    workbooks = {
        ImportEntityType.DEPARTMENTS: _workbook(
            ImportEntityType.DEPARTMENTS,
            [["DPT", "Department", None, True]],
        ),
        ImportEntityType.SECTIONS: _workbook(
            ImportEntityType.SECTIONS,
            [["DPT", "SEC", "Section", None, True]],
        ),
        ImportEntityType.DOCUMENT_TYPES: _workbook(
            ImportEntityType.DOCUMENT_TYPES,
            [["SOP", "Standard Procedure", "PROCEDURE", None, True, True]],
        ),
        ImportEntityType.DOCUMENT_STATUSES: _workbook(
            ImportEntityType.DOCUMENT_STATUSES,
            [["DRAFT", "Draft", None, 10, True, False, False, True]],
        ),
        ImportEntityType.VALIDATION_RULES: _workbook(
            ImportEntityType.VALIDATION_RULES,
            [
                [
                    "SOP-3LANG",
                    "SOP Three Language",
                    "SOP",
                    True,
                    True,
                    True,
                    95,
                    95,
                    95,
                    True,
                    "id,en,zh",
                    True,
                    "TITLE,PURPOSE",
                    False,
                    95,
                    70,
                    True,
                    True,
                ]
            ],
        ),
    }
    # Dependency order is intentional: section references DPT, and rule
    # references SOP.
    for entity_type in (
        ImportEntityType.DEPARTMENTS,
        ImportEntityType.SECTIONS,
        ImportEntityType.DOCUMENT_TYPES,
        ImportEntityType.DOCUMENT_STATUSES,
        ImportEntityType.VALIDATION_RULES,
    ):
        response = await api_client.post(
            "/api/v1/master-data/import/confirm",
            headers=headers,
            data={
                "entityType": entity_type.value,
                "mode": "CREATE_ONLY",
            },
            files=_files(workbooks[entity_type]),
        )
        assert response.status_code == 200, (
            entity_type.value,
            response.text,
        )
        assert response.json()["data"]["created"] == 1

        exported = await api_client.get(
            f"/api/v1/master-data/export/{entity_type.value}",
            headers=headers,
        )
        assert exported.status_code == 200, (
            entity_type.value,
            exported.text,
        )
        workbook = load_workbook(BytesIO(exported.content), data_only=True)
        assert tuple(cell.value for cell in workbook.active[1]) == (
            TEMPLATE_HEADERS[entity_type]
        )
        assert workbook.active.max_row == 2
